import sys
import os
import json
from datetime import datetime, date, timedelta
from pathlib import Path
import shutil

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit,
    QComboBox, QPushButton, QMessageBox, QVBoxLayout, QFormLayout,
    QHBoxLayout, QDialog, QTabWidget, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QMenu,
    QGroupBox, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit,
    QFileDialog, QProgressBar, QTreeWidget, QTreeWidgetItem,
    QSplitter, QStatusBar, QToolBar, QToolButton, QGridLayout,
    QScrollArea, QFrame, QListWidget, QListWidgetItem,
    QRadioButton, QButtonGroup, QInputDialog, QDialogButtonBox,
    QStackedWidget, QCalendarWidget, QSizePolicy, QStyle
)
from PySide6.QtCore import (
    Qt, QTimer, QDateTime, QDate, QTime, QThread, Signal,
    QSettings, QPoint, QSize, QRect, QEvent, QPropertyAnimation,
    QEasingCurve, QDir
)
from PySide6.QtGui import (
    QPalette, QColor, QIcon, QAction, QFont, QFontMetrics,
    QPainter, QBrush, QPen, QLinearGradient, QPixmap, QImage,
    QKeySequence, QShortcut, QCursor, QMouseEvent, QKeyEvent,
    QDesktopServices
)

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ================= CESTY A NASTAVENÍ =================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

APP_DIR = os.path.join(os.environ["LOCALAPPDATA"], "HEM_InventoryManager")
os.makedirs(APP_DIR, exist_ok=True)

SETTINGS_FILE = os.path.join(APP_DIR, "settings.json")
WELLNESS_DATA_FILE = os.path.join(APP_DIR, "wellness_data.json")
MINIBAR_DATA_FILE = os.path.join(APP_DIR, "minibar_data.json")
LOBBY_DATA_FILE = os.path.join(APP_DIR, "lobby_data.json")
WELLNESS_ITEMS_FILE = os.path.join(APP_DIR, "wellness_items.json")
MINIBAR_ITEMS_FILE = os.path.join(APP_DIR, "minibar_items.json")
LOBBY_ITEMS_FILE = os.path.join(APP_DIR, "lobby_items.json")
ARCHIVE_DATA_FILE = os.path.join(APP_DIR, "archive_data.json")

# ================= DEFAULT NASTAVENÍ =================
DEFAULT_SETTINGS = {
    "company_name": "Wellness Hotel Beethoven****",
    "company_address": "Beethovenova 1146, 430 01 Chomutov",
    "num_rooms": 30,
    "auto_backup": True,
    "backup_dir": os.path.join(Path.home(), "Desktop", "HEM_Inventory_Backups"),
    "theme": "light",
    "report_language": "cs",
    "currency": "CZK"
}

# ================= DEFAULT POLOŽKY =================
DEFAULT_WELLNESS_ITEMS = [
    {"id": 1, "name": "Víno", "unit": "ks", "active": True, "category": "nápoje"},
    {"id": 2, "name": "Voda s mátou", "unit": "ks", "active": True, "category": "nápoje"},
    {"id": 3, "name": "Obložená mísa", "unit": "ks", "active": True, "category": "jídlo"},
    {"id": 4, "name": "Čaj bylinkový", "unit": "ks", "active": True, "category": "nápoje"},
    {"id": 5, "name": "Káva", "unit": "ks", "active": True, "category": "nápoje"}
]

DEFAULT_MINIBAR_ITEMS = [
    {"id": 1, "name": "Bonaqua", "unit": "ks", "active": True, "category": "voda", "price": 45},
    {"id": 2, "name": "Voda Evian", "unit": "ks", "active": True, "category": "voda", "price": 85},
    {"id": 3, "name": "Tiger", "unit": "ks", "active": True, "category": "pivo", "price": 65},
    {"id": 4, "name": "Pilsner Urquell", "unit": "ks", "active": True, "category": "pivo", "price": 70},
    {"id": 5, "name": "Coca cola", "unit": "ks", "active": True, "category": "limonáda", "price": 55},
    {"id": 6, "name": "Red Bull", "unit": "ks", "active": True, "category": "energetický", "price": 75},
    {"id": 7, "name": "Stella Artois", "unit": "ks", "active": True, "category": "pivo", "price": 75},
    {"id": 8, "name": "Víno bílé", "unit": "ks", "active": True, "category": "víno", "price": 120},
    {"id": 9, "name": "Víno červené", "unit": "ks", "active": True, "category": "víno", "price": 120},
    {"id": 10, "name": "Kešu ořechy", "unit": "ks", "active": True, "category": "snack", "price": 45},
    {"id": 11, "name": "Brambůrky", "unit": "ks", "active": True, "category": "snack", "price": 40},
    {"id": 12, "name": "Sušené ovoce", "unit": "ks", "active": True, "category": "snack", "price": 35}
]

DEFAULT_LOBBY_ITEMS = [
    {"id": 1, "name": "Káva", "unit": "ks", "active": True, "category": "nápoje", "has_price": True, "default_price": 55},
    {"id": 2, "name": "Čaj", "unit": "ks", "active": True, "category": "nápoje", "has_price": True, "default_price": 45},
    {"id": 3, "name": "Voda", "unit": "ks", "active": True, "category": "nápoje", "has_price": True, "default_price": 35},
    {"id": 4, "name": "Deštník", "unit": "ks", "active": True, "category": "služby", "has_price": False},
]

# ================= CUSTOM SPINBOX S WATERMARK =================
class WatermarkSpinBox(QSpinBox):
    """SpinBox s vodoznakem "počet" na pozadí při hodnotě 0"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimum(0)
        self.setMaximum(1000)
        self.setValue(0)
        
    def paintEvent(self, event):
        """Překreslí widget s vodoznakem"""
        super().paintEvent(event)
        
        if self.value() == 0:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.TextAntialiasing)
            
            # Nastavení průhledného textu
            painter.setPen(QColor(180, 180, 180, 150))  # Světle šedá s průhledností
            font = self.font()
            font.setItalic(True)
            painter.setFont(font)
            
            # Vycentrování textu
            text_rect = self.rect().adjusted(5, 0, -25, 0)  # Odsazení pro šipky
            painter.drawText(text_rect, Qt.AlignLeft | Qt.AlignVCenter, "počet")
            
            # Skryjeme text "0" které vykresluje QSpinBox
            painter.setPen(Qt.transparent)
            painter.drawText(text_rect, Qt.AlignLeft | Qt.AlignVCenter, "0")

# ================= POMOCNÉ FUNKCE =================
def load_json_file(filepath, default_data):
    """Načte JSON soubor, pokud neexistuje vytvoří s defaultními daty"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(default_data, f, indent=2, ensure_ascii=False)
            return default_data.copy()
    except Exception as e:
        print(f"Chyba při načítání {filepath}: {e}")
        return default_data.copy()

def save_json_file(filepath, data):
    """Uloží data do JSON souboru"""
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Chyba při ukládání {filepath}: {e}")
        return False

def load_settings():
    """Načte nastavení aplikace"""
    return load_json_file(SETTINGS_FILE, DEFAULT_SETTINGS)

def save_settings(data):
    """Uloží nastavení aplikace"""
    return save_json_file(SETTINGS_FILE, data)

def load_wellness_items():
    """Načte položky wellness"""
    return load_json_file(WELLNESS_ITEMS_FILE, DEFAULT_WELLNESS_ITEMS)

def save_wellness_items(data):
    """Uloží položky wellness"""
    return save_json_file(WELLNESS_ITEMS_FILE, data)

def load_minibar_items():
    """Načte položky minibarů"""
    return load_json_file(MINIBAR_ITEMS_FILE, DEFAULT_MINIBAR_ITEMS)

def save_minibar_items(data):
    """Uloží položky minibarů"""
    return save_json_file(MINIBAR_ITEMS_FILE, data)

def load_lobby_items():
    """Načte položky lobby"""
    return load_json_file(LOBBY_ITEMS_FILE, DEFAULT_LOBBY_ITEMS)

def save_lobby_items(data):
    """Uloží položky lobby"""
    return save_json_file(LOBBY_ITEMS_FILE, data)

def load_wellness_data():
    """Načte data wellness"""
    return load_json_file(WELLNESS_DATA_FILE, {})

def save_wellness_data(data):
    """Uloží data wellness"""
    return save_json_file(WELLNESS_DATA_FILE, data)

def load_minibar_data():
    """Načte data minibarů"""
    return load_json_file(MINIBAR_DATA_FILE, {})

def save_minibar_data(data):
    """Uloží data minibarů"""
    return save_json_file(MINIBAR_DATA_FILE, data)

def load_lobby_data():
    """Načte data lobby"""
    return load_json_file(LOBBY_DATA_FILE, {})

def save_lobby_data(data):
    """Uloží data lobby"""
    return save_json_file(LOBBY_DATA_FILE, data)

def load_archive_data():
    """Načte archivní data"""
    return load_json_file(ARCHIVE_DATA_FILE, [])

def save_archive_data(data):
    """Uloží archivní data"""
    return save_json_file(ARCHIVE_DATA_FILE, data)

# ================= TÉMA APLIKACE =================
def setup_theme(app, theme_name):
    """Nastaví téma aplikace"""
    if theme_name == "dark":
        app.setStyle("Fusion")
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.WindowText, Qt.white)
        dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
        dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
        dark_palette.setColor(QPalette.ToolTipText, Qt.white)
        dark_palette.setColor(QPalette.Text, Qt.white)
        dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ButtonText, Qt.white)
        dark_palette.setColor(QPalette.BrightText, Qt.red)
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, Qt.black)
        app.setPalette(dark_palette)
    elif theme_name == "light":
        app.setStyle("Fusion")
        light_palette = QPalette()
        light_palette.setColor(QPalette.Window, QColor(240, 240, 240))
        light_palette.setColor(QPalette.WindowText, Qt.black)
        light_palette.setColor(QPalette.Base, Qt.white)
        light_palette.setColor(QPalette.AlternateBase, QColor(240, 240, 240))
        light_palette.setColor(QPalette.ToolTipBase, Qt.white)
        light_palette.setColor(QPalette.ToolTipText, Qt.black)
        light_palette.setColor(QPalette.Text, Qt.black)
        light_palette.setColor(QPalette.Button, QColor(240, 240, 240))
        light_palette.setColor(QPalette.ButtonText, Qt.black)
        light_palette.setColor(QPalette.BrightText, Qt.red)
        light_palette.setColor(QPalette.Link, QColor(0, 100, 200))
        light_palette.setColor(QPalette.Highlight, QColor(0, 100, 200))
        light_palette.setColor(QPalette.HighlightedText, Qt.white)
        app.setPalette(light_palette)
    else:
        app.setStyle("")

# ================= DIALOG PRO SPRÁVU POLOŽEK =================
class ManageItemsDialog(QDialog):
    def __init__(self, item_type, parent=None):
        super().__init__(parent)
        self.item_type = item_type  # "wellness", "minibar" nebo "lobby"
        self.parent = parent
        self.setWindowTitle(f"Správa položek - {item_type.capitalize()}")
        self.setMinimumSize(600, 500)
        
        if item_type == "wellness":
            self.items = load_wellness_items()
            self.save_function = save_wellness_items
        elif item_type == "minibar":
            self.items = load_minibar_items()
            self.save_function = save_minibar_items
        else:  # lobby
            self.items = load_lobby_items()
            self.save_function = save_lobby_items
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Tabulka položek
        if self.item_type == "lobby":
            self.table = QTableWidget()
            self.table.setColumnCount(6)  # ID, Název, Jednotka, Kategorie, Cena?, Aktivní
            headers = ["ID", "Název", "Jednotka", "Kategorie", "Cena", "Aktivní"]
        elif self.item_type == "minibar":
            self.table = QTableWidget()
            self.table.setColumnCount(5)
            headers = ["ID", "Název", "Jednotka", "Kategorie", "Cena (Kč)", "Aktivní"]
        else:  # wellness
            self.table = QTableWidget()
            self.table.setColumnCount(5)
            headers = ["ID", "Název", "Jednotka", "Kategorie", "Aktivní"]
        
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        self.load_items_to_table()
        
        layout.addWidget(self.table)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        
        btn_add = QPushButton("➕ Přidat")
        btn_edit = QPushButton("✏️ Upravit")
        btn_delete = QPushButton("🗑 Smazat")
        btn_toggle = QPushButton("🔄 Aktivovat/Deaktivovat")
        
        btn_add.clicked.connect(self.add_item)
        btn_edit.clicked.connect(self.edit_item)
        btn_delete.clicked.connect(self.delete_item)
        btn_toggle.clicked.connect(self.toggle_active)
        
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_edit)
        button_layout.addWidget(btn_delete)
        button_layout.addWidget(btn_toggle)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Informace
        info_label = QLabel("💡 Tip: Deaktivované položky se nebudou zobrazovat při zadávání.")
        info_label.setStyleSheet("color: gray; font-style: italic; padding: 5px;")
        layout.addWidget(info_label)
        
        self.setLayout(layout)
    
    def load_items_to_table(self):
        self.table.setRowCount(len(self.items))
        
        for row, item in enumerate(self.items):
            self.table.setItem(row, 0, QTableWidgetItem(str(item.get("id", row+1))))
            self.table.setItem(row, 1, QTableWidgetItem(item.get("name", "")))
            self.table.setItem(row, 2, QTableWidgetItem(item.get("unit", "ks")))
            self.table.setItem(row, 3, QTableWidgetItem(item.get("category", "")))
            
            col_offset = 0
            if self.item_type == "lobby":
                # Sloupec Cena: pokud has_price True, zobrazit cenu, jinak "ne"
                if item.get("has_price", False):
                    price_text = str(item.get("default_price", 0))
                else:
                    price_text = "ne"
                self.table.setItem(row, 4, QTableWidgetItem(price_text))
                active_col = 5
            elif self.item_type == "minibar":
                self.table.setItem(row, 4, QTableWidgetItem(str(item.get("price", 0))))
                active_col = 5
            else:  # wellness
                active_col = 4
            
            # Checkbox pro aktivní
            active_checkbox = QCheckBox()
            active_checkbox.setChecked(item.get("active", True))
            self.table.setCellWidget(row, active_col, active_checkbox)
            
            # Označit neaktivní položky šedě
            if not item.get("active", True):
                for col in range(self.table.columnCount()):
                    if col != active_col:  # Nepřebarvit checkbox
                        item_widget = self.table.item(row, col)
                        if item_widget:
                            item_widget.setForeground(QColor("gray"))
    
    def add_item(self):
        dialog = ItemEditDialog(self.item_type, parent=self)
        if dialog.exec():
            new_item = dialog.get_item()
            # Najít max ID
            max_id = max([item.get("id", 0) for item in self.items], default=0)
            new_item["id"] = max_id + 1
            self.items.append(new_item)
            self.save_function(self.items)
            self.load_items_to_table()
            if self.parent:
                self.parent.refresh_item_lists()
    
    def edit_item(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0 and selected_row < len(self.items):
            item = self.items[selected_row]
            dialog = ItemEditDialog(self.item_type, item, parent=self)
            if dialog.exec():
                updated_item = dialog.get_item()
                self.items[selected_row].update(updated_item)
                self.save_function(self.items)
                self.load_items_to_table()
                if self.parent:
                    self.parent.refresh_item_lists()
        else:
            QMessageBox.warning(self, "Upozornění", "Vyberte položku k úpravě.")
    
    def delete_item(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0 and selected_row < len(self.items):
            item_name = self.items[selected_row].get("name", "")
            reply = QMessageBox.question(
                self, "Smazat položku",
                f"Opravdu chcete smazat položku '{item_name}'?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                del self.items[selected_row]
                self.save_function(self.items)
                self.load_items_to_table()
                if self.parent:
                    self.parent.refresh_item_lists()
        else:
            QMessageBox.warning(self, "Upozornění", "Vyberte položku ke smazání.")
    
    def toggle_active(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0 and selected_row < len(self.items):
            current_state = self.items[selected_row].get("active", True)
            self.items[selected_row]["active"] = not current_state
            self.save_function(self.items)
            self.load_items_to_table()
            if self.parent:
                self.parent.refresh_item_lists()
        else:
            QMessageBox.warning(self, "Upozornění", "Vyberte položku.")

class ItemEditDialog(QDialog):
    def __init__(self, item_type, item_data=None, parent=None):
        super().__init__(parent)
        self.item_type = item_type
        self.item_data = item_data or {}
        self.is_edit = bool(item_data)
        
        title = "Upravit položku" if self.is_edit else "Nová položka"
        self.setWindowTitle(f"{title} - {item_type.capitalize()}")
        self.setFixedWidth(400)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QFormLayout()
        
        # Název
        self.name_edit = QLineEdit(self.item_data.get("name", ""))
        layout.addRow("Název položky *:", self.name_edit)
        
        # Jednotka
        self.unit_combo = QComboBox()
        self.unit_combo.addItems(["ks", "l", "kg", "g", "ml"])
        self.unit_combo.setCurrentText(self.item_data.get("unit", "ks"))
        layout.addRow("Jednotka:", self.unit_combo)
        
        # Kategorie
        self.category_edit = QLineEdit(self.item_data.get("category", ""))
        layout.addRow("Kategorie:", self.category_edit)
        
        # Pro lobby: checkbox "Sledovat cenu" a cena
        if self.item_type == "lobby":
            self.has_price_check = QCheckBox("Sledovat cenu")
            self.has_price_check.setChecked(self.item_data.get("has_price", False))
            layout.addRow("", self.has_price_check)
            
            self.price_spin = QDoubleSpinBox()
            self.price_spin.setRange(0, 10000)
            self.price_spin.setValue(self.item_data.get("default_price", 0))
            self.price_spin.setSuffix(" Kč")
            self.price_spin.setEnabled(self.has_price_check.isChecked())
            self.has_price_check.toggled.connect(self.price_spin.setEnabled)
            layout.addRow("Cena (Kč):", self.price_spin)
        
        # Pro minibar: cena
        elif self.item_type == "minibar":
            self.price_spin = QDoubleSpinBox()
            self.price_spin.setRange(0, 10000)
            self.price_spin.setValue(self.item_data.get("price", 0))
            self.price_spin.setSuffix(" Kč")
            layout.addRow("Cena:", self.price_spin)
        
        # Aktivní
        self.active_checkbox = QCheckBox("Aktivní")
        self.active_checkbox.setChecked(self.item_data.get("active", True))
        layout.addRow("", self.active_checkbox)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_item)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)
        
        self.setLayout(layout)
    
    def save_item(self):
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "Chyba", "Zadejte název položky.")
            return
        
        self.item_data = {
            "name": self.name_edit.text().strip(),
            "unit": self.unit_combo.currentText(),
            "category": self.category_edit.text().strip(),
            "active": self.active_checkbox.isChecked()
        }
        
        if self.item_type == "minibar":
            self.item_data["price"] = self.price_spin.value()
        elif self.item_type == "lobby":
            self.item_data["has_price"] = self.has_price_check.isChecked()
            if self.has_price_check.isChecked():
                self.item_data["default_price"] = self.price_spin.value()
            else:
                self.item_data["default_price"] = 0
        
        self.accept()
    
    def get_item(self):
        return self.item_data

# ================= DIALOG PRO EDITACI ZÁZNAMU =================
class EditArchiveDialog(QDialog):
    def __init__(self, entry_data, wellness_items, minibar_items, lobby_items, parent=None):
        super().__init__(parent)
        self.entry_data = entry_data.copy()
        self.wellness_items = wellness_items
        self.minibar_items = minibar_items
        self.lobby_items = lobby_items
        self.data_type = entry_data.get("type", "wellness")
        
        self.setWindowTitle(f"Upravit záznam - {entry_data.get('date', '')}")
        self.setMinimumSize(500, 400)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Informace o záznamu
        info_label = QLabel(f"Úprava záznamu z {self.entry_data.get('date', '')} ({self.data_type})")
        info_label.setStyleSheet("font-weight: bold; padding: 5px;")
        layout.addWidget(info_label)
        
        # Formulář pro editaci
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        # Data pro editaci (položky)
        self.item_widgets = {}
        data = self.entry_data.get("data", {})
        
        if self.data_type == "wellness":
            items = self.wellness_items
        elif self.data_type == "minibar":
            items = self.minibar_items
        else:  # lobby
            items = self.lobby_items
        
        # Aktivní položky
        active_items = [item for item in items if item.get("active", True)]
        
        for item in active_items:
            item_id = str(item['id'])
            item_name = item['name']
            unit = item.get('unit', 'ks')
            
            item_layout = QHBoxLayout()
            item_label = QLabel(f"{item_name} ({unit}):")
            
            # Použít WatermarkSpinBox
            item_spin = WatermarkSpinBox()
            item_spin.setValue(data.get(item_id, 0) if isinstance(data.get(item_id), int) else 0)
            
            item_layout.addWidget(item_label)
            item_layout.addWidget(item_spin)
            
            # Pro lobby položky s cenou
            if self.data_type == "lobby" and item.get("has_price", False):
                price_label = QLabel("Cena:")
                price_spin = QDoubleSpinBox()
                price_spin.setRange(0, 100000)
                price_spin.setValue(data.get(item_id, {}).get("price", item.get("default_price", 0)) if isinstance(data.get(item_id), dict) else item.get("default_price", 0))
                price_spin.setSuffix(" Kč")
                item_layout.addWidget(price_label)
                item_layout.addWidget(price_spin)
                self.item_widgets[item_id] = (item_spin, price_spin)
            else:
                item_layout.addStretch()
                self.item_widgets[item_id] = item_spin
            
            scroll_layout.addLayout(item_layout)
        
        # Pokud lobby, přidat sekci pro "Na přání"
        if self.data_type == "lobby":
            scroll_layout.addWidget(QLabel("Na přání:"))
            self.custom_table = QTableWidget()
            self.custom_table.setColumnCount(4)
            self.custom_table.setHorizontalHeaderLabels(["Popis", "Množství", "Cena", ""])
            self.custom_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            self.custom_table.setMaximumHeight(150)
            
            # Naplnit existující data
            custom_requests = data.get("custom", []) if isinstance(data, dict) else []
            self.custom_table.setRowCount(len(custom_requests))
            for i, req in enumerate(custom_requests):
                self.custom_table.setItem(i, 0, QTableWidgetItem(req.get("description", "")))
                self.custom_table.setItem(i, 1, QTableWidgetItem(str(req.get("qty", 0))))
                self.custom_table.setItem(i, 2, QTableWidgetItem(str(req.get("price", 0))))
                
                btn_del = QPushButton("🗑")
                btn_del.clicked.connect(lambda checked, r=i: self.custom_table.removeRow(r))
                self.custom_table.setCellWidget(i, 3, btn_del)
            
            scroll_layout.addWidget(self.custom_table)
            
            btn_add_custom = QPushButton("➕ Přidat řádek")
            btn_add_custom.clicked.connect(self.add_custom_row)
            scroll_layout.addWidget(btn_add_custom)
        
        # Poznámka
        scroll_layout.addWidget(QLabel("Poznámka:"))
        self.note_edit = QLineEdit(data.get("note", ""))
        scroll_layout.addWidget(self.note_edit)
        
        scroll_widget.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(300)
        
        layout.addWidget(scroll_area)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit změny")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_changes)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def add_custom_row(self):
        row = self.custom_table.rowCount()
        self.custom_table.insertRow(row)
        self.custom_table.setItem(row, 0, QTableWidgetItem(""))
        self.custom_table.setItem(row, 1, QTableWidgetItem("0"))
        self.custom_table.setItem(row, 2, QTableWidgetItem("0"))
        btn_del = QPushButton("🗑")
        btn_del.clicked.connect(lambda checked, r=row: self.custom_table.removeRow(r))
        self.custom_table.setCellWidget(row, 3, btn_del)
    
    def save_changes(self):
        """Uloží provedené změny"""
        # Shromáždit data
        updated_data = {}
        
        for item_id, widget in self.item_widgets.items():
            if isinstance(widget, tuple):
                qty_spin, price_spin = widget
                qty = qty_spin.value()
                if qty > 0:
                    updated_data[item_id] = {"qty": qty, "price": price_spin.value()}
            else:
                qty = widget.value()
                if qty > 0:
                    updated_data[item_id] = qty
        
        # Přidat custom položky
        custom_list = []
        for row in range(self.custom_table.rowCount()):
            desc_item = self.custom_table.item(row, 0)
            qty_item = self.custom_table.item(row, 1)
            price_item = self.custom_table.item(row, 2)
            if desc_item and qty_item and price_item:
                desc = desc_item.text().strip()
                qty = qty_item.text().strip()
                price = price_item.text().strip()
                if desc and qty.isdigit() and price.replace('.','',1).isdigit():
                    custom_list.append({
                        "description": desc,
                        "qty": int(qty),
                        "price": float(price)
                    })
        if custom_list:
            updated_data["custom"] = custom_list
        
        # Přidat poznámku
        note = self.note_edit.text().strip()
        if note:
            updated_data["note"] = note
        
        # Zachovat původní metadata
        updated_data["timestamp"] = self.entry_data.get("data", {}).get("timestamp", datetime.now().isoformat())
        updated_data["user"] = self.entry_data.get("data", {}).get("user", os.getlogin())
        
        # Aktualizovat vstupní data
        self.entry_data["data"] = updated_data
        
        # Vytvořit popis
        items_list = []
        for key, value in updated_data.items():
            if key not in ["note", "timestamp", "user", "custom"]:
                if self.data_type == "wellness":
                    item_name = next((item['name'] for item in self.wellness_items 
                                    if str(item['id']) == key), f"Položka {key}")
                    items_list.append(f"{item_name}: {value}")
                elif self.data_type == "minibar":
                    item_name = next((item['name'] for item in self.minibar_items 
                                    if str(item['id']) == key), f"Položka {key}")
                    items_list.append(f"{item_name}: {value}")
                else:  # lobby
                    item = next((item for item in self.lobby_items if str(item['id']) == key), None)
                    if item:
                        if isinstance(value, dict):
                            items_list.append(f"{item['name']}: {value['qty']} ks (cena {value['price']} Kč)")
                        else:
                            items_list.append(f"{item['name']}: {value} ks")
        
        if "custom" in updated_data:
            for c in updated_data["custom"]:
                items_list.append(f"Na přání - {c['description']}: {c['qty']} ks (cena {c['price']} Kč)")
        
        if items_list:
            description = f"{self.data_type.capitalize()} - " + ", ".join(items_list[:3])
            if len(items_list) > 3:
                description += f" + {len(items_list) - 3} další"
        else:
            description = f"{self.data_type.capitalize()} - žádné položky"
        
        self.entry_data["description"] = description
        self.entry_data["timestamp"] = datetime.now().isoformat()
        
        self.accept()
    
    def get_updated_entry(self):
        """Vrátí aktualizovaný záznam"""
        return self.entry_data

# ================= HLAVNÍ OKNO APLIKACE =================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = load_settings()
        self.setWindowTitle("HEM - Inventory & Write-off Manager v0.2.0")
        self.resize(1200, 800)
        
        self.wellness_items = load_wellness_items()
        self.minibar_items = load_minibar_items()
        self.lobby_items = load_lobby_items()
        self.wellness_data = load_wellness_data()
        self.minibar_data = load_minibar_data()
        self.lobby_data = load_lobby_data()
        
        self.current_date = QDate.currentDate()
        
        self.init_ui()
    
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        
        # Horní panel s datem a tlačítky
        top_layout = QHBoxLayout()
        
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Datum:"))
        
        self.date_edit = QDateEdit()
        self.date_edit.setDate(self.current_date)
        self.date_edit.setCalendarPopup(True)
        self.date_edit.dateChanged.connect(self.date_changed)
        date_layout.addWidget(self.date_edit)
        
        btn_today = QPushButton("Dnes")
        btn_today.clicked.connect(self.set_today)
        date_layout.addWidget(btn_today)
        
        top_layout.addLayout(date_layout)
        top_layout.addStretch()
        
        main_layout.addLayout(top_layout)
        
        # Hlavní záložky
        self.tabs = QTabWidget()
        
        # Wellness záložka
        self.wellness_tab = self.create_wellness_tab()
        self.tabs.addTab(self.wellness_tab, "🏊 Wellness Odpisy")
        
        # Minibary záložka
        self.minibar_tab = self.create_minibar_writeoff_tab()
        self.tabs.addTab(self.minibar_tab, "🥤 Minibary Odpisy")
        
        # Lobby záložka (nová)
        self.lobby_tab = self.create_lobby_tab()
        self.tabs.addTab(self.lobby_tab, "🛋️ Lobby Odpisy")
        
        # Měsíční report záložka
        self.report_tab = self.create_report_tab()
        self.tabs.addTab(self.report_tab, "📊 Měsíční Report")
        
        # Zásobník záložka
        self.archive_tab = self.create_archive_tab()
        self.tabs.addTab(self.archive_tab, "📋 Zásobník")
        
        # Nastavení záložka
        self.settings_tab = self.create_settings_tab()
        self.tabs.addTab(self.settings_tab, "⚙ Nastavení")
        
        main_layout.addWidget(self.tabs)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.update_status_bar()
        
        central_widget.setLayout(main_layout)
        
        # Načíst data pro aktuální datum
        self.load_current_date_data()
    
    def create_wellness_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Denní evidence wellness odpisů")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Formulář pro zadání
        form_group = QGroupBox("Zadejte spotřebu")
        form_layout = QGridLayout()
        
        self.wellness_widgets = {}
        row = 0
        
        # Načíst aktivní wellness položky
        active_items = [item for item in self.wellness_items if item.get("active", True)]
        
        for item in active_items:
            label = QLabel(f"{item['name']} ({item['unit']}):")
            # Použít WatermarkSpinBox místo QSpinBox
            spinbox = WatermarkSpinBox()
            spinbox.setValue(0)
            
            form_layout.addWidget(label, row, 0)
            form_layout.addWidget(spinbox, row, 1)
            
            self.wellness_widgets[item['id']] = {
                'name': item['name'],
                'unit': item['unit'],
                'widget': spinbox
            }
            row += 1
        
        # Poznámka
        form_layout.addWidget(QLabel("Poznámka:"), row, 0)
        self.wellness_note = QLineEdit()
        form_layout.addWidget(self.wellness_note, row, 1)
        row += 1
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("💾 Uložit")
        btn_save.clicked.connect(self.save_wellness_data)
        
        btn_clear = QPushButton("🗑 Vymazat")
        btn_clear.clicked.connect(self.clear_wellness_form)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_clear)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        tab.setLayout(layout)
        return tab
    
    def create_minibar_writeoff_tab(self):
        """Vytvoří záložku pro minibar odpisy (stejná struktura jako wellness)"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Denní evidence minibar odpisů")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Formulář pro zadání
        form_group = QGroupBox("Zadejte spotřebu minibarů")
        form_layout = QGridLayout()
        
        self.minibar_widgets = {}
        row = 0
        
        # Načíst aktivní minibar položky
        active_items = [item for item in self.minibar_items if item.get("active", True)]
        
        for item in active_items:
            label = QLabel(f"{item['name']} ({item['unit']}):")
            # Použít WatermarkSpinBox místo QSpinBox
            spinbox = WatermarkSpinBox()
            spinbox.setValue(0)
            
            form_layout.addWidget(label, row, 0)
            form_layout.addWidget(spinbox, row, 1)
            
            self.minibar_widgets[item['id']] = {
                'name': item['name'],
                'unit': item['unit'],
                'price': item.get('price', 0),
                'widget': spinbox
            }
            row += 1
        
        # Poznámka
        form_layout.addWidget(QLabel("Poznámka:"), row, 0)
        self.minibar_note = QLineEdit()
        form_layout.addWidget(self.minibar_note, row, 1)
        row += 1
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("💾 Uložit")
        btn_save.clicked.connect(self.save_minibar_data)
        
        btn_clear = QPushButton("🗑 Vymazat")
        btn_clear.clicked.connect(self.clear_minibar_form)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_clear)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        tab.setLayout(layout)
        return tab
    
    def create_lobby_tab(self):
        """Vytvoří záložku pro lobby odpisy s možností zadání ceny a položek na přání"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Denní evidence lobby odpisů")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Formulář pro zadání
        form_group = QGroupBox("Zadejte spotřebu lobby")
        form_layout = QGridLayout()
        
        self.lobby_widgets = {}  # {id: {'qty': spin, 'price': spin (if has_price)}}
        row = 0
        
        # Načíst aktivní lobby položky
        active_items = [item for item in self.lobby_items if item.get("active", True)]
        
        for item in active_items:
            item_id = item['id']
            name = item['name']
            unit = item.get('unit', 'ks')
            has_price = item.get('has_price', False)
            
            label = QLabel(f"{name} ({unit}):")
            form_layout.addWidget(label, row, 0)
            
            qty_spin = WatermarkSpinBox()
            qty_spin.setValue(0)
            form_layout.addWidget(qty_spin, row, 1)
            
            if has_price:
                price_label = QLabel("Cena:")
                form_layout.addWidget(price_label, row, 2)
                
                price_spin = QDoubleSpinBox()
                price_spin.setRange(0, 100000)
                price_spin.setValue(item.get('default_price', 0))
                price_spin.setSuffix(" Kč")
                form_layout.addWidget(price_spin, row, 3)
                
                self.lobby_widgets[item_id] = {
                    'qty': qty_spin,
                    'price': price_spin,
                    'name': name,
                    'has_price': True
                }
            else:
                self.lobby_widgets[item_id] = {
                    'qty': qty_spin,
                    'name': name,
                    'has_price': False
                }
                # Přesunout placeholder na místo
                form_layout.addWidget(QLabel(""), row, 2)  # prázdné místo
                form_layout.addWidget(QLabel(""), row, 3)
            
            row += 1
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Sekce pro "Na přání"
        custom_group = QGroupBox("Položky na přání")
        custom_layout = QVBoxLayout()
        
        self.lobby_custom_table = QTableWidget()
        self.lobby_custom_table.setColumnCount(4)
        self.lobby_custom_table.setHorizontalHeaderLabels(["Popis", "Množství", "Cena", ""])
        self.lobby_custom_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.lobby_custom_table.setMaximumHeight(150)
        custom_layout.addWidget(self.lobby_custom_table)
        
        btn_add_custom = QPushButton("➕ Přidat řádek")
        btn_add_custom.clicked.connect(self.add_lobby_custom_row)
        custom_layout.addWidget(btn_add_custom)
        
        custom_group.setLayout(custom_layout)
        layout.addWidget(custom_group)
        
        # Poznámka
        note_layout = QHBoxLayout()
        note_layout.addWidget(QLabel("Poznámka:"))
        self.lobby_note = QLineEdit()
        note_layout.addWidget(self.lobby_note)
        layout.addLayout(note_layout)
        
        # Tlačítka
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("💾 Uložit")
        btn_save.clicked.connect(self.save_lobby_data)
        
        btn_clear = QPushButton("🗑 Vymazat")
        btn_clear.clicked.connect(self.clear_lobby_form)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_clear)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        tab.setLayout(layout)
        return tab
    
    def add_lobby_custom_row(self):
        row = self.lobby_custom_table.rowCount()
        self.lobby_custom_table.insertRow(row)
        self.lobby_custom_table.setItem(row, 0, QTableWidgetItem(""))
        self.lobby_custom_table.setItem(row, 1, QTableWidgetItem("0"))
        self.lobby_custom_table.setItem(row, 2, QTableWidgetItem("0"))
        btn_del = QPushButton("🗑")
        btn_del.clicked.connect(lambda checked, r=row: self.lobby_custom_table.removeRow(r))
        self.lobby_custom_table.setCellWidget(row, 3, btn_del)
    
    def create_report_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Měsíční report - Souhrn spotřeb")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Výběr měsíce a roku
        period_layout = QHBoxLayout()
        
        period_layout.addWidget(QLabel("Měsíc:"))
        self.report_month = QComboBox()
        self.report_month.addItems([
            "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
            "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"
        ])
        self.report_month.setCurrentIndex(QDate.currentDate().month() - 1)
        period_layout.addWidget(self.report_month)
        
        period_layout.addWidget(QLabel("Rok:"))
        self.report_year = QSpinBox()
        self.report_year.setRange(2020, 2030)
        self.report_year.setValue(QDate.currentDate().year())
        period_layout.addWidget(self.report_year)
        
        period_layout.addStretch()
        layout.addLayout(period_layout)
        
        # Tlačítka pro kontrolu náhledu
        preview_controls = QHBoxLayout()
        btn_preview = QPushButton("🔍 Zobrazit náhled")
        btn_preview.clicked.connect(self.preview_report)
        
        btn_clear_preview = QPushButton("👁 Smaž náhled")
        btn_clear_preview.clicked.connect(self.clear_preview)
        
        preview_controls.addWidget(btn_preview)
        preview_controls.addWidget(btn_clear_preview)
        preview_controls.addStretch()
        
        layout.addLayout(preview_controls)
        
        # Oblast pro náhled reportu
        self.report_preview = QTextEdit()
        self.report_preview.setReadOnly(True)
        layout.addWidget(self.report_preview)
        
        # Tlačítka pro export
        export_layout = QHBoxLayout()
        
        btn_export_pdf = QPushButton("📄 Export do PDF")
        btn_export_pdf.clicked.connect(self.export_report_pdf)
        
        btn_export_excel = QPushButton("📊 Export do Excel")
        btn_export_excel.clicked.connect(self.export_report_excel)
        
        export_layout.addWidget(btn_export_pdf)
        export_layout.addWidget(btn_export_excel)
        export_layout.addStretch()
        
        layout.addLayout(export_layout)
        
        tab.setLayout(layout)
        return tab
    
    def create_archive_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Zásobník - Historie záznamů")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Horní panel s filtrem a tlačítky
        top_panel = QHBoxLayout()
        
        # Filtr data
        filter_layout = QHBoxLayout()
        
        filter_layout.addWidget(QLabel("Od:"))
        self.archive_from = QDateEdit()
        self.archive_from.setDate(QDate.currentDate().addMonths(-1))
        self.archive_from.setCalendarPopup(True)
        filter_layout.addWidget(self.archive_from)
        
        filter_layout.addWidget(QLabel("Do:"))
        self.archive_to = QDateEdit()
        self.archive_to.setDate(QDate.currentDate())
        self.archive_to.setCalendarPopup(True)
        filter_layout.addWidget(self.archive_to)
        
        btn_filter = QPushButton("Filtrovat")
        btn_filter.clicked.connect(self.filter_archive)
        filter_layout.addWidget(btn_filter)
        
        top_panel.addLayout(filter_layout)
        top_panel.addStretch()
        
        # Tlačítko pro obnovení dat
        btn_refresh = QPushButton("🔄 Obnovit data")
        btn_refresh.clicked.connect(self.refresh_archive)
        top_panel.addWidget(btn_refresh)
        
        layout.addLayout(top_panel)
        
        # Tabulka archivu - nastavit resize policy
        self.archive_table = QTableWidget()
        self.archive_table.setColumnCount(5)
        self.archive_table.setHorizontalHeaderLabels(["Datum", "Typ", "Položky", "Množství", "Akce"])
        self.archive_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.archive_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Nastavit resize policy pro sloupce
        header = self.archive_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Datum - podle obsahu
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Typ - podle obsahu
        header.setSectionResizeMode(2, QHeaderView.Stretch)           # Položky - roztažitelné
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Množství - podle obsahu
        header.setSectionResizeMode(4, QHeaderView.Stretch)           # Akce - roztažitelné
        
        # Nastavit minimální šířky pro sloupce s tlačítky
        self.archive_table.setColumnWidth(4, 200)  # Minimální šířka pro sloupec Akce
        
        layout.addWidget(self.archive_table)
        
        # Načíst data archivu
        self.load_archive_data()
        
        tab.setLayout(layout)
        return tab
    
    def create_settings_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Nadpis
        title_label = QLabel("Nastavení aplikace")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Záložky v nastavení
        settings_tabs = QTabWidget()
        
        # Obecné nastavení
        general_tab = QWidget()
        general_layout = QFormLayout()
        
        self.company_name = QLineEdit(self.settings.get("company_name", ""))
        general_layout.addRow("Název společnosti:", self.company_name)
        
        self.company_address = QLineEdit(self.settings.get("company_address", ""))
        general_layout.addRow("Adresa:", self.company_address)
        
        self.num_rooms = QSpinBox()
        self.num_rooms.setRange(1, 100)
        self.num_rooms.setValue(self.settings.get("num_rooms", 30))
        general_layout.addRow("Počet pokojů:", self.num_rooms)
        
        self.auto_backup = QCheckBox("Automatické zálohování")
        self.auto_backup.setChecked(self.settings.get("auto_backup", True))
        general_layout.addRow("", self.auto_backup)
        
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Světlý", "Tmavý"])
        current_theme = self.settings.get("theme", "light")
        self.theme_combo.setCurrentText("Světlý" if current_theme == "light" else "Tmavý")
        # Přidat signál pro změnu tématu
        self.theme_combo.currentTextChanged.connect(self.theme_changed)
        general_layout.addRow("Téma:", self.theme_combo)
        
        general_tab.setLayout(general_layout)
        
        # Správa wellness položek
        wellness_items_tab = QWidget()
        wellness_layout = QVBoxLayout()
        
        btn_manage_wellness = QPushButton("🛠 Spravovat wellness položky")
        btn_manage_wellness.clicked.connect(lambda: self.manage_items("wellness"))
        wellness_layout.addWidget(btn_manage_wellness)
        
        wellness_info = QLabel("Zde můžete přidávat, upravovat a mazat položky pro wellness odpisy.")
        wellness_info.setStyleSheet("color: gray; font-style: italic; padding: 10px;")
        wellness_layout.addWidget(wellness_info)
        
        wellness_layout.addStretch()
        wellness_items_tab.setLayout(wellness_layout)
        
        # Správa minibar položek
        minibar_items_tab = QWidget()
        minibar_layout = QVBoxLayout()
        
        btn_manage_minibar = QPushButton("🛠 Spravovat minibar položky")
        btn_manage_minibar.clicked.connect(lambda: self.manage_items("minibar"))
        minibar_layout.addWidget(btn_manage_minibar)
        
        minibar_info = QLabel("Zde můžete přidávat, upravovat a mazat položky pro minibary.")
        minibar_info.setStyleSheet("color: gray; font-style: italic; padding: 10px;")
        minibar_layout.addWidget(minibar_info)
        
        minibar_layout.addStretch()
        minibar_items_tab.setLayout(minibar_layout)
        
        # Správa lobby položek (nové)
        lobby_items_tab = QWidget()
        lobby_layout = QVBoxLayout()
        
        btn_manage_lobby = QPushButton("🛠 Spravovat lobby položky")
        btn_manage_lobby.clicked.connect(lambda: self.manage_items("lobby"))
        lobby_layout.addWidget(btn_manage_lobby)
        
        lobby_info = QLabel("Zde můžete přidávat, upravovat a mazat položky pro lobby. U každé položky lze nastavit, zda se má evidovat cena.")
        lobby_info.setStyleSheet("color: gray; font-style: italic; padding: 10px;")
        lobby_layout.addWidget(lobby_info)
        
        lobby_layout.addStretch()
        lobby_items_tab.setLayout(lobby_layout)
        
        # Zálohování
        backup_tab = QWidget()
        backup_layout = QVBoxLayout()
        
        btn_backup_now = QPushButton("💾 Vytvořit zálohu dat")
        btn_backup_now.clicked.connect(self.create_backup)
        backup_layout.addWidget(btn_backup_now)
        
        btn_restore = QPushButton("🔄 Obnovit ze zálohy")
        btn_restore.clicked.connect(self.restore_backup)
        backup_layout.addWidget(btn_restore)
        
        backup_info = QLabel("Zálohy se ukládají do: " + self.settings.get("backup_dir", ""))
        backup_info.setWordWrap(True)
        backup_info.setStyleSheet("color: gray; font-style: italic; padding: 10px;")
        backup_layout.addWidget(backup_info)
        
        backup_layout.addStretch()
        backup_tab.setLayout(backup_layout)
        
        # O programu
        about_tab = QWidget()
        about_layout = QVBoxLayout()
        
        about_text = QLabel(
            "HEM - Inventory & Write-off Manager\n\n"
            "Verze: 0.2.0\n"
            "Vývojář: JAMAsoft\n"
            "Web: www.jamasoft.cz\n\n"
            "Systém pro správu inventářů a odpisů\n"
            "Wellness, Minibary, Lobby a ostatní služby\n\n"
            "© 2025 HEM - Hotel Easy Manager\n\n"
            "Používáno ve Wellness Hotelu Beethoven"
        )
        about_text.setWordWrap(True)
        about_text.setAlignment(Qt.AlignCenter)
        about_layout.addWidget(about_text)
        
        about_layout.addStretch()
        about_tab.setLayout(about_layout)
        
        settings_tabs.addTab(general_tab, "Obecné")
        settings_tabs.addTab(wellness_items_tab, "Wellness položky")
        settings_tabs.addTab(minibar_items_tab, "Minibar položky")
        settings_tabs.addTab(lobby_items_tab, "Lobby položky")
        settings_tabs.addTab(backup_tab, "Zálohování")
        settings_tabs.addTab(about_tab, "O programu")
        
        layout.addWidget(settings_tabs)
        
        # Tlačítko uložit nastavení
        btn_save_settings = QPushButton("💾 Uložit nastavení")
        btn_save_settings.clicked.connect(self.save_settings)
        layout.addWidget(btn_save_settings, alignment=Qt.AlignRight)
        
        tab.setLayout(layout)
        return tab
    
    def theme_changed(self, theme_text):
        """Změna tématu - zobrazení upozornění"""
        QMessageBox.information(self, "Změna tématu", 
                              "Změna vzhledu se projeví po restartování aplikace.")
    
    def clear_preview(self):
        """Smaže náhled reportu"""
        self.report_preview.clear()
    
    def refresh_item_lists(self):
        """Obnoví seznamy položek v aplikaci"""
        self.wellness_items = load_wellness_items()
        self.minibar_items = load_minibar_items()
        self.lobby_items = load_lobby_items()
        
        # Obnovit wellness tab
        self.wellness_tab = self.create_wellness_tab()
        self.tabs.removeTab(0)
        self.tabs.insertTab(0, self.wellness_tab, "🏊 Wellness Odpisy")
        
        # Obnovit minibar tab
        self.minibar_tab = self.create_minibar_writeoff_tab()
        self.tabs.removeTab(1)
        self.tabs.insertTab(1, self.minibar_tab, "🥤 Minibary Odpisy")
        
        # Obnovit lobby tab
        self.lobby_tab = self.create_lobby_tab()
        self.tabs.removeTab(2)
        self.tabs.insertTab(2, self.lobby_tab, "🛋️ Lobby Odpisy")
    
    def manage_items(self, item_type):
        """Otevře dialog pro správu položek"""
        dialog = ManageItemsDialog(item_type, parent=self)
        dialog.exec()
    
    def date_changed(self, new_date):
        """Zpracování změny data"""
        self.current_date = new_date
        self.load_current_date_data()
        self.update_status_bar()
    
    def set_today(self):
        """Nastaví dnešní datum"""
        self.date_edit.setDate(QDate.currentDate())
    
    def load_current_date_data(self):
        """Načte data pro aktuální datum"""
        date_str = self.current_date.toString("yyyy-MM-dd")
        
        # Načíst wellness data
        wellness_data_for_date = self.wellness_data.get(date_str, {})
        for item_id, widgets in self.wellness_widgets.items():
            value = wellness_data_for_date.get(str(item_id), 0)
            widgets['widget'].setValue(value)
        self.wellness_note.setText(wellness_data_for_date.get("note", ""))
        
        # Načíst minibar data
        minibar_data_for_date = self.minibar_data.get(date_str, {})
        for item_id, widgets in self.minibar_widgets.items():
            value = minibar_data_for_date.get(str(item_id), 0)
            widgets['widget'].setValue(value)
        self.minibar_note.setText(minibar_data_for_date.get("note", ""))
        
        # Načíst lobby data
        lobby_data_for_date = self.lobby_data.get(date_str, {})
        items_data = lobby_data_for_date.get("items", {})
        for item_id, widget_dict in self.lobby_widgets.items():
            str_id = str(item_id)
            if str_id in items_data:
                if widget_dict['has_price']:
                    qty = items_data[str_id].get("qty", 0)
                    price = items_data[str_id].get("price", 0)
                    widget_dict['qty'].setValue(qty)
                    widget_dict['price'].setValue(price)
                else:
                    widget_dict['qty'].setValue(items_data[str_id])
            else:
                widget_dict['qty'].setValue(0)
                if widget_dict['has_price']:
                    widget_dict['price'].setValue(0)
        
        # Načíst custom položky
        custom_list = lobby_data_for_date.get("custom", [])
        self.lobby_custom_table.setRowCount(len(custom_list))
        for i, req in enumerate(custom_list):
            self.lobby_custom_table.setItem(i, 0, QTableWidgetItem(req.get("description", "")))
            self.lobby_custom_table.setItem(i, 1, QTableWidgetItem(str(req.get("qty", 0))))
            self.lobby_custom_table.setItem(i, 2, QTableWidgetItem(str(req.get("price", 0))))
            btn_del = QPushButton("🗑")
            btn_del.clicked.connect(lambda checked, r=i: self.lobby_custom_table.removeRow(r))
            self.lobby_custom_table.setCellWidget(i, 3, btn_del)
        
        self.lobby_note.setText(lobby_data_for_date.get("note", ""))
    
    def save_wellness_data(self):
        """Uloží wellness data"""
        date_str = self.current_date.toString("yyyy-MM-dd")
        
        wellness_data_for_date = {}
        for item_id, widgets in self.wellness_widgets.items():
            value = widgets['widget'].value()
            if value > 0:
                wellness_data_for_date[str(item_id)] = value
        
        note = self.wellness_note.text().strip()
        if note:
            wellness_data_for_date["note"] = note
        
        wellness_data_for_date["timestamp"] = datetime.now().isoformat()
        wellness_data_for_date["user"] = os.getlogin()
        
        self.wellness_data[date_str] = wellness_data_for_date
        save_wellness_data(self.wellness_data)
        
        self.add_to_archive(date_str, "wellness", wellness_data_for_date)
        
        QMessageBox.information(self, "Uloženo", "Wellness data byla úspěšně uložena.")
        self.update_status_bar()
    
    def save_minibar_data(self):
        """Uloží minibar data"""
        date_str = self.current_date.toString("yyyy-MM-dd")
        
        minibar_data_for_date = {}
        for item_id, widgets in self.minibar_widgets.items():
            value = widgets['widget'].value()
            if value > 0:
                minibar_data_for_date[str(item_id)] = value
        
        note = self.minibar_note.text().strip()
        if note:
            minibar_data_for_date["note"] = note
        
        minibar_data_for_date["timestamp"] = datetime.now().isoformat()
        minibar_data_for_date["user"] = os.getlogin()
        
        self.minibar_data[date_str] = minibar_data_for_date
        save_minibar_data(self.minibar_data)
        
        self.add_to_archive(date_str, "minibar", minibar_data_for_date)
        
        QMessageBox.information(self, "Uloženo", "Minibar data byla úspěšně uložena.")
        self.update_status_bar()
    
    def save_lobby_data(self):
        """Uloží lobby data včetně cen a položek na přání"""
        date_str = self.current_date.toString("yyyy-MM-dd")
        
        items_data = {}
        for item_id, widget_dict in self.lobby_widgets.items():
            qty = widget_dict['qty'].value()
            if qty > 0:
                if widget_dict['has_price']:
                    price = widget_dict['price'].value()
                    items_data[str(item_id)] = {"qty": qty, "price": price}
                else:
                    items_data[str(item_id)] = qty
        
        custom_list = []
        for row in range(self.lobby_custom_table.rowCount()):
            desc_item = self.lobby_custom_table.item(row, 0)
            qty_item = self.lobby_custom_table.item(row, 1)
            price_item = self.lobby_custom_table.item(row, 2)
            if desc_item and qty_item and price_item:
                desc = desc_item.text().strip()
                qty = qty_item.text().strip()
                price = price_item.text().strip()
                if desc and qty.isdigit() and price.replace('.','',1).isdigit():
                    custom_list.append({
                        "description": desc,
                        "qty": int(qty),
                        "price": float(price)
                    })
        
        note = self.lobby_note.text().strip()
        
        lobby_data_for_date = {
            "items": items_data,
            "custom": custom_list,
            "timestamp": datetime.now().isoformat(),
            "user": os.getlogin()
        }
        if note:
            lobby_data_for_date["note"] = note
        
        self.lobby_data[date_str] = lobby_data_for_date
        save_lobby_data(self.lobby_data)
        
        self.add_to_archive(date_str, "lobby", lobby_data_for_date)
        
        QMessageBox.information(self, "Uloženo", "Lobby data byla úspěšně uložena.")
        self.update_status_bar()
    
    def clear_wellness_form(self):
        """Vymaže wellness formulář"""
        for item_id, widgets in self.wellness_widgets.items():
            widgets['widget'].setValue(0)
        self.wellness_note.clear()
    
    def clear_minibar_form(self):
        """Vymaže minibar formulář"""
        for item_id, widgets in self.minibar_widgets.items():
            widgets['widget'].setValue(0)
        self.minibar_note.clear()
    
    def clear_lobby_form(self):
        """Vymaže lobby formulář"""
        for widget_dict in self.lobby_widgets.values():
            widget_dict['qty'].setValue(0)
            if widget_dict['has_price']:
                widget_dict['price'].setValue(0)
        self.lobby_custom_table.setRowCount(0)
        self.lobby_note.clear()
    
    def add_to_archive(self, date_str, data_type, data):
        """Přidá záznam do archivu"""
        archive_data = load_archive_data()
        
        # Vytvořit popis záznamu
        description = ""
        items_list = []
        
        if data_type == "wellness" or data_type == "minibar":
            for item_id, value in data.items():
                if item_id not in ["note", "timestamp", "user"]:
                    if data_type == "wellness":
                        item_name = next((item['name'] for item in self.wellness_items 
                                        if str(item['id']) == item_id), f"Položka {item_id}")
                    else:
                        item_name = next((item['name'] for item in self.minibar_items 
                                        if str(item['id']) == item_id), f"Položka {item_id}")
                    items_list.append(f"{item_name}: {value}")
        else:  # lobby
            items = data.get("items", {})
            for item_id, value in items.items():
                item = next((item for item in self.lobby_items if str(item['id']) == item_id), None)
                if item:
                    if isinstance(value, dict):
                        items_list.append(f"{item['name']}: {value['qty']} ks (cena {value['price']} Kč)")
                    else:
                        items_list.append(f"{item['name']}: {value} ks")
            for c in data.get("custom", []):
                items_list.append(f"Na přání - {c['description']}: {c['qty']} ks (cena {c['price']} Kč)")
        
        if items_list:
            description = f"{data_type.capitalize()} - " + ", ".join(items_list[:3])
            if len(items_list) > 3:
                description += f" + {len(items_list) - 3} další"
        else:
            description = f"{data_type.capitalize()} - žádné položky"
        
        archive_entry = {
            "date": date_str,
            "type": data_type,
            "data": data,
            "description": description,
            "timestamp": datetime.now().isoformat()
        }
        
        archive_data.append(archive_entry)
        save_archive_data(archive_data)
    
    def load_archive_data(self):
        """Načte data do archivu"""
        archive_data = load_archive_data()
        self.archive_table.setRowCount(len(archive_data))
        
        for row, entry in enumerate(archive_data):
            # Datum
            date_item = QTableWidgetItem(entry["date"])
            self.archive_table.setItem(row, 0, date_item)
            
            # Typ
            type_item = QTableWidgetItem(entry["type"].capitalize())
            self.archive_table.setItem(row, 1, type_item)
            
            # Popis
            desc_item = QTableWidgetItem(entry.get("description", ""))
            self.archive_table.setItem(row, 2, desc_item)
            
            # Množství
            total = 0
            if entry["type"] in ["wellness", "minibar"]:
                total = sum(int(v) for k, v in entry["data"].items() 
                          if k not in ["note", "timestamp", "user"])
            else:  # lobby
                items = entry["data"].get("items", {})
                for v in items.values():
                    if isinstance(v, dict):
                        total += v.get("qty", 0)
                    else:
                        total += v
                total += sum(c.get("qty", 0) for c in entry["data"].get("custom", []))
            
            qty_item = QTableWidgetItem(str(total))
            self.archive_table.setItem(row, 3, qty_item)
            
            # Akce - vytvořit widget s tlačítky
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(2, 2, 2, 2)
            action_layout.setSpacing(5)
            
            btn_edit = QPushButton("✏️ Upravit")
            btn_edit.setFixedSize(80, 25)
            btn_edit.clicked.connect(lambda checked, r=row: self.edit_archive_entry(r))
            
            btn_delete = QPushButton("🗑 Smazat")
            btn_delete.setFixedSize(80, 25)
            btn_delete.clicked.connect(lambda checked, r=row: self.delete_archive_entry(r))
            
            action_layout.addWidget(btn_edit)
            action_layout.addWidget(btn_delete)
            action_layout.addStretch()
            
            action_widget.setLayout(action_layout)
            self.archive_table.setCellWidget(row, 4, action_widget)
    
    def refresh_archive(self):
        """Obnoví data v archivu"""
        self.load_archive_data()
        QMessageBox.information(self, "Obnoveno", "Data archivu byla obnovena.")
    
    def filter_archive(self):
        """Filtruje archiv podle data"""
        from_date = self.archive_from.date().toString("yyyy-MM-dd")
        to_date = self.archive_to.date().toString("yyyy-MM-dd")
        
        archive_data = load_archive_data()
        filtered_data = [
            entry for entry in archive_data
            if from_date <= entry["date"] <= to_date
        ]
        
        self.archive_table.setRowCount(len(filtered_data))
        for row, entry in enumerate(filtered_data):
            self.archive_table.setItem(row, 0, QTableWidgetItem(entry["date"]))
            self.archive_table.setItem(row, 1, QTableWidgetItem(entry["type"].capitalize()))
            self.archive_table.setItem(row, 2, QTableWidgetItem(entry.get("description", "")))
            
            total = 0
            if entry["type"] in ["wellness", "minibar"]:
                total = sum(int(v) for k, v in entry["data"].items() 
                          if k not in ["note", "timestamp", "user"])
            else:  # lobby
                items = entry["data"].get("items", {})
                for v in items.values():
                    if isinstance(v, dict):
                        total += v.get("qty", 0)
                    else:
                        total += v
                total += sum(c.get("qty", 0) for c in entry["data"].get("custom", []))
            
            self.archive_table.setItem(row, 3, QTableWidgetItem(str(total)))
            
            # Akce - vytvořit widget s tlačítky
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(2, 2, 2, 2)
            action_layout.setSpacing(5)
            
            btn_edit = QPushButton("✏️ Upravit")
            btn_edit.setFixedSize(80, 25)
            btn_edit.clicked.connect(lambda checked, r=row: self.edit_archive_entry_filtered(r, filtered_data))
            
            btn_delete = QPushButton("🗑 Smazat")
            btn_delete.setFixedSize(80, 25)
            btn_delete.clicked.connect(lambda checked, r=row: self.delete_archive_entry_filtered(r, filtered_data))
            
            action_layout.addWidget(btn_edit)
            action_layout.addWidget(btn_delete)
            action_layout.addStretch()
            
            action_widget.setLayout(action_layout)
            self.archive_table.setCellWidget(row, 4, action_widget)
    
    def edit_archive_entry(self, row):
        """Editace záznamu v archivu - otevře dialog"""
        archive_data = load_archive_data()
        if 0 <= row < len(archive_data):
            entry = archive_data[row]
            self.open_edit_dialog(entry, row)
    
    def edit_archive_entry_filtered(self, row, filtered_data):
        """Editace filtrovaného záznamu v archivu"""
        if 0 <= row < len(filtered_data):
            entry = filtered_data[row]
            # Najít původní index v plných datech
            archive_data = load_archive_data()
            original_index = None
            for i, item in enumerate(archive_data):
                if item["date"] == entry["date"] and item["type"] == entry["type"] and item.get("timestamp") == entry.get("timestamp"):
                    original_index = i
                    break
            
            if original_index is not None:
                self.open_edit_dialog(entry, original_index)
    
    def open_edit_dialog(self, entry, original_index):
        """Otevře dialog pro editaci záznamu"""
        dialog = EditArchiveDialog(entry, self.wellness_items, self.minibar_items, self.lobby_items, self)
        if dialog.exec():
            updated_entry = dialog.get_updated_entry()
            
            # Aktualizovat data v paměti
            if updated_entry["type"] == "wellness":
                date_str = updated_entry["date"]
                self.wellness_data[date_str] = updated_entry["data"]
                save_wellness_data(self.wellness_data)
            elif updated_entry["type"] == "minibar":
                date_str = updated_entry["date"]
                self.minibar_data[date_str] = updated_entry["data"]
                save_minibar_data(self.minibar_data)
            else:  # lobby
                date_str = updated_entry["date"]
                self.lobby_data[date_str] = updated_entry["data"]
                save_lobby_data(self.lobby_data)
            
            # Aktualizovat archiv
            archive_data = load_archive_data()
            if 0 <= original_index < len(archive_data):
                archive_data[original_index] = updated_entry
                save_archive_data(archive_data)
                
                # Znovu načíst archiv
                self.load_archive_data()
                
                QMessageBox.information(self, "Upraveno", "Záznam byl úspěšně upraven.")
            else:
                QMessageBox.warning(self, "Chyba", "Nepodařilo se najít záznam k úpravě.")
    
    def delete_archive_entry(self, row):
        """Smazání záznamu z archivu"""
        reply = QMessageBox.question(
            self, "Smazat záznam",
            "Opravdu chcete smazat tento záznam?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            archive_data = load_archive_data()
            if 0 <= row < len(archive_data):
                # Také smazat z hlavních dat
                entry = archive_data[row]
                date_str = entry["date"]
                data_type = entry["type"]
                
                if data_type == "wellness" and date_str in self.wellness_data:
                    del self.wellness_data[date_str]
                    save_wellness_data(self.wellness_data)
                elif data_type == "minibar" and date_str in self.minibar_data:
                    del self.minibar_data[date_str]
                    save_minibar_data(self.minibar_data)
                elif data_type == "lobby" and date_str in self.lobby_data:
                    del self.lobby_data[date_str]
                    save_lobby_data(self.lobby_data)
                
                # Smazat z archivu
                del archive_data[row]
                save_archive_data(archive_data)
                
                # Znovu načíst archiv
                self.load_archive_data()
                
                QMessageBox.information(self, "Smazáno", "Záznam byl úspěšně smazán.")
    
    def delete_archive_entry_filtered(self, row, filtered_data):
        """Smazání filtrovaného záznamu z archivu"""
        if 0 <= row < len(filtered_data):
            entry = filtered_data[row]
            
            reply = QMessageBox.question(
                self, "Smazat záznam",
                f"Opravdu chcete smazat záznam z {entry['date']}?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Smazat z hlavních dat
                date_str = entry["date"]
                data_type = entry["type"]
                
                if data_type == "wellness" and date_str in self.wellness_data:
                    del self.wellness_data[date_str]
                    save_wellness_data(self.wellness_data)
                elif data_type == "minibar" and date_str in self.minibar_data:
                    del self.minibar_data[date_str]
                    save_minibar_data(self.minibar_data)
                elif data_type == "lobby" and date_str in self.lobby_data:
                    del self.lobby_data[date_str]
                    save_lobby_data(self.lobby_data)
                
                # Smazat z archivu
                archive_data = load_archive_data()
                new_archive_data = []
                for item in archive_data:
                    if not (item["date"] == date_str and item["type"] == data_type and item.get("timestamp") == entry.get("timestamp")):
                        new_archive_data.append(item)
                
                save_archive_data(new_archive_data)
                
                # Znovu načíst archiv
                self.load_archive_data()
                
                QMessageBox.information(self, "Smazáno", "Záznam byl úspěšně smazán.")
    
    def preview_report(self):
        """Náhled měsíčního reportu včetně lobby"""
        month = self.report_month.currentIndex() + 1
        year = self.report_year.value()
        
        # Získat data pro daný měsíc
        wellness_month_data = {}
        minibar_month_data = {}
        lobby_month_data = {"items": {}, "custom": [], "total_value": 0}
        
        # Wellness data
        for date_str, data in self.wellness_data.items():
            try:
                entry_date = datetime.strptime(date_str, "%Y-%m-%d")
                if entry_date.year == year and entry_date.month == month:
                    for item_id, value in data.items():
                        if item_id not in ["note", "timestamp", "user"]:
                            if item_id not in wellness_month_data:
                                wellness_month_data[item_id] = 0
                            wellness_month_data[item_id] += int(value)
            except:
                continue
        
        # Minibar data
        for date_str, data in self.minibar_data.items():
            try:
                entry_date = datetime.strptime(date_str, "%Y-%m-%d")
                if entry_date.year == year and entry_date.month == month:
                    for item_id, value in data.items():
                        if item_id not in ["note", "timestamp", "user"]:
                            if item_id not in minibar_month_data:
                                minibar_month_data[item_id] = 0
                            minibar_month_data[item_id] += int(value)
            except:
                continue
        
        # Lobby data
        for date_str, data in self.lobby_data.items():
            try:
                entry_date = datetime.strptime(date_str, "%Y-%m-%d")
                if entry_date.year == year and entry_date.month == month:
                    items = data.get("items", {})
                    for item_id, val in items.items():
                        if isinstance(val, dict):
                            qty = val.get("qty", 0)
                            price = val.get("price", 0)
                            if item_id not in lobby_month_data["items"]:
                                lobby_month_data["items"][item_id] = {"qty": 0, "value": 0}
                            lobby_month_data["items"][item_id]["qty"] += qty
                            lobby_month_data["items"][item_id]["value"] += qty * price
                            lobby_month_data["total_value"] += qty * price
                        else:
                            qty = val
                            if item_id not in lobby_month_data["items"]:
                                lobby_month_data["items"][item_id] = {"qty": 0, "value": 0}
                            lobby_month_data["items"][item_id]["qty"] += qty
                            # Pokud není cena, nepočítáme hodnotu
                    
                    for c in data.get("custom", []):
                        lobby_month_data["custom"].append(c)
                        lobby_month_data["total_value"] += c["qty"] * c["price"]
            except:
                continue
        
        # Uložit data pro export
        self.report_data = {
            'wellness': wellness_month_data,
            'minibar': minibar_month_data,
            'lobby': lobby_month_data,
            'month': month,
            'year': year
        }
        
        # Generovat náhled
        preview_text = f"# MĚSÍČNÍ REPORT - {month}/{year}\n\n"
        preview_text += f"Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        preview_text += "=" * 50 + "\n\n"
        
        # Wellness sekce
        preview_text += "## WELLNESS ODPISY\n\n"
        if wellness_month_data:
            total_wellness = 0
            for item_id, total in wellness_month_data.items():
                item_name = next((item['name'] for item in self.wellness_items 
                                if str(item['id']) == item_id), f"Položka {item_id}")
                preview_text += f"- {item_name}: {total} ks\n"
                total_wellness += total
            preview_text += f"\n**Celkem wellness položek: {total_wellness} ks**\n"
        else:
            preview_text += "Žádná data pro tento měsíc.\n"
        
        preview_text += "\n" + "=" * 50 + "\n\n"
        
        # Minibar sekce
        preview_text += "## MINIBARY ODPISY\n\n"
        if minibar_month_data:
            total_minibar = 0
            total_value = 0
            
            for item_id, total in minibar_month_data.items():
                item = next((item for item in self.minibar_items 
                           if str(item['id']) == item_id), None)
                if item:
                    item_name = item['name']
                    price = item.get('price', 0)
                    value = total * price
                    
                    preview_text += f"- {item_name}: {total} ks"
                    if price > 0:
                        preview_text += f" (hodnota: {value:,} Kč)"
                    preview_text += "\n"
                    
                    total_minibar += total
                    total_value += value
            
            preview_text += f"\n**Celkem minibar položek: {total_minibar} ks**\n"
            if total_value > 0:
                preview_text += f"**Celková hodnota: {total_value:,} Kč**\n"
        else:
            preview_text += "Žádná data pro tento měsíc.\n"
        
        preview_text += "\n" + "=" * 50 + "\n\n"
        
        # Lobby sekce
        preview_text += "## LOBBY ODPISY\n\n"
        if lobby_month_data["items"] or lobby_month_data["custom"]:
            total_lobby_qty = 0
            total_lobby_value = lobby_month_data["total_value"]
            
            # Předdefinované položky
            for item_id, vals in lobby_month_data["items"].items():
                item = next((item for item in self.lobby_items if str(item['id']) == item_id), None)
                if item:
                    item_name = item['name']
                    qty = vals["qty"]
                    value = vals["value"]
                    preview_text += f"- {item_name}: {qty} ks"
                    if value > 0:
                        preview_text += f" (hodnota: {value:,} Kč)"
                    preview_text += "\n"
                    total_lobby_qty += qty
            
            # Položky na přání
            for c in lobby_month_data["custom"]:
                preview_text += f"- Na přání - {c['description']}: {c['qty']} ks (cena {c['price']} Kč, hodnota {c['qty'] * c['price']:,} Kč)\n"
                total_lobby_qty += c['qty']
            
            preview_text += f"\n**Celkem lobby položek: {total_lobby_qty} ks**\n"
            if total_lobby_value > 0:
                preview_text += f"**Celková hodnota: {total_lobby_value:,} Kč**\n"
        else:
            preview_text += "Žádná data pro tento měsíc.\n"
        
        preview_text += "\n" + "=" * 50 + "\n\n"
        preview_text += f"© {self.settings.get('company_name', 'Wellness Hotel Beethoven')}"
        
        self.report_preview.setPlainText(preview_text)
    
    def export_report_pdf(self):
        """Export reportu do PDF včetně lobby"""
        if not hasattr(self, 'report_data'):
            QMessageBox.warning(self, "Upozornění", "Nejdříve vygenerujte náhled reportu.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit PDF report", "", "PDF Files (*.pdf)"
        )
        
        if file_path:
            try:
                pdfmetrics.registerFont(TTFont("Arial", r"C:\Windows\Fonts\arial.ttf"))
                
                c = canvas.Canvas(file_path, pagesize=A4)
                w, h = A4
                
                # Nadpis
                c.setFont("Arial", 16)
                month_name = self.report_month.currentText()
                c.drawString(50, h - 50, f"MĚSÍČNÍ REPORT - {month_name} {self.report_data['year']}")
                
                # Informace
                c.setFont("Arial", 10)
                c.drawString(50, h - 80, f"Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
                c.drawString(50, h - 95, f"Společnost: {self.settings.get('company_name', '')}")
                
                c.line(50, h - 110, w - 50, h - 110)
                
                y = h - 140
                
                # Wellness sekce
                c.setFont("Arial", 12)
                c.drawString(50, y, "WELLNESS ODPISY")
                y -= 25
                
                c.setFont("Arial", 10)
                total_wellness = 0
                for item_id, total in self.report_data['wellness'].items():
                    item_name = next((item['name'] for item in self.wellness_items 
                                    if str(item['id']) == item_id), f"Položka {item_id}")
                    c.drawString(70, y, f"{item_name}: {total} ks")
                    y -= 20
                    total_wellness += total
                if total_wellness > 0:
                    y -= 10
                    c.setFont("Arial", 10)
                    c.drawString(70, y, f"Celkem: {total_wellness} ks")
                    y -= 25
                else:
                    c.drawString(70, y, "Žádná data")
                    y -= 25
                
                y -= 10
                c.line(50, y, w - 50, y)
                y -= 25
                
                # Minibar sekce
                c.setFont("Arial", 12)
                c.drawString(50, y, "MINIBARY ODPISY")
                y -= 25
                
                c.setFont("Arial", 10)
                total_minibar = 0
                total_minibar_value = 0
                for item_id, total in self.report_data['minibar'].items():
                    item = next((item for item in self.minibar_items 
                               if str(item['id']) == item_id), None)
                    if item:
                        item_name = item['name']
                        price = item.get('price', 0)
                        value = total * price
                        c.drawString(70, y, f"{item_name}: {total} ks")
                        if price > 0:
                            c.drawString(200, y, f"Hodnota: {value:,} Kč")
                        y -= 20
                        total_minibar += total
                        total_minibar_value += value
                if total_minibar > 0:
                    y -= 10
                    c.drawString(70, y, f"Celkem položek: {total_minibar} ks")
                    y -= 20
                    if total_minibar_value > 0:
                        c.drawString(70, y, f"Celková hodnota: {total_minibar_value:,} Kč")
                        y -= 25
                else:
                    c.drawString(70, y, "Žádná data")
                    y -= 25
                
                y -= 10
                c.line(50, y, w - 50, y)
                y -= 25
                
                # Lobby sekce
                c.setFont("Arial", 12)
                c.drawString(50, y, "LOBBY ODPISY")
                y -= 25
                
                c.setFont("Arial", 10)
                lobby_data = self.report_data['lobby']
                total_lobby_qty = 0
                total_lobby_value = lobby_data["total_value"]
                
                # Předdefinované položky
                for item_id, vals in lobby_data["items"].items():
                    item = next((item for item in self.lobby_items if str(item['id']) == item_id), None)
                    if item:
                        item_name = item['name']
                        qty = vals["qty"]
                        value = vals["value"]
                        c.drawString(70, y, f"{item_name}: {qty} ks")
                        if value > 0:
                            c.drawString(200, y, f"Hodnota: {value:,} Kč")
                        y -= 20
                        total_lobby_qty += qty
                
                # Položky na přání
                for c_item in lobby_data["custom"]:
                    c.drawString(70, y, f"Na přání - {c_item['description']}: {c_item['qty']} ks")
                    c.drawString(200, y, f"Hodnota: {c_item['qty'] * c_item['price']:,} Kč")
                    y -= 20
                    total_lobby_qty += c_item['qty']
                
                if total_lobby_qty > 0:
                    y -= 10
                    c.drawString(70, y, f"Celkem položek: {total_lobby_qty} ks")
                    y -= 20
                    if total_lobby_value > 0:
                        c.drawString(70, y, f"Celková hodnota: {total_lobby_value:,} Kč")
                        y -= 25
                else:
                    c.drawString(70, y, "Žádná data")
                    y -= 25
                
                # Patička
                y -= 30
                c.line(50, y, w - 50, y)
                y -= 20
                c.drawString(50, y, f"© {self.settings.get('company_name', 'Wellness Hotel Beethoven')}")
                c.drawString(w - 150, y, f"Strana 1/1")
                
                c.save()
                
                QMessageBox.information(self, "Hotovo", f"PDF report byl uložen:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se vytvořit PDF: {str(e)}")
    
    def export_report_excel(self):
        """Export reportu do Excelu včetně lobby"""
        if not hasattr(self, 'report_data'):
            QMessageBox.warning(self, "Upozornění", "Nejdříve vygenerujte náhled reportu.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit Excel report", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Report_{self.report_data['month']}_{self.report_data['year']}"
                
                # Nadpis
                ws['A1'] = f"MĚSÍČNÍ REPORT - {self.report_month.currentText()} {self.report_data['year']}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:E1')
                
                # Informace
                ws['A3'] = f"Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                ws['A4'] = f"Společnost: {self.settings.get('company_name', '')}"
                
                # Wellness sekce
                ws['A6'] = "WELLNESS ODPISY"
                ws['A6'].font = Font(bold=True, size=12)
                
                row = 7
                total_wellness = 0
                for item_id, total in self.report_data['wellness'].items():
                    item_name = next((item['name'] for item in self.wellness_items 
                                    if str(item['id']) == item_id), f"Položka {item_id}")
                    ws[f'A{row}'] = item_name
                    ws[f'B{row}'] = total
                    ws[f'C{row}'] = "ks"
                    total_wellness += total
                    row += 1
                if total_wellness > 0:
                    ws[f'A{row}'] = "CELKEM"
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = total_wellness
                    ws[f'B{row}'].font = Font(bold=True)
                    ws[f'C{row}'] = "ks"
                else:
                    ws[f'A{row}'] = "Žádná data"
                
                row += 3
                
                # Minibar sekce
                ws[f'A{row}'] = "MINIBARY ODPISY"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Položka"
                ws[f'B{row}'] = "Množství"
                ws[f'C{row}'] = "Jednotka"
                ws[f'D{row}'] = "Cena/ks"
                ws[f'E{row}'] = "Celková hodnota"
                header_font = Font(bold=True)
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                total_minibar = 0
                total_minibar_value = 0
                for item_id, total in self.report_data['minibar'].items():
                    item = next((item for item in self.minibar_items 
                               if str(item['id']) == item_id), None)
                    if item:
                        ws[f'A{row}'] = item['name']
                        ws[f'B{row}'] = total
                        ws[f'C{row}'] = item.get('unit', 'ks')
                        price = item.get('price', 0)
                        ws[f'D{row}'] = price
                        value = total * price
                        ws[f'E{row}'] = value
                        total_minibar += total
                        total_minibar_value += value
                        row += 1
                if total_minibar > 0:
                    ws[f'A{row}'] = "CELKEM"
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = total_minibar
                    ws[f'B{row}'].font = Font(bold=True)
                    ws[f'C{row}'] = "ks"
                    ws[f'E{row}'] = total_minibar_value
                    ws[f'E{row}'].font = Font(bold=True)
                else:
                    ws[f'A{row}'] = "Žádná data"
                
                row += 3
                
                # Lobby sekce
                ws[f'A{row}'] = "LOBBY ODPISY"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Položka"
                ws[f'B{row}'] = "Množství"
                ws[f'C{row}'] = "Jednotka"
                ws[f'D{row}'] = "Cena/ks"
                ws[f'E{row}'] = "Celková hodnota"
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                lobby_data = self.report_data['lobby']
                total_lobby_qty = 0
                total_lobby_value = lobby_data["total_value"]
                
                # Předdefinované položky
                for item_id, vals in lobby_data["items"].items():
                    item = next((item for item in self.lobby_items if str(item['id']) == item_id), None)
                    if item:
                        ws[f'A{row}'] = item['name']
                        ws[f'B{row}'] = vals["qty"]
                        ws[f'C{row}'] = item.get('unit', 'ks')
                        # Průměrná cena za kus? Nebo zobrazit celkovou hodnotu
                        ws[f'D{row}'] = vals["value"] / vals["qty"] if vals["qty"] > 0 else 0
                        ws[f'E{row}'] = vals["value"]
                        total_lobby_qty += vals["qty"]
                        row += 1
                
                # Položky na přání
                for c_item in lobby_data["custom"]:
                    ws[f'A{row}'] = f"Na přání: {c_item['description']}"
                    ws[f'B{row}'] = c_item['qty']
                    ws[f'C{row}'] = "ks"
                    ws[f'D{row}'] = c_item['price']
                    ws[f'E{row}'] = c_item['qty'] * c_item['price']
                    total_lobby_qty += c_item['qty']
                    row += 1
                
                if total_lobby_qty > 0:
                    ws[f'A{row}'] = "CELKEM"
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = total_lobby_qty
                    ws[f'B{row}'].font = Font(bold=True)
                    ws[f'C{row}'] = "ks"
                    ws[f'E{row}'] = total_lobby_value
                    ws[f'E{row}'].font = Font(bold=True)
                else:
                    ws[f'A{row}'] = "Žádná data"
                
                # Nastavit šířky sloupců
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                
                QMessageBox.information(self, "Hotovo", f"Excel report byl uložen:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se vytvořit Excel: {str(e)}")
    
    def save_settings(self):
        """Uloží nastavení"""
        self.settings = {
            "company_name": self.company_name.text(),
            "company_address": self.company_address.text(),
            "num_rooms": self.num_rooms.value(),
            "auto_backup": self.auto_backup.isChecked(),
            "backup_dir": self.settings.get("backup_dir", ""),
            "theme": "light" if self.theme_combo.currentText() == "Světlý" else "dark",
            "report_language": "cs",
            "currency": "CZK"
        }
        
        save_settings(self.settings)
        
        # Znovu načíst nastavení
        self.settings = load_settings()
        
        QMessageBox.information(self, "Nastavení", "Nastavení bylo úspěšně uloženo.")
    
    def create_backup(self):
        """Vytvoří zálohu dat"""
        try:
            backup_dir = self.settings.get("backup_dir", "")
            if not backup_dir:
                backup_dir = os.path.join(Path.home(), "Desktop", "HEM_Inventory_Backups")
                self.settings["backup_dir"] = backup_dir
                save_settings(self.settings)
            
            os.makedirs(backup_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}")
            os.makedirs(backup_path, exist_ok=True)
            
            # Zkopírovat všechny JSON soubory
            files_to_backup = [
                SETTINGS_FILE, WELLNESS_DATA_FILE, MINIBAR_DATA_FILE, LOBBY_DATA_FILE,
                WELLNESS_ITEMS_FILE, MINIBAR_ITEMS_FILE, LOBBY_ITEMS_FILE, ARCHIVE_DATA_FILE
            ]
            
            for file_path in files_to_backup:
                if os.path.exists(file_path):
                    shutil.copy2(file_path, os.path.join(backup_path, os.path.basename(file_path)))
            
            QMessageBox.information(self, "Záloha", f"Záloha byla vytvořena:\n{backup_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se vytvořit zálohu: {str(e)}")
    
    def restore_backup(self):
        """Obnoví data ze zálohy"""
        try:
            backup_dir = self.settings.get("backup_dir", "")
            if not backup_dir or not os.path.exists(backup_dir):
                QMessageBox.warning(self, "Upozornění", "Nenalezena žádná záloha.")
                return
            
            # Najít nejnovější zálohu
            backups = []
            for item in os.listdir(backup_dir):
                item_path = os.path.join(backup_dir, item)
                if os.path.isdir(item_path) and item.startswith("backup_"):
                    backups.append((item_path, os.path.getmtime(item_path)))
            
            if not backups:
                QMessageBox.warning(self, "Upozornění", "Nenalezena žádná záloha.")
                return
            
            backups.sort(key=lambda x: x[1], reverse=True)
            latest_backup = backups[0][0]
            
            reply = QMessageBox.question(
                self, "Obnovit zálohu",
                f"Obnovit data ze zálohy:\n{latest_backup}?\n\nAktuální data budou přepsána.",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Obnovit soubory
                for filename in os.listdir(latest_backup):
                    if filename.endswith('.json'):
                        src = os.path.join(latest_backup, filename)
                        dst = os.path.join(APP_DIR, filename)
                        shutil.copy2(src, dst)
                
                # Znovu načíst data
                self.settings = load_settings()
                self.wellness_items = load_wellness_items()
                self.minibar_items = load_minibar_items()
                self.lobby_items = load_lobby_items()
                self.wellness_data = load_wellness_data()
                self.minibar_data = load_minibar_data()
                self.lobby_data = load_lobby_data()
                
                # Obnovit UI
                self.refresh_item_lists()
                self.load_archive_data()
                
                QMessageBox.information(self, "Obnoveno", "Data byla úspěšně obnovena ze zálohy.")
                
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se obnovit zálohu: {str(e)}")
    
    def update_status_bar(self):
        """Aktualizuje status bar"""
        date_str = self.current_date.toString("dd.MM.yyyy")
        
        # Spočítat záznamy pro aktuální datum
        wellness_count = 0
        minibar_count = 0
        lobby_count = 0
        
        date_key = self.current_date.toString("yyyy-MM-dd")
        
        if date_key in self.wellness_data:
            wellness_data = self.wellness_data[date_key]
            wellness_count = sum(1 for k in wellness_data.keys() 
                               if k not in ["note", "timestamp", "user"])
        
        if date_key in self.minibar_data:
            minibar_data = self.minibar_data[date_key]
            minibar_count = sum(1 for k in minibar_data.keys() 
                              if k not in ["note", "timestamp", "user"])
        
        if date_key in self.lobby_data:
            lobby_data = self.lobby_data[date_key]
            lobby_count = len(lobby_data.get("items", {})) + len(lobby_data.get("custom", []))
        
        status_text = f"Datum: {date_str} | "
        status_text += f"Wellness: {wellness_count} položek | "
        status_text += f"Minibary: {minibar_count} položek | "
        status_text += f"Lobby: {lobby_count} položek"
        
        self.status_bar.showMessage(status_text)

# ================= SPUŠTĚNÍ APLIKACE =================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Nastavit téma
    settings = load_settings()
    theme = settings.get("theme", "light")
    setup_theme(app, theme)
    
    # Vytvořit hlavní okno
    window = MainWindow()
    window.setWindowTitle("HEM - Inventory & Write-off Manager v0.2.0")
    window.show()
    
    sys.exit(app.exec())