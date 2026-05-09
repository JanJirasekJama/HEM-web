import os
import sys
import json
from datetime import datetime, timedelta
import shutil
import subprocess
import csv
import sqlite3
import hashlib
import webbrowser
from io import BytesIO
from decimal import Decimal
import random
import string

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit,
    QComboBox, QPushButton, QMessageBox, QVBoxLayout, QFormLayout,
    QHBoxLayout, QDialog, QTabWidget, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QMenu,
    QGroupBox, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit,
    QFileDialog, QProgressBar, QTreeWidget, QTreeWidgetItem,
    QSplitter, QStatusBar, QToolBar, QToolButton, QGridLayout,
    QScrollArea, QFrame, QListWidget, QListWidgetItem,
    QRadioButton, QButtonGroup, QInputDialog, QDialogButtonBox,
    QStackedWidget, QCalendarWidget
)
from PySide6.QtCore import (
    Qt, QTimer, QDateTime, QDate, QTime, QThread, Signal, 
    QSettings, QPoint, QSize, QRect, QEvent, QPropertyAnimation,
    QEasingCurve, QDir
)
from PySide6.QtGui import (
    QPalette, QColor, QIcon, QAction, QFont, QFontMetrics,
    QPainter, QBrush, QPen, QLinearGradient, QPixmap, QImage,
    QKeySequence, QShortcut, QCursor, QMouseEvent, QKeyEvent,
    QDesktopServices
)
from PySide6.QtPrintSupport import QPrintDialog, QPrinter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm

# ================= NOVÉ IMPORTY PRO POKROČILÉ FUNKCE =================
from reportlab.lib.utils import ImageReader
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib
matplotlib.use('Qt5Agg')
import numpy as np
from collections import Counter, defaultdict
import requests
from bs4 import BeautifulSoup
import zipfile
import tempfile
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

# ================= CESTY =================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

APP_DIR = os.path.join(os.environ["LOCALAPPDATA"], "HEM_ZalohoveFaktury")
os.makedirs(APP_DIR, exist_ok=True)

SETTINGS_FILE = os.path.join(APP_DIR, "settings.json")
COUNTER_FILE = os.path.join(APP_DIR, "faktury_counter.txt")
ARCHIVE_DIR = os.path.join(APP_DIR, "archiv_faktur")
ARCHIVE_DATA_FILE = os.path.join(APP_DIR, "archiv_data.json")
EMAIL_SETTINGS_FILE = os.path.join(APP_DIR, "email_settings.json")
CASH_REGISTER_FILE = os.path.join(APP_DIR, "cash_register.json")
TAX_REPORT_FILE = os.path.join(APP_DIR, "tax_reports.json")
TEMPLATES_DIR = os.path.join(APP_DIR, "templates")
USERS_FILE = os.path.join(APP_DIR, "users.json")
BACKUP_DIR = os.path.join(APP_DIR, "backups")
CUSTOMERS_FILE = os.path.join(APP_DIR, "customers.json")
RECOVERY_FILE = os.path.join(APP_DIR, "recovery_data.json")
SERVICES_FILE = os.path.join(APP_DIR, "services.json")
SPLATNOSTI_FILE = os.path.join(APP_DIR, "splatnosti.json")
LAST_LOGIN_FILE = os.path.join(APP_DIR, "last_login.json")

COUNTER_PASSWORD = "061004"

# ================= DEFAULT NASTAVENÍ =================
DEFAULT_SETTINGS = {
    "open_pdf": True,
    "output_dir": BASE_DIR,
    "theme": "system",
    "email_notifications": False,
    "auto_backup": True,
    "backup_interval_days": 7,
    "qr_code_enabled": False,
    "auto_sync": False,
    "reminder_enabled": False,
    "tax_rate": 21,
    "currency": "CZK",
    "company_name": "TRIDENT GROUP 007 s.r.o.",
    "company_address": "Nám. Jiřího z Lobkovic 2406/9, 130 00 Praha 3 – Vinohrady",
    "company_id": "27262405",
    "company_vat": "CZ27262405",
    "branch_same": True,
    "branch_name": "",
    "branch_address": "",
}

# ================= DEFAULT SLUŽBY =================
DEFAULT_SERVICES = {
    "Wellness": [
        {"nazev": "Vířivka", "cena": 2000, "aktivni": True, "typ": "wellness"},
        {"nazev": "Vířivka + sauna", "cena": 3000, "aktivni": True, "typ": "wellness"},
        {"nazev": "Pouze sauna", "cena": 1500, "aktivni": True, "typ": "wellness"}
    ],
    "Ubytování": [
        {"nazev": "Pokoj 1 lůžkový", "cena": 1000, "aktivni": True, "typ": "ubytovani"},
        {"nazev": "Pokoj 2 lůžkový", "cena": 1800, "aktivni": True, "typ": "ubytovani"},
        {"nazev": "Apartmá", "cena": 3500, "aktivni": True, "typ": "ubytovani"}
    ],
    "Ostatní služby": [
        {"nazev": "Balíček Restaurace", "cena": 500, "aktivni": True, "typ": "ostatni"},
        {"nazev": "Balíček Ubytování + Restaurace", "cena": 2200, "aktivni": True, "typ": "ostatni"},
        {"nazev": "Balíček Ubytování + Wellness", "cena": 4500, "aktivni": True, "typ": "ostatni"}
    ]
}

# ================= DEFAULT SPLATNOSTI (s jednotkou) =================
DEFAULT_SPLATNOSTI = [
    {"nazev": "okamžitě", "hodiny": 0, "jednotka": "hodiny", "aktivni": True},
    {"nazev": "4 hodiny", "hodiny": 4, "jednotka": "hodiny", "aktivni": True},
    {"nazev": "24 hodiny", "hodiny": 24, "jednotka": "hodiny", "aktivni": True},
    {"nazev": "1 den", "hodiny": 1, "jednotka": "dny", "aktivni": True},
    {"nazev": "3 dny", "hodiny": 3, "jednotka": "dny", "aktivni": True},
    {"nazev": "7 dní", "hodiny": 7, "jednotka": "dny", "aktivni": True},
]

# ================= POMOCNÉ FUNKCE =================
def desktop_path():
    return os.path.join(os.path.expanduser("~"), "Desktop")

def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_SETTINGS, f, indent=2, ensure_ascii=False)
        return DEFAULT_SETTINGS.copy()
    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_settings(data):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

settings = load_settings()

def load_services():
    if not os.path.exists(SERVICES_FILE):
        with open(SERVICES_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_SERVICES, f, indent=2, ensure_ascii=False)
        return DEFAULT_SERVICES.copy()
    with open(SERVICES_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_services(data):
    with open(SERVICES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

services = load_services()

def load_splatnosti():
    """Načte seznam splatností ze souboru, doplní chybějící jednotku."""
    if not os.path.exists(SPLATNOSTI_FILE):
        with open(SPLATNOSTI_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_SPLATNOSTI, f, indent=2, ensure_ascii=False)
        return DEFAULT_SPLATNOSTI.copy()
    with open(SPLATNOSTI_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Zpětná kompatibilita – pokud chybí jednotka, přidáme "hodiny"
    for item in data:
        if "jednotka" not in item:
            item["jednotka"] = "hodiny"
    return data

def save_splatnosti(data):
    """Uloží seznam splatností do souboru"""
    with open(SPLATNOSTI_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_last_login():
    """Načte poslední přihlášení ze souboru"""
    if not os.path.exists(LAST_LOGIN_FILE):
        return {"username": "", "password": ""}
    try:
        with open(LAST_LOGIN_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {"username": "", "password": ""}

def save_last_login(username, password):
    """Uloží poslední přihlášení do souboru"""
    # Neukládáme admin účet
    if username == "admin":
        return
    data = {"username": username, "password": password}
    with open(LAST_LOGIN_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def clear_last_login():
    """Vymaže uložené přihlášení"""
    if os.path.exists(LAST_LOGIN_FILE):
        os.remove(LAST_LOGIN_FILE)

def get_all_active_services():
    all_services = []
    for category, service_list in services.items():
        for service in service_list:
            if service.get("aktivni", True):
                all_services.append({
                    "nazev": service["nazev"],
                    "cena": service["cena"],
                    "kategorie": category,
                    "typ": service.get("typ", "ostatni")
                })
    return all_services

def get_all_active_splatnosti():
    """Vrátí seznam aktivních splatností"""
    splatnosti = load_splatnosti()
    return [s for s in splatnosti if s.get("aktivni", True)]

# ================= VYTVOŘENÍ VŠECH SLOŽEK =================
for directory in [ARCHIVE_DIR, TEMPLATES_DIR, BACKUP_DIR]:
    os.makedirs(directory, exist_ok=True)

# ================= ARCHIV FAKTUR =================
def ensure_archive_dir():
    """Vytvoří složku pro archiv faktur pokud neexistuje"""
    os.makedirs(ARCHIVE_DIR, exist_ok=True)

def load_archive_data():
    """Načte data archivu faktur"""
    if not os.path.exists(ARCHIVE_DATA_FILE):
        return []
    try:
        with open(ARCHIVE_DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

def save_archive_data(data):
    """Uloží data archivu faktur"""
    with open(ARCHIVE_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def add_to_archive(faktura_data, cislo_faktury, filepath, due_date_str, vydal=None):
    """Přidá fakturu do archivu"""
    ensure_archive_dir()
    
    # Zkopírovat PDF do archivu
    archive_filepath = os.path.join(ARCHIVE_DIR, f"faktura_{cislo_faktury}.pdf")
    shutil.copy2(filepath, archive_filepath)
    
    # Načíst stávající data
    archive_data = load_archive_data()
    
    # Určit stav faktury podle času vytvoření (nyní)
    vytvoreni = datetime.now()
    stav = get_payment_status(faktura_data["splatnost"], vytvoreni)
    
    # Přidat nový záznam
    new_record = {
        "cislo_faktury": cislo_faktury,
        "jmeno": faktura_data["jmeno"],
        "termin": faktura_data["termin"],
        "datum_vytvoreni": vytvoreni.strftime("%d.%m.%Y %H:%M:%S"),
        "stav": stav,
        "splatnost": faktura_data["splatnost"],          # název splatnosti
        "due_date": due_date_str,                        # konkrétní datum splatnosti
        "cena": faktura_data["cena"],
        "sluzba": faktura_data["sluzba"],
        "poznamka": faktura_data.get("poznamka", ""),
        "archiv_path": archive_filepath,
        "email": faktura_data.get("email", ""),
        "telefon": faktura_data.get("telefon", ""),
        "vydal": vydal if vydal else "Neznámý"
    }
    
    archive_data.append(new_record)
    save_archive_data(archive_data)
    
    return new_record

def get_payment_status(splatnost_nazev, vytvoreni, current_time=None):
    """
    Určí stav platby na základě splatnosti, času vytvoření a aktuálního času.
    - vytvoreni: datetime vytvoření faktury
    - current_time: čas pro porovnání (None = nyní)
    Vrací: 0 = po splatnosti, 1 = uhrazeno (nepoužívá se zde), 2 = v termínu
    """
    if current_time is None:
        current_time = datetime.now()
    
    splatnosti_data = load_splatnosti()
    
    # Najít splatnost v seznamu
    splatnost_obj = None
    for s in splatnosti_data:
        if s["nazev"] == splatnost_nazev:
            splatnost_obj = s
            break
    
    if not splatnost_obj:
        return 0
    
    hodiny = splatnost_obj.get("hodiny", 0)
    jednotka = splatnost_obj.get("jednotka", "hodiny")
    
    if jednotka == "hodiny":
        if hodiny == 0:
            # do konce dne
            deadline = vytvoreni.replace(hour=23, minute=59, second=59)
        else:
            deadline = vytvoreni + timedelta(hours=hodiny)
    else:  # dny
        if hodiny == 0:
            deadline = vytvoreni.replace(hour=23, minute=59, second=59)
        else:
            # konec dne za hodiny dní
            deadline = vytvoreni.replace(hour=23, minute=59, second=59) + timedelta(days=hodiny)
    
    if current_time <= deadline:
        return 2   # v termínu (k uhrazení)
    else:
        return 0   # po splatnosti (neuhrazeno)

def update_archive_statuses():
    """Aktualizuje stavy všech faktur v archivu podle aktuálního času"""
    archive_data = load_archive_data()
    updated = False
    
    for item in archive_data:
        if item["stav"] == 1:
            continue  # uhrazené neměníme
            
        try:
            vytvoreni = datetime.strptime(item["datum_vytvoreni"], "%d.%m.%Y %H:%M:%S")
        except:
            vytvoreni = datetime.now()
            
        new_status = get_payment_status(item["splatnost"], vytvoreni)
        
        if item["stav"] != new_status:
            item["stav"] = new_status
            updated = True
    
    if updated:
        save_archive_data(archive_data)
    
    return archive_data

def remove_from_archive(cislo_faktury):
    """Odstraní fakturu z archivu"""
    archive_data = load_archive_data()
    new_data = []
    removed = False
    
    for item in archive_data:
        if item["cislo_faktury"] == cislo_faktury:
            if os.path.exists(item["archiv_path"]):
                os.remove(item["archiv_path"])
            removed = True
        else:
            new_data.append(item)
    
    if removed:
        save_archive_data(new_data)
    
    return removed

def update_payment_status(cislo_faktury, new_status):
    """Aktualizuje stav platby faktury (ruční změna)"""
    archive_data = load_archive_data()
    
    for item in archive_data:
        if item["cislo_faktury"] == cislo_faktury:
            item["stav"] = new_status
            save_archive_data(archive_data)
            return True
    
    return False

# ================= NASTAVENÍ TÉMAT =================
def setup_theme(app, theme_name):
    if theme_name == "dark":
        app.setStyle("Fusion")
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.WindowText, Qt.white)
        dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
        dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
        dark_palette.setColor(QPalette.ToolTipText, Qt.white)
        dark_palette.setColor(QPalette.Text, Qt.white)
        dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ButtonText, Qt.white)
        dark_palette.setColor(QPalette.BrightText, Qt.red)
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, Qt.black)
        app.setPalette(dark_palette)
    elif theme_name == "light":
        app.setStyle("Fusion")
        light_palette = QPalette()
        light_palette.setColor(QPalette.Window, QColor(240, 240, 240))
        light_palette.setColor(QPalette.WindowText, Qt.black)
        light_palette.setColor(QPalette.Base, Qt.white)
        light_palette.setColor(QPalette.AlternateBase, QColor(240, 240, 240))
        light_palette.setColor(QPalette.ToolTipBase, Qt.white)
        light_palette.setColor(QPalette.ToolTipText, Qt.black)
        light_palette.setColor(QPalette.Text, Qt.black)
        light_palette.setColor(QPalette.Button, QColor(240, 240, 240))
        light_palette.setColor(QPalette.ButtonText, Qt.black)
        light_palette.setColor(QPalette.BrightText, Qt.red)
        light_palette.setColor(QPalette.Link, QColor(0, 100, 200))
        light_palette.setColor(QPalette.Highlight, QColor(0, 100, 200))
        light_palette.setColor(QPalette.HighlightedText, Qt.white)
        app.setPalette(light_palette)
    else:
        app.setStyle("")

# ================= KONSTANTY =================
UCET = "20220297/5500"
KS = "1004"

# ================= ČÍSLO FAKTURY =================
def nove_cislo_faktury():
    rok = datetime.now().year
    yy = str(rok)[-2:]

    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            ulozeny_rok, poradi = f.read().split(";")
            poradi = int(poradi)
    else:
        ulozeny_rok = str(rok)
        poradi = 0

    if ulozeny_rok != str(rok):
        poradi = 0

    poradi += 1

    with open(COUNTER_FILE, "w", encoding="utf-8") as f:
        f.write(f"{rok};{poradi}")

    return f"{yy}{poradi:04d}"

# ================= STORNO =================
def storno_text(termin_str):
    dnes = datetime.now().date()
    try:
        # Zkusíme parsovat datum (může být s časem)
        if ' ' in termin_str:
            termin = datetime.strptime(termin_str, "%d.%m.%Y %H:%M").date()
        else:
            termin = datetime.strptime(termin_str, "%d.%m.%Y").date()
    except:
        termin = dnes
    rozdil = (termin - dnes).days
    lhuta = "nejpozději 2 hodiny před konáním" if rozdil <= 1 else "nejpozději 24 hodin před konáním"

    return [
        "Platba je nevratná.",
        f"1× možnost změny termínu ({lhuta}).",
        "https://hotelbeethoven.cz/obchodni-podminky/"
    ]

# ================= PDF S KONKRÉTNÍM DATEM SPLATNOSTI =================
def vytvor_pdf(data, due_date_str, vydal=None):
    """Vytvoří PDF fakturu s konkrétním datem splatnosti"""
    pdfmetrics.registerFont(TTFont("Arial", r"C:\Windows\Fonts\arial.ttf"))

    cislo_faktury = nove_cislo_faktury()
    cesta = os.path.join(settings["output_dir"], f"zalohovafaktura_{cislo_faktury}.pdf")

    c = canvas.Canvas(cesta, pagesize=A4)
    w, h = A4

    c.setFont("Arial", 16)
    c.drawString(50, h - 50, "ZÁLOHOVÁ FAKTURA")

    c.setFont("Arial", 10)
    c.drawString(50, h - 80, f"Číslo faktury: {cislo_faktury}")
    c.drawString(50, h - 95, f"Datum vystavení: {datetime.now().strftime('%d.%m.%Y')}")
    
    if vydal:
        c.drawString(50, h - 110, f"Vydal: {vydal}")

    c.line(50, h - 125, w - 50, h - 125)

    left_x, right_x = 50, 320
    y = h - 155

    dodavatel = [
        "DODAVATEL",
        settings.get("company_name", "TRIDENT GROUP 007 s.r.o."),
        settings.get("company_address", "Nám. Jiřího z Lobkovic 2406/9, 130 00 Praha 3 – Vinohrady"),
        f"IČO: {settings.get('company_id', '27262405')} | DIČ: {settings.get('company_vat', 'CZ27262405')}",
        f"Číslo účtu: {UCET}",
    ]

    provozovna = [
        "PROVOZOVNA",
        "WELLNESS HOTEL BEETHOVEN****",
        "Beethovenova 1146",
        "430 01 Chomutov",
        "recepce@hotelbeethoven.cz",
        "+420 774 775 599",
    ]

    yl = y
    for r in dodavatel:
        c.drawString(left_x, yl, r)
        yl -= 14

    yr = y
    for r in provozovna:
        c.drawString(right_x, yr, r)
        yr -= 14

    y = min(yl, yr) - 20
    c.line(50, y, w - 50, y)
    y -= 25

    c.drawString(50, y, "ODBĚRATEL")
    y -= 15
    c.drawString(70, y, data["jmeno"])
    y -= 15

    if data["email"]:
        c.drawString(70, y, f"E-mail: {data['email']}")
        y -= 15
    if data["telefon"]:
        c.drawString(70, y, f"Telefon: {data['telefon']}")

    y -= 25
    c.line(50, y, w - 50, y)
    y -= 20

    c.drawString(50, y, f"Služba: {data['sluzba']}")
    y -= 15
    c.drawString(70, y, f"Termín: {data['termin']}")
    y -= 15
    c.drawString(70, y, f"Cena služby: {data['cena']} CZK")
    y -= 15
    # Poznámka
    if data.get("poznamka"):
        c.drawString(70, y, f"Poznámka: {data['poznamka']}")
        y -= 15

    y -= 25
    c.line(50, y, w - 50, y)
    y -= 20

    c.drawString(50, y, "PLATEBNÍ ÚDAJE")
    y -= 15
    c.drawString(70, y, f"Číslo účtu: {UCET}")
    y -= 15
    c.drawString(70, y, f"Variabilní symbol: {cislo_faktury}")
    y -= 15
    c.drawString(70, y, f"Konstantní symbol: {KS}")
    y -= 15
    # Zobrazení konkrétního data splatnosti
    c.drawString(70, y, f"Splatnost: {due_date_str}")

    y -= 80
    c.line(50, y, w - 50, y)
    y -= 20

    c.drawString(50, y, "STORNO PODMÍNKY")
    y -= 15
    for r in storno_text(data["termin"]):
        c.drawString(70, y, r)
        y -= 15

    c.save()

    # Přidat do archivu (due_date_str již máme)
    add_to_archive(data, cislo_faktury, cesta, due_date_str, vydal)

    if settings["open_pdf"]:
        os.startfile(cesta)
    
    return cesta, cislo_faktury

# ================= 1. STATISTIKY A REPORTY =================
class StatisticsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Statistiky a Reporty")
        self.setMinimumSize(1000, 700)
        
        self.init_ui()
        self.load_statistics()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Period selection
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Období:"))
        
        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        period_layout.addWidget(self.date_from)
        
        period_layout.addWidget(QLabel("do"))
        
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        period_layout.addWidget(self.date_to)
        
        self.btn_refresh = QPushButton("Aktualizovat")
        self.btn_refresh.clicked.connect(self.load_statistics)
        period_layout.addWidget(self.btn_refresh)
        
        period_layout.addStretch()
        layout.addLayout(period_layout)
        
        # Tabs
        tabs = QTabWidget()
        
        # Summary tab
        summary_tab = QWidget()
        summary_layout = QVBoxLayout()
        
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(2)
        self.summary_table.setHorizontalHeaderLabels(["Ukazatel", "Hodnota"])
        summary_layout.addWidget(self.summary_table)
        
        summary_tab.setLayout(summary_layout)
        
        # Charts tab
        charts_tab = QWidget()
        charts_layout = QVBoxLayout()
        
        self.figure = plt.figure(figsize=(10, 8))
        self.canvas = FigureCanvas(self.figure)
        charts_layout.addWidget(self.canvas)
        
        charts_tab.setLayout(charts_layout)
        
        # Monthly breakdown tab
        monthly_tab = QWidget()
        monthly_layout = QVBoxLayout()
        
        self.monthly_table = QTableWidget()
        self.monthly_table.setColumnCount(4)
        self.monthly_table.setHorizontalHeaderLabels(["Měsíc", "Počet faktur", "Celková částka", "Průměrná cena"])
        monthly_layout.addWidget(self.monthly_table)
        
        monthly_tab.setLayout(monthly_layout)
        
        tabs.addTab(summary_tab, "Přehled")
        tabs.addTab(charts_tab, "Grafy")
        tabs.addTab(monthly_tab, "Měsíční rozpis")
        
        layout.addWidget(tabs)
        
        # Export buttons
        export_layout = QHBoxLayout()
        self.btn_export_csv = QPushButton("Export do CSV")
        self.btn_export_pdf = QPushButton("Export do PDF")
        self.btn_export_excel = QPushButton("Export do Excel")
        
        self.btn_export_csv.clicked.connect(self.export_csv)
        self.btn_export_pdf.clicked.connect(self.export_pdf)
        self.btn_export_excel.clicked.connect(self.export_excel)
        
        export_layout.addWidget(self.btn_export_csv)
        export_layout.addWidget(self.btn_export_pdf)
        export_layout.addWidget(self.btn_export_excel)
        export_layout.addStretch()
        
        layout.addLayout(export_layout)
        self.setLayout(layout)
    
    def load_statistics(self):
        archive_data = load_archive_data()
        date_from = self.date_from.date()
        date_to = self.date_to.date()
        
        # Filter data by date
        filtered_data = []
        for item in archive_data:
            try:
                item_date = datetime.strptime(item["datum_vytvoreni"], "%d.%m.%Y %H:%M:%S").date()
                if date_from <= QDate(item_date.year, item_date.month, item_date.day) <= date_to:
                    filtered_data.append(item)
            except:
                pass
        
        # Calculate statistics
        total_invoices = len(filtered_data)
        total_amount = sum(item["cena"] for item in filtered_data)
        paid_invoices = sum(1 for item in filtered_data if item["stav"] == 1)
        unpaid_invoices = sum(1 for item in filtered_data if item["stav"] == 0)
        pending_invoices = sum(1 for item in filtered_data if item["stav"] == 2)
        
        # Service breakdown
        service_counts = {}
        service_amounts = {}
        for item in filtered_data:
            service = item["sluzba"]
            service_counts[service] = service_counts.get(service, 0) + 1
            service_amounts[service] = service_amounts.get(service, 0) + item["cena"]
        
        # Monthly breakdown
        monthly_data = defaultdict(lambda: {"count": 0, "amount": 0})
        for item in filtered_data:
            try:
                item_date = datetime.strptime(item["datum_vytvoreni"], "%d.%m.%Y %H:%M:%S")
                month_key = f"{item_date.year}-{item_date.month:02d}"
                monthly_data[month_key]["count"] += 1
                monthly_data[month_key]["amount"] += item["cena"]
            except:
                pass
        
        # Update summary table
        self.summary_table.setRowCount(8)
        summary_data = [
            ["Celkový počet faktur", str(total_invoices)],
            ["Celková částka", f"{total_amount:,} CZK"],
            ["Uhrazené faktury", f"{paid_invoices} ({paid_invoices/total_invoices*100:.1f}%)" if total_invoices > 0 else "0"],
            ["Neuhrazené faktury", f"{unpaid_invoices} ({unpaid_invoices/total_invoices*100:.1f}%)" if total_invoices > 0 else "0"],
            ["Faktury k uhrazení", f"{pending_invoices} ({pending_invoices/total_invoices*100:.1f}%)" if total_invoices > 0 else "0"],
            ["Průměrná faktura", f"{total_amount/total_invoices:,.0f} CZK" if total_invoices > 0 else "0"],
            ["Nejčastější služba", max(service_counts, key=service_counts.get) if service_counts else "Žádná"],
            ["Nejvyšší obrat", max(service_amounts, key=service_amounts.get) if service_amounts else "Žádná"]
        ]
        
        for i, (label, value) in enumerate(summary_data):
            self.summary_table.setItem(i, 0, QTableWidgetItem(label))
            self.summary_table.setItem(i, 1, QTableWidgetItem(value))
        
        # Update monthly table
        sorted_months = sorted(monthly_data.keys())
        self.monthly_table.setRowCount(len(sorted_months))
        for i, month in enumerate(sorted_months):
            data = monthly_data[month]
            year, month_num = map(int, month.split('-'))
            month_name = QDate(year, month_num, 1).toString("MMMM yyyy")
            
            self.monthly_table.setItem(i, 0, QTableWidgetItem(month_name))
            self.monthly_table.setItem(i, 1, QTableWidgetItem(str(data["count"])))
            self.monthly_table.setItem(i, 2, QTableWidgetItem(f"{data['amount']:,} CZK"))
            self.monthly_table.setItem(i, 3, QTableWidgetItem(f"{data['amount']/data['count']:,.0f} CZK" if data["count"] > 0 else "0"))
        
        # Update charts
        self.update_charts(service_counts, service_amounts, monthly_data)
    
    def update_charts(self, service_counts, service_amounts, monthly_data):
        self.figure.clear()
        
        # Service distribution pie chart
        ax1 = self.figure.add_subplot(221)
        if service_counts:
            labels = list(service_counts.keys())
            sizes = list(service_counts.values())
            ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax1.set_title('Rozdělení podle služeb')
        
        # Monthly revenue bar chart
        ax2 = self.figure.add_subplot(222)
        if monthly_data:
            months = sorted(monthly_data.keys())
            revenues = [monthly_data[m]["amount"] for m in months]
            month_labels = [f"{m.split('-')[1]}/{m.split('-')[0][2:]}" for m in months]
            ax2.bar(month_labels, revenues)
            ax2.set_title('Měsíční obrat')
            ax2.set_ylabel('CZK')
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # Payment status
        ax3 = self.figure.add_subplot(223)
        status_labels = ['Uhrazeno', 'Neuhrazeno', 'K uhrazení']
        archive_data = load_archive_data()
        status_counts = [
            sum(1 for item in archive_data if item["stav"] == 1),
            sum(1 for item in archive_data if item["stav"] == 0),
            sum(1 for item in archive_data if item["stav"] == 2)
        ]
        colors = ['green', 'red', 'blue']
        ax3.bar(status_labels, status_counts, color=colors)
        ax3.set_title('Stav plateb')
        
        # Service revenue
        ax4 = self.figure.add_subplot(224)
        if service_amounts:
            services = list(service_amounts.keys())
            amounts = list(service_amounts.values())
            ax4.barh(services, amounts)
            ax4.set_title('Obrat podle služeb')
            ax4.set_xlabel('CZK')
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def export_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit jako CSV", "", "CSV Files (*.csv)"
        )
        if file_path:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(['Ukazatel', 'Hodnota'])
                for row in range(self.summary_table.rowCount()):
                    label = self.summary_table.item(row, 0).text()
                    value = self.summary_table.item(row, 1).text()
                    writer.writerow([label, value])
            QMessageBox.information(self, "Export", "Data byla exportována do CSV.")
    
    def export_pdf(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit jako PDF", "", "PDF Files (*.pdf)"
        )
        if file_path:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            elements.append(Paragraph("Statistika faktur", styles['Title']))
            elements.append(Spacer(1, 12))
            
            # Summary table
            data = [['Ukazatel', 'Hodnota']]
            for row in range(self.summary_table.rowCount()):
                label = self.summary_table.item(row, 0).text()
                value = self.summary_table.item(row, 1).text()
                data.append([label, value])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            
            doc.build(elements)
            QMessageBox.information(self, "Export", "Report byl exportován do PDF.")
    
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Uložit jako Excel", "", "Excel Files (*.xlsx)"
            )
            if file_path:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Statistika"
                
                # Write summary
                ws['A1'] = "Statistika faktur"
                ws['A1'].font = Font(bold=True, size=14)
                
                for row in range(self.summary_table.rowCount()):
                    ws.cell(row=row+3, column=1, value=self.summary_table.item(row, 0).text())
                    ws.cell(row=row+3, column=2, value=self.summary_table.item(row, 1).text())
                
                wb.save(file_path)
                QMessageBox.information(self, "Export", "Data byla exportována do Excel.")
        except ImportError:
            QMessageBox.warning(self, "Chyba", "Pro export do Excel nainstalujte knihovnu openpyxl.")

# ================= 2. E-MAILOVÉ NOTIFIKACE =================
class EmailSettingsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Nastavení e-mailu")
        self.setMinimumWidth(500)
        
        self.load_settings()
        self.init_ui()
    
    def load_settings(self):
        default_settings = {
            "enabled": False,
            "smtp_server": "smtp.gmail.com",
            "smtp_port": 587,
            "username": "",
            "password": "",
            "from_email": "",
            "ssl_required": True,
            "send_reminders": True,
            "reminder_days": 3,
            "invoice_template": "Vážený zákazníku,\n\npřikládáme zálohovou fakturu č. {invoice_number}.\n\nS pozdravem,\n{company_name}"
        }
        
        if os.path.exists(EMAIL_SETTINGS_FILE):
            with open(EMAIL_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                self.settings = json.load(f)
                # Merge with defaults for missing keys
                for key, value in default_settings.items():
                    if key not in self.settings:
                        self.settings[key] = value
        else:
            self.settings = default_settings
    
    def save_settings(self):
        with open(EMAIL_SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.settings, f, indent=2, ensure_ascii=False)
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Enable/disable
        self.enable_checkbox = QCheckBox("Povolit e-mailové notifikace")
        self.enable_checkbox.setChecked(self.settings["enabled"])
        layout.addWidget(self.enable_checkbox)
        
        # SMTP Settings group
        smtp_group = QGroupBox("SMTP Nastavení")
        smtp_layout = QFormLayout()
        
        self.smtp_server = QLineEdit(self.settings["smtp_server"])
        self.smtp_port = QSpinBox()
        self.smtp_port.setRange(1, 65535)
        self.smtp_port.setValue(self.settings["smtp_port"])
        self.username = QLineEdit(self.settings["username"])
        self.password = QLineEdit(self.settings["password"])
        self.password.setEchoMode(QLineEdit.Password)
        self.from_email = QLineEdit(self.settings["from_email"])
        self.ssl_checkbox = QCheckBox("Použít SSL/TLS")
        self.ssl_checkbox.setChecked(self.settings["ssl_required"])
        
        smtp_layout.addRow("SMTP Server:", self.smtp_server)
        smtp_layout.addRow("Port:", self.smtp_port)
        smtp_layout.addRow("Uživatelské jméno:", self.username)
        smtp_layout.addRow("Heslo:", self.password)
        smtp_layout.addRow("Odesílatel:", self.from_email)
        smtp_layout.addRow("", self.ssl_checkbox)
        
        smtp_group.setLayout(smtp_layout)
        layout.addWidget(smtp_group)
        
        # Test button
        test_button = QPushButton("Otestovat připojení")
        test_button.clicked.connect(self.test_connection)
        layout.addWidget(test_button)
        
        # Reminders group
        reminders_group = QGroupBox("Připomenutí")
        reminders_layout = QFormLayout()
        
        self.send_reminders = QCheckBox("Posílat připomenutí splatnosti")
        self.send_reminders.setChecked(self.settings["send_reminders"])
        self.reminder_days = QSpinBox()
        self.reminder_days.setRange(1, 30)
        self.reminder_days.setValue(self.settings["reminder_days"])
        
        reminders_layout.addRow("", self.send_reminders)
        reminders_layout.addRow("Dny před splatností:", self.reminder_days)
        
        reminders_group.setLayout(reminders_layout)
        layout.addWidget(reminders_group)
        
        # Template
        template_group = QGroupBox("Šablona e-mailu")
        template_layout = QVBoxLayout()
        
        self.template_edit = QTextEdit()
        self.template_edit.setPlainText(self.settings["invoice_template"])
        self.template_edit.setMaximumHeight(150)
        
        template_layout.addWidget(QLabel("Dostupné proměnné: {invoice_number}, {customer_name}, {amount}, {due_date}, {company_name}"))
        template_layout.addWidget(self.template_edit)
        
        template_group.setLayout(template_layout)
        layout.addWidget(template_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_and_close)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def test_connection(self):
        try:
            server = smtplib.SMTP(self.smtp_server.text(), self.smtp_port.value())
            if self.ssl_checkbox.isChecked():
                server.starttls()
            server.login(self.username.text(), self.password.text())
            server.quit()
            QMessageBox.information(self, "Test připojení", "Připojení k SMTP serveru bylo úspěšné!")
        except Exception as e:
            QMessageBox.critical(self, "Chyba připojení", f"Nelze se připojit k SMTP serveru:\n{str(e)}")
    
    def save_and_close(self):
        self.settings.update({
            "enabled": self.enable_checkbox.isChecked(),
            "smtp_server": self.smtp_server.text(),
            "smtp_port": self.smtp_port.value(),
            "username": self.username.text(),
            "password": self.password.text(),
            "from_email": self.from_email.text(),
            "ssl_required": self.ssl_checkbox.isChecked(),
            "send_reminders": self.send_reminders.isChecked(),
            "reminder_days": self.reminder_days.value(),
            "invoice_template": self.template_edit.toPlainText()
        })
        self.save_settings()
        self.accept()

def send_invoice_email(customer_email, invoice_path, invoice_data):
    """Odešle fakturu e-mailem a pošle kopii odesílateli"""
    try:
        # Load email settings
        if not os.path.exists(EMAIL_SETTINGS_FILE):
            return False, "E-mailové nastavení není nakonfigurováno"
        
        with open(EMAIL_SETTINGS_FILE, 'r', encoding='utf-8') as f:
            settings = json.load(f)
        
        if not settings.get("enabled", False):
            return False, "E-mailové notifikace jsou vypnuty"
        
        from_email = settings.get('from_email', '')
        if not from_email:
            return False, "Není nastavena adresa odesílatele"
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = customer_email
        msg['Cc'] = from_email  # kopie odesílateli
        msg['Subject'] = f"Zálohová faktura {invoice_data['cislo_faktury']}"
        
        # Create email body from template
        template = settings.get('invoice_template', 'Vážený zákazníku,\n\npřikládáme fakturu č. {invoice_number}.\n\nS pozdravem,\n{company_name}')
        body = template.format(
            invoice_number=invoice_data['cislo_faktury'],
            customer_name=invoice_data['jmeno'],
            amount=invoice_data['cena'],
            due_date=invoice_data['due_date'],  # due_date již máme
            company_name=settings.get('company_name', 'Wellness Hotel Beethoven')
        )
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Attach PDF
        with open(invoice_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(invoice_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(invoice_path)}"'
            msg.attach(part)
        
        # Odeslat všem (To a Cc)
        recipients = [customer_email, from_email]
        
        server = smtplib.SMTP(settings['smtp_server'], settings['smtp_port'])
        if settings.get('ssl_required', True):
            server.starttls()
        server.login(settings['username'], settings['password'])
        server.send_message(msg, from_email, recipients)
        server.quit()
        
        return True, "E-mail byl úspěšně odeslán (včetně kopie odesílateli)"
    except Exception as e:
        return False, f"Chyba při odesílání e-mailu: {str(e)}"

# ================= 3. DAŇOVÉ PŘIZNÁNÍ =================
class TaxReportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Daňové reporty")
        self.setMinimumSize(800, 600)
        
        self.init_ui()
        self.load_tax_data()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Period selection
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Období:"))
        
        self.year_combo = QComboBox()
        current_year = datetime.now().year
        for year in range(current_year - 5, current_year + 1):
            self.year_combo.addItem(str(year))
        self.year_combo.setCurrentText(str(current_year))
        
        self.quarter_combo = QComboBox()
        self.quarter_combo.addItems(["Celý rok", "Q1", "Q2", "Q3", "Q4"])
        
        period_layout.addWidget(self.year_combo)
        period_layout.addWidget(self.quarter_combo)
        
        self.btn_generate = QPushButton("Generovat report")
        self.btn_generate.clicked.connect(self.generate_report)
        period_layout.addWidget(self.btn_generate)
        
        period_layout.addStretch()
        layout.addLayout(period_layout)
        
        # Report display
        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        layout.addWidget(self.report_text)
        
        # Export buttons
        export_layout = QHBoxLayout()
        self.btn_export_tax = QPushButton("Export pro daňové přiznání")
        self.btn_export_vat = QPushButton("Export pro DPH")
        self.btn_print = QPushButton("Tisk")
        
        self.btn_export_tax.clicked.connect(self.export_tax_report)
        self.btn_export_vat.clicked.connect(self.export_vat_report)
        self.btn_print.clicked.connect(self.print_report)
        
        export_layout.addWidget(self.btn_export_tax)
        export_layout.addWidget(self.btn_export_vat)
        export_layout.addWidget(self.btn_print)
        export_layout.addStretch()
        
        layout.addLayout(export_layout)
        self.setLayout(layout)
    
    def load_tax_data(self):
        if not os.path.exists(TAX_REPORT_FILE):
            self.tax_data = []
        else:
            with open(TAX_REPORT_FILE, 'r', encoding='utf-8') as f:
                self.tax_data = json.load(f)
    
    def generate_report(self):
        year = int(self.year_combo.currentText())
        quarter = self.quarter_combo.currentText()
        
        archive_data = load_archive_data()
        
        # Filter by year and quarter
        filtered_data = []
        for item in archive_data:
            try:
                item_date = datetime.strptime(item["datum_vytvoreni"], "%d.%m.%Y %H:%M:%S")
                if item_date.year != year:
                    continue
                
                if quarter != "Celý rok":
                    quarter_num = int(quarter[1])
                    item_quarter = (item_date.month - 1) // 3 + 1
                    if item_quarter != quarter_num:
                        continue
                
                filtered_data.append(item)
            except:
                continue
        
        # Calculate tax report
        total_revenue = sum(item["cena"] for item in filtered_data)
        tax_rate = settings.get("tax_rate", 21)
        tax_amount = total_revenue * tax_rate / 100
        net_revenue = total_revenue - tax_amount
        
        # Service breakdown for tax purposes
        service_revenue = {}
        for item in filtered_data:
            service = item["sluzba"]
            service_revenue[service] = service_revenue.get(service, 0) + item["cena"]
        
        # Generate report text
        report = f"DAŇOVÝ REPORT\n"
        report += f"===============\n\n"
        report += f"Období: {year} {quarter}\n"
        report += f"Datum generování: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n\n"
        
        report += f"PŘEHLED TRŽEB\n"
        report += f"--------------\n"
        report += f"Celkové tržby: {total_revenue:,.0f} {settings.get('currency', 'CZK')}\n"
        report += f"Sazba DPH: {tax_rate}%\n"
        report += f"Částka DPH: {tax_amount:,.0f} {settings.get('currency', 'CZK')}\n"
        report += f"Čisté tržby (bez DPH): {net_revenue:,.0f} {settings.get('currency', 'CZK')}\n\n"
        
        report += f"ROZPIS PODLE SLUŽEB\n"
        report += f"-------------------\n"
        for service, revenue in service_revenue.items():
            service_tax = revenue * tax_rate / 100
            service_net = revenue - service_tax
            report += f"{service}:\n"
            report += f"  Tržby s DPH: {revenue:,.0f} {settings.get('currency', 'CZK')}\n"
            report += f"  DPH: {service_tax:,.0f} {settings.get('currency', 'CZK')}\n"
            report += f"  Čisté tržby: {service_net:,.0f} {settings.get('currency', 'CZK')}\n\n"
        
        report += f"STATISTIKA\n"
        report += f"-----------\n"
        report += f"Počet faktur: {len(filtered_data)}\n"
        
        # Payment status statistics
        paid_count = sum(1 for item in filtered_data if item["stav"] == 1)
        unpaid_count = sum(1 for item in filtered_data if item["stav"] == 0)
        pending_count = sum(1 for item in filtered_data if item["stav"] == 2)
        
        report += f"Uhrazené faktury: {paid_count}\n"
        report += f"Neuhrazené faktury: {unpaid_count}\n"
        report += f"Faktury k uhrazení: {pending_count}\n"
        
        self.report_text.setPlainText(report)
    
    def export_tax_report(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit daňový report", "", "CSV Files (*.csv)"
        )
        if file_path:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(['Daňový report', self.year_combo.currentText(), self.quarter_combo.currentText()])
                writer.writerow([])
                writer.writerow(['Položka', 'Hodnota', 'Měna'])
                
                lines = self.report_text.toPlainText().split('\n')
                for line in lines:
                    if ':' in line:
                        parts = line.split(':', 1)
                        writer.writerow([parts[0].strip(), parts[1].strip(), settings.get('currency', 'CZK')])
            
            QMessageBox.information(self, "Export", "Daňový report byl exportován.")
    
    def export_vat_report(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Uložit report DPH", "", "PDF Files (*.pdf)"
        )
        if file_path:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            elements.append(Paragraph("REPORT PRO DPH", styles['Title']))
            elements.append(Spacer(1, 12))
            
            # Add report content
            report_text = self.report_text.toPlainText()
            for line in report_text.split('\n'):
                if line.strip():
                    elements.append(Paragraph(line, styles['Normal']))
            
            doc.build(elements)
            QMessageBox.information(self, "Export", "Report DPH byl exportován do PDF.")
    
    def print_report(self):
        printer = QPrinter()
        print_dialog = QPrintDialog(printer, self)
        if print_dialog.exec() == QPrintDialog.Accepted:
            document = QTextDocument()
            document.setPlainText(self.report_text.toPlainText())
            document.print_(printer)

# ================= 4. MULTI-USER PODPORA (UPRAVENÉ ROLE) =================
class UserManager:
    def __init__(self):
        self.users_file = USERS_FILE
        self.current_user = None
        self.load_users()
    
    def load_users(self):
        if not os.path.exists(self.users_file):
            # Admin účet s hard coded heslem
            self.users = {
                "admin": {
                    "password": hashlib.sha256("061004".encode()).hexdigest(),
                    "role": "admin",
                    "permissions": ["all"],
                    "created": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    "last_login": None,
                    "cannot_delete": True
                },
                "recepce": {
                    "password": hashlib.sha256("recepce123".encode()).hexdigest(),
                    "role": "recepční",
                    "permissions": ["view", "create", "mark_paid", "reports", "backup", "services", "recurring"],
                    "created": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    "last_login": None,
                    "cannot_delete": False
                },
                "ucetni": {
                    "password": hashlib.sha256("ucetni123".encode()).hexdigest(),
                    "role": "účetní",
                    "permissions": ["view", "create", "mark_paid", "reports", "export", "settings"],
                    "created": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    "last_login": None,
                    "cannot_delete": False
                }
            }
            self.save_users()
        else:
            with open(self.users_file, 'r', encoding='utf-8') as f:
                self.users = json.load(f)
    
    def save_users(self):
        with open(self.users_file, 'w', encoding='utf-8') as f:
            json.dump(self.users, f, indent=2, ensure_ascii=False)
    
    def authenticate(self, username, password):
        if username in self.users:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            if self.users[username]["password"] == hashed_password:
                self.current_user = username
                self.users[username]["last_login"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                self.save_users()
                
                # Uložit poslední přihlášení (kromě admin)
                save_last_login(username, password)
                
                return True
        return False
    
    def has_permission(self, permission):
        if not self.current_user:
            return False
        user_data = self.users.get(self.current_user, {})
        if user_data.get("role") == "admin":
            return True
        return permission in user_data.get("permissions", [])
    
    def add_user(self, username, password, role, permissions):
        if username in self.users:
            return False
        self.users[username] = {
            "password": hashlib.sha256(password.encode()).hexdigest(),
            "role": role,
            "permissions": permissions,
            "created": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            "last_login": None,
            "cannot_delete": False
        }
        self.save_users()
        return True
    
    def delete_user(self, username):
        if username in self.users and not self.users[username].get("cannot_delete", False):
            del self.users[username]
            self.save_users()
            return True
        return False

class LoginDialog(QDialog):
    def __init__(self, user_manager):
        super().__init__()
        self.user_manager = user_manager
        self.setWindowTitle("Přihlášení - HEM Zálohové faktury")
        self.setFixedWidth(400)
        self.setFixedHeight(220)
        
        self.init_ui()
        
        # Načíst poslední přihlášení
        self.load_last_login()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Logo/header
        header_label = QLabel("HEM - Zálohové faktury")
        header_font = QFont()
        header_font.setPointSize(14)
        header_font.setBold(True)
        header_label.setFont(header_font)
        header_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(header_label)
        
        # Version
        version_label = QLabel("Verze 2.0.1")
        version_label.setAlignment(Qt.AlignCenter)
        version_label.setStyleSheet("color: gray;")
        layout.addWidget(version_label)
        
        # Username
        layout.addWidget(QLabel("Uživatelské jméno:"))
        self.username_edit = QLineEdit()
        layout.addWidget(self.username_edit)
        
        # Password
        layout.addWidget(QLabel("Heslo:"))
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.password_edit.returnPressed.connect(self.login)
        layout.addWidget(self.password_edit)
        
        # Tlačítko pro smazání uložených údajů
        clear_btn = QPushButton("Smazat uložené přihlašovací údaje")
        clear_btn.clicked.connect(self.clear_saved_login)
        clear_btn.setStyleSheet("font-size: 10px;")
        layout.addWidget(clear_btn)
        
        # Buttons
        button_layout = QHBoxLayout()
        login_button = QPushButton("Přihlásit")
        cancel_button = QPushButton("Zrušit")
        
        login_button.setDefault(True)
        login_button.setAutoDefault(True)
        login_button.clicked.connect(self.login)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(login_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        login_button.setFocus()
    
    def load_last_login(self):
        """Načte poslední uložené přihlášení"""
        last_login = load_last_login()
        if last_login["username"]:
            self.username_edit.setText(last_login["username"])
            self.password_edit.setText(last_login["password"])
    
    def clear_saved_login(self):
        """Smaže uložené přihlašovací údaje"""
        clear_last_login()
        self.username_edit.clear()
        self.password_edit.clear()
        QMessageBox.information(self, "Hotovo", "Uložené přihlašovací údaje byly smazány.")
        self.username_edit.setFocus()
    
    def login(self):
        username = self.username_edit.text()
        password = self.password_edit.text()
        
        if self.user_manager.authenticate(username, password):
            self.accept()
        else:
            QMessageBox.warning(self, "Chyba", "Neplatné přihlašovací údaje")
            self.password_edit.clear()
            self.password_edit.setFocus()

class UserManagementDialog(QDialog):
    def __init__(self, user_manager):
        super().__init__()
        self.user_manager = user_manager
        self.setWindowTitle("Správa uživatelů")
        self.setMinimumSize(500, 400)
        
        self.init_ui()
        self.load_users()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Users list
        self.users_list = QListWidget()
        layout.addWidget(self.users_list)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.btn_add = QPushButton("Přidat uživatele")
        self.btn_edit = QPushButton("Upravit")
        self.btn_delete = QPushButton("Smazat")
        
        self.btn_add.clicked.connect(self.add_user)
        self.btn_edit.clicked.connect(self.edit_user)
        self.btn_delete.clicked.connect(self.delete_user)
        
        button_layout.addWidget(self.btn_add)
        button_layout.addWidget(self.btn_edit)
        button_layout.addWidget(self.btn_delete)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def load_users(self):
        self.users_list.clear()
        for username, data in self.user_manager.users.items():
            cannot_delete = data.get("cannot_delete", False)
            role = data['role']
            last_login = data.get('last_login', 'Nikdy')
            
            item_text = f"{username} ({role})"
            if cannot_delete:
                item_text += " [chráněný]"
            
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, username)
            
            if cannot_delete:
                item.setForeground(QColor("blue"))
            
            self.users_list.addItem(item)
    
    def add_user(self):
        dialog = UserEditDialog(self.user_manager, None)
        if dialog.exec():
            self.load_users()
    
    def edit_user(self):
        selected = self.users_list.currentItem()
        if selected:
            username = selected.data(Qt.UserRole)
            if self.user_manager.users[username].get("cannot_delete", False):
                QMessageBox.warning(self, "Chyba", "Tento účet nelze upravovat")
                return
            
            dialog = UserEditDialog(self.user_manager, username)
            if dialog.exec():
                self.load_users()
    
    def delete_user(self):
        selected = self.users_list.currentItem()
        if selected:
            username = selected.data(Qt.UserRole)
            
            if self.user_manager.users[username].get("cannot_delete", False):
                QMessageBox.warning(self, "Chyba", "Tento účet nelze smazat")
                return
            
            if username == self.user_manager.current_user:
                QMessageBox.warning(self, "Chyba", "Nemůžete smazat svůj vlastní účet")
                return
            
            reply = QMessageBox.question(self, "Smazat uživatele", 
                                       f"Opravdu chcete smazat uživatele {username}?",
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                if self.user_manager.delete_user(username):
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Chyba", "Nelze smazat uživatele")

class UserEditDialog(QDialog):
    def __init__(self, user_manager, username=None):
        super().__init__()
        self.user_manager = user_manager
        self.username = username
        self.is_edit = username is not None
        
        self.setWindowTitle("Přidat uživatele" if not self.is_edit else "Upravit uživatele")
        self.setFixedWidth(400)
        
        self.init_ui()
        if self.is_edit:
            self.load_user_data()
    
    def init_ui(self):
        layout = QFormLayout()
        
        self.username_edit = QLineEdit()
        if self.is_edit:
            self.username_edit.setText(self.username)
            self.username_edit.setReadOnly(True)
        layout.addRow("Uživatelské jméno:", self.username_edit)
        
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        if not self.is_edit:
            layout.addRow("Heslo:", self.password_edit)
        else:
            layout.addRow("Nové heslo (ponechte prázdné):", self.password_edit)
        
        # Pouze 3 role podle požadavku
        self.role_combo = QComboBox()
        self.role_combo.addItems(["admin", "recepční", "účetní"])
        layout.addRow("Role:", self.role_combo)
        
        # Permissions podle rolí
        self.permissions_group = QGroupBox("Oprávnění (automaticky podle role)")
        permissions_layout = QVBoxLayout()
        
        self.permissions_label = QLabel("Oprávnění se automaticky přiřadí podle zvolené role.")
        permissions_layout.addWidget(self.permissions_label)
        
        # Info o oprávněních podle rolí
        info_text = """
        Admin: Všechna oprávnění
        Recepční: Zásobník, Nastavení, Statistiky, Daňové reporty, Zálohy, Služby, Pravidelné
        Účetní: Zásobník, Nastavení, Statistiky, Daňové reporty, Export
        """
        self.permissions_info = QLabel(info_text)
        self.permissions_info.setStyleSheet("color: gray; font-size: 10pt;")
        permissions_layout.addWidget(self.permissions_info)
        
        self.permissions_group.setLayout(permissions_layout)
        layout.addRow(self.permissions_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_user)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)
        
        self.setLayout(layout)
    
    def load_user_data(self):
        if self.username in self.user_manager.users:
            user_data = self.user_manager.users[self.username]
            self.role_combo.setCurrentText(user_data["role"])
    
    def save_user(self):
        username = self.username_edit.text()
        if not username:
            QMessageBox.warning(self, "Chyba", "Zadejte uživatelské jméno")
            return
        
        password = self.password_edit.text() if not self.is_edit else None
        if not self.is_edit and not password:
            QMessageBox.warning(self, "Chyba", "Zadejte heslo")
            return
        
        role = self.role_combo.currentText()
        
        # Automatické přiřazení oprávnění podle role
        if role == "admin":
            permissions = ["all"]
        elif role == "recepční":
            permissions = ["view", "create", "mark_paid", "reports", "backup", "services", "recurring", "settings"]
        elif role == "účetní":
            permissions = ["view", "create", "mark_paid", "reports", "export", "settings"]
        else:
            permissions = []
        
        if self.is_edit:
            # Update existing user
            if password:
                self.user_manager.users[username]["password"] = hashlib.sha256(password.encode()).hexdigest()
            self.user_manager.users[username]["role"] = role
            self.user_manager.users[username]["permissions"] = permissions
        else:
            # Add new user
            if not self.user_manager.add_user(username, password, role, permissions):
                QMessageBox.warning(self, "Chyba", "Uživatel již existuje")
                return
        
        self.user_manager.save_users()
        self.accept()

# ================= 5. AUTOMATICKÉ ZÁLOHOVÁNÍ =================
class BackupManager:
    def __init__(self):
        self.backup_dir = BACKUP_DIR
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def create_backup(self, target_dir=None):
        """Vytvoří kompletní zálohu aplikace do zadané složky"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if target_dir:
            backup_dir = target_dir
        else:
            # Výchozí cesta: Plocha/HEM_zaloha/datum_cas
            desktop = desktop_path()
            base_backup_dir = os.path.join(desktop, "HEM_zaloha")
            os.makedirs(base_backup_dir, exist_ok=True)
            backup_dir = os.path.join(base_backup_dir, timestamp)
        
        os.makedirs(backup_dir, exist_ok=True)
        
        files_to_backup = [
            SETTINGS_FILE,
            COUNTER_FILE,
            ARCHIVE_DATA_FILE,
            EMAIL_SETTINGS_FILE,
            CASH_REGISTER_FILE,
            TAX_REPORT_FILE,
            USERS_FILE,
            SERVICES_FILE,
            SPLATNOSTI_FILE,
            ARCHIVE_DIR
        ]
        
        try:
            for file_path in files_to_backup:
                if os.path.exists(file_path):
                    if os.path.isdir(file_path):
                        # Zálohovat celou složku
                        dir_name = os.path.basename(file_path)
                        dest_dir = os.path.join(backup_dir, dir_name)
                        shutil.copytree(file_path, dest_dir, dirs_exist_ok=True)
                    else:
                        # Zálohovat soubor
                        dest_file = os.path.join(backup_dir, os.path.basename(file_path))
                        shutil.copy2(file_path, dest_file)
            
            # Také zálohovat PDF faktury z output_dir pokud jsou jinde než v ARCHIVE_DIR
            output_dir = settings.get("output_dir", BASE_DIR)
            if output_dir != ARCHIVE_DIR and os.path.exists(output_dir):
                for filename in os.listdir(output_dir):
                    if filename.startswith("zalohovafaktura_"):
                        src = os.path.join(output_dir, filename)
                        dest = os.path.join(backup_dir, "faktury", filename)
                        os.makedirs(os.path.dirname(dest), exist_ok=True)
                        shutil.copy2(src, dest)
            
            return True, f"Záloha vytvořena: {backup_dir}"
        except Exception as e:
            return False, f"Chyba při vytváření zálohy: {str(e)}"
    
    def restore_backup(self, backup_dir):
        """Obnoví data ze zálohy"""
        try:
            # Restore files
            for item in os.listdir(backup_dir):
                source = os.path.join(backup_dir, item)
                
                if item == "archiv_faktur":
                    destination = ARCHIVE_DIR
                elif item == "faktury":
                    # Obnovit faktury do output_dir
                    output_dir = settings.get("output_dir", BASE_DIR)
                    for faktura_file in os.listdir(source):
                        shutil.copy2(os.path.join(source, faktura_file), 
                                   os.path.join(output_dir, faktura_file))
                    continue
                elif item.endswith('.json') or item.endswith('.txt'):
                    destination = os.path.join(APP_DIR, item)
                else:
                    continue
                
                if os.path.isdir(source):
                    if os.path.exists(destination):
                        shutil.rmtree(destination)
                    shutil.copytree(source, destination)
                else:
                    shutil.copy2(source, destination)
            
            return True, "Záloha úspěšně obnovena"
        except Exception as e:
            return False, f"Chyba při obnově zálohy: {str(e)}"
    
    def cleanup_old_backups(self, keep_last=10):
        """Smaže staré zálohy, zachová posledních 'keep_last'"""
        desktop = desktop_path()
        base_backup_dir = os.path.join(desktop, "HEM_zaloha")
        
        if not os.path.exists(base_backup_dir):
            return
        
        backups = []
        for folder in os.listdir(base_backup_dir):
            folder_path = os.path.join(base_backup_dir, folder)
            if os.path.isdir(folder_path):
                backups.append((folder_path, os.path.getmtime(folder_path)))
        
        backups.sort(key=lambda x: x[1], reverse=True)
        
        for folder_path, _ in backups[keep_last:]:
            try:
                shutil.rmtree(folder_path)
            except:
                pass
    
    def get_backup_list(self):
        """Vrátí seznam dostupných záloh"""
        desktop = desktop_path()
        base_backup_dir = os.path.join(desktop, "HEM_zaloha")
        
        if not os.path.exists(base_backup_dir):
            return []
        
        backups = []
        for folder in os.listdir(base_backup_dir):
            folder_path = os.path.join(base_backup_dir, folder)
            if os.path.isdir(folder_path):
                try:
                    # Pokusit se parsovat datum z názvu složky
                    dt = datetime.strptime(folder, "%Y%m%d_%H%M%S")
                    date_str = dt.strftime("%d.%m.%Y %H:%M:%S")
                except:
                    date_str = folder
                
                # Spočítat velikost
                total_size = 0
                for dirpath, dirnames, filenames in os.walk(folder_path):
                    for f in filenames:
                        fp = os.path.join(dirpath, f)
                        total_size += os.path.getsize(fp)
                
                backups.append({
                    'name': folder,
                    'path': folder_path,
                    'date': date_str,
                    'size': total_size
                })
        
        backups.sort(key=lambda x: x['name'], reverse=True)
        return backups

class BackupDialog(QDialog):
    def __init__(self, backup_manager):
        super().__init__()
        self.backup_manager = backup_manager
        self.setWindowTitle("Správa záloh")
        self.setMinimumSize(600, 400)
        
        self.init_ui()
        self.load_backups()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Backup list
        self.backup_table = QTableWidget()
        self.backup_table.setColumnCount(4)
        self.backup_table.setHorizontalHeaderLabels(["Název", "Datum", "Velikost", "Akce"])
        self.backup_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.backup_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.btn_create = QPushButton("Vytvořit zálohu")
        self.btn_restore = QPushButton("Obnovit vybranou")
        self.btn_delete = QPushButton("Smazat vybranou")
        self.btn_choose_dir = QPushButton("Vybrat cílovou složku")
        
        self.btn_create.clicked.connect(self.create_backup)
        self.btn_restore.clicked.connect(self.restore_backup)
        self.btn_delete.clicked.connect(self.delete_backup)
        self.btn_choose_dir.clicked.connect(self.choose_backup_dir)
        
        button_layout.addWidget(self.btn_create)
        button_layout.addWidget(self.btn_restore)
        button_layout.addWidget(self.btn_delete)
        button_layout.addWidget(self.btn_choose_dir)
        
        layout.addLayout(button_layout)
        
        # Cílová složka info
        self.target_dir_label = QLabel("Cílová složka: Plocha/HEM_zaloha")
        layout.addWidget(self.target_dir_label)
        
        self.target_dir = None
        
        self.setLayout(layout)
    
    def load_backups(self):
        backups = self.backup_manager.get_backup_list()
        self.backup_table.setRowCount(len(backups))
        
        for row, backup in enumerate(backups):
            self.backup_table.setItem(row, 0, QTableWidgetItem(backup['name']))
            self.backup_table.setItem(row, 1, QTableWidgetItem(backup['date']))
            self.backup_table.setItem(row, 2, QTableWidgetItem(f"{backup['size'] / 1024 / 1024:.2f} MB"))
            
            # Action button
            restore_button = QPushButton("Obnovit")
            restore_button.clicked.connect(lambda checked, b=backup: self.restore_selected_backup(b))
            self.backup_table.setCellWidget(row, 3, restore_button)
    
    def choose_backup_dir(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Vyberte cílovou složku pro zálohu", desktop_path())
        if dir_path:
            self.target_dir = dir_path
            self.target_dir_label.setText(f"Cílová složka: {dir_path}")
    
    def create_backup(self):
        success, message = self.backup_manager.create_backup(self.target_dir)
        if success:
            QMessageBox.information(self, "Záloha", message)
            self.load_backups()
        else:
            QMessageBox.critical(self, "Chyba", message)
    
    def restore_backup(self):
        selected = self.backup_table.currentRow()
        if selected >= 0:
            backup_name = self.backup_table.item(selected, 0).text()
            
            # Najít cestu k záloze
            desktop = desktop_path()
            base_backup_dir = os.path.join(desktop, "HEM_zaloha")
            backup_path = os.path.join(base_backup_dir, backup_name)
            
            if os.path.exists(backup_path):
                self.restore_selected_backup({'path': backup_path})
            else:
                QMessageBox.warning(self, "Chyba", "Záloha nebyla nalezena")
    
    def restore_selected_backup(self, backup):
        reply = QMessageBox.question(
            self, "Obnovit zálohu",
            f"Opravdu chcete obnovit zálohu {backup['name'] if 'name' in backup else backup['path']}?\n"
            "Aktuální data budou přepsána.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success, message = self.backup_manager.restore_backup(backup['path'])
            if success:
                QMessageBox.information(self, "Obnovení", message)
                self.accept()  # Close dialog, app should restart
            else:
                QMessageBox.critical(self, "Chyba", message)
    
    def delete_backup(self):
        selected = self.backup_table.currentRow()
        if selected >= 0:
            backup_name = self.backup_table.item(selected, 0).text()
            
            desktop = desktop_path()
            base_backup_dir = os.path.join(desktop, "HEM_zaloha")
            backup_path = os.path.join(base_backup_dir, backup_name)
            
            if not os.path.exists(backup_path):
                QMessageBox.warning(self, "Chyba", "Záloha nebyla nalezena")
                return
            
            reply = QMessageBox.question(
                self, "Smazat zálohu",
                f"Opravdu chcete smazat zálohu {backup_name}?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                try:
                    shutil.rmtree(backup_path)
                    self.load_backups()
                except Exception as e:
                    QMessageBox.critical(self, "Chyba", f"Nelze smazat zálohu: {str(e)}")

class AutoBackupSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Automatické zálohování")
        self.setFixedWidth(400)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Enable auto backup
        self.enable_checkbox = QCheckBox("Povolit automatické zálohování")
        self.enable_checkbox.setChecked(settings.get("auto_backup", True))
        layout.addWidget(self.enable_checkbox)
        
        # Interval
        interval_layout = QHBoxLayout()
        interval_layout.addWidget(QLabel("Interval zálohování:"))
        
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 30)
        self.interval_spin.setValue(settings.get("backup_interval_days", 7))
        self.interval_spin.setSuffix(" dní")
        
        interval_layout.addWidget(self.interval_spin)
        interval_layout.addStretch()
        layout.addLayout(interval_layout)
        
        # Last backup info
        self.last_backup_label = QLabel("Poslední záloha: N/A")
        layout.addWidget(self.last_backup_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_settings)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        self.load_last_backup_info()
    
    def load_last_backup_info(self):
        backup_manager = BackupManager()
        backups = backup_manager.get_backup_list()
        if backups:
            self.last_backup_label.setText(f"Poslední záloha: {backups[0]['date']}")
    
    def save_settings(self):
        settings["auto_backup"] = self.enable_checkbox.isChecked()
        settings["backup_interval_days"] = self.interval_spin.value()
        save_settings(settings)
        self.accept()

# ================= 6. PLÁNOVAČ PRAVIDELNÝCH FAKTUR =================
class RecurringInvoiceDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Pravidelné faktury")
        self.setMinimumSize(600, 400)
        
        self.init_ui()
        self.load_recurring_invoices()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # List of recurring invoices
        self.invoice_table = QTableWidget()
        self.invoice_table.setColumnCount(6)
        self.invoice_table.setHorizontalHeaderLabels(["Zákazník", "Služba", "Frekvence", "Další generování", "Aktivní", "Akce"])
        self.invoice_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.invoice_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.btn_add = QPushButton("Přidat pravidelnou fakturu")
        self.btn_edit = QPushButton("Upravit")
        self.btn_delete = QPushButton("Smazat")
        self.btn_generate = QPushButton("Generovat nyní")
        
        self.btn_add.clicked.connect(self.add_recurring_invoice)
        self.btn_edit.clicked.connect(self.edit_recurring_invoice)
        self.btn_delete.clicked.connect(self.delete_recurring_invoice)
        self.btn_generate.clicked.connect(self.generate_now)
        
        button_layout.addWidget(self.btn_add)
        button_layout.addWidget(self.btn_edit)
        button_layout.addWidget(self.btn_delete)
        button_layout.addWidget(self.btn_generate)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def load_recurring_invoices(self):
        # This would load from a file, using empty list for now
        self.recurring_invoices = []
        self.invoice_table.setRowCount(0)
    
    def add_recurring_invoice(self):
        dialog = RecurringInvoiceEditDialog(self)
        if dialog.exec():
            # Add new recurring invoice
            self.load_recurring_invoices()
    
    def edit_recurring_invoice(self):
        selected = self.invoice_table.currentRow()
        if selected >= 0:
            dialog = RecurringInvoiceEditDialog(self, self.recurring_invoices[selected])
            if dialog.exec():
                self.load_recurring_invoices()
    
    def delete_recurring_invoice(self):
        selected = self.invoice_table.currentRow()
        if selected >= 0:
            reply = QMessageBox.question(
                self, "Smazat pravidelnou fakturu",
                "Opravdu chcete smazat tuto pravidelnou fakturu?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                del self.recurring_invoices[selected]
                self.load_recurring_invoices()
    
    def generate_now(self):
        selected = self.invoice_table.currentRow()
        if selected >= 0:
            invoice = self.recurring_invoices[selected]
            QMessageBox.information(self, "Generování", "Faktura byla vygenerována")

class RecurringInvoiceEditDialog(QDialog):
    def __init__(self, parent=None, invoice_data=None):
        super().__init__(parent)
        self.invoice_data = invoice_data or {}
        self.is_edit = bool(invoice_data)
        
        self.setWindowTitle("Upravit pravidelnou fakturu" if self.is_edit else "Nová pravidelná faktura")
        self.setFixedWidth(500)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QFormLayout()
        
        # Customer
        self.customer_edit = QLineEdit(self.invoice_data.get("customer", ""))
        layout.addRow("Zákazník:", self.customer_edit)
        
        # Service
        self.service_combo = QComboBox()
        # Načíst služby
        all_services = get_all_active_services()
        for service in all_services:
            self.service_combo.addItem(f"{service['nazev']} ({service['cena']} Kč)")
        
        if "service" in self.invoice_data:
            self.service_combo.setCurrentText(self.invoice_data["service"])
        layout.addRow("Služba:", self.service_combo)
        
        # Frequency
        self.frequency_combo = QComboBox()
        self.frequency_combo.addItems(["Měsíčně", "Čtvrtletně", "Pololetně", "Ročně"])
        if "frequency" in self.invoice_data:
            self.frequency_combo.setCurrentText(self.invoice_data["frequency"])
        layout.addRow("Frekvence:", self.frequency_combo)
        
        # Start date
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate())
        if "start_date" in self.invoice_data:
            self.start_date.setDate(QDate.fromString(self.invoice_data["start_date"], "dd.MM.yyyy"))
        layout.addRow("Datum začátku:", self.start_date)
        
        # End date (optional)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate().addYears(1))
        self.end_date.setCalendarPopup(True)
        layout.addRow("Datum konce (nepovinné):", self.end_date)
        
        # Amount
        self.amount_spin = QDoubleSpinBox()
        self.amount_spin.setRange(0, 100000)
        self.amount_spin.setValue(self.invoice_data.get("amount", 0))
        self.amount_spin.setPrefix("CZK ")
        layout.addRow("Částka:", self.amount_spin)
        
        # Active
        self.active_checkbox = QCheckBox("Aktivní")
        self.active_checkbox.setChecked(self.invoice_data.get("active", True))
        layout.addRow("", self.active_checkbox)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_invoice)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)
        
        self.setLayout(layout)
    
    def save_invoice(self):
        if not self.customer_edit.text():
            QMessageBox.warning(self, "Chyba", "Zadejte jméno zákazníka")
            return
        
        self.invoice_data = {
            "customer": self.customer_edit.text(),
            "service": self.service_combo.currentText(),
            "frequency": self.frequency_combo.currentText(),
            "start_date": self.start_date.date().toString("dd.MM.yyyy"),
            "end_date": self.end_date.date().toString("dd.MM.yyyy"),
            "amount": self.amount_spin.value(),
            "active": self.active_checkbox.isChecked(),
            "last_generated": None,
            "next_generation": self.calculate_next_generation()
        }
        
        self.accept()
    
    def calculate_next_generation(self):
        today = QDate.currentDate()
        frequency = self.frequency_combo.currentText()
        
        if frequency == "Měsíčně":
            return today.addMonths(1).toString("dd.MM.yyyy")
        elif frequency == "Čtvrtletně":
            return today.addMonths(3).toString("dd.MM.yyyy")
        elif frequency == "Pololetně":
            return today.addMonths(6).toString("dd.MM.yyyy")
        elif frequency == "Ročně":
            return today.addYears(1).toString("dd.MM.yyyy")
        
        return today.toString("dd.MM.yyyy")

# ================= 7. SPRÁVA SPLATNOSTÍ (upraveno o jednotku) =================
class SplatnostiDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Správa splatností")
        self.setMinimumSize(600, 400)
        
        self.splatnosti = load_splatnosti()
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Tabulka splatností
        self.splatnosti_table = QTableWidget()
        self.splatnosti_table.setColumnCount(5)  # přidán sloupec Jednotka
        self.splatnosti_table.setHorizontalHeaderLabels(["Název", "Počet", "Jednotka", "Aktivní", "Akce"])
        self.splatnosti_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        layout.addWidget(self.splatnosti_table)
        
        # Tlačítka
        buttons_layout = QHBoxLayout()
        
        btn_add = QPushButton("➕ Přidat splatnost")
        btn_edit = QPushButton("✏️ Upravit")
        btn_delete = QPushButton("🗑 Smazat")
        btn_toggle = QPushButton("🔄 Aktivovat/Deaktivovat")
        
        btn_add.clicked.connect(self.add_splatnost)
        btn_edit.clicked.connect(self.edit_splatnost)
        btn_delete.clicked.connect(self.delete_splatnost)
        btn_toggle.clicked.connect(self.toggle_active)
        
        buttons_layout.addWidget(btn_add)
        buttons_layout.addWidget(btn_edit)
        buttons_layout.addWidget(btn_delete)
        buttons_layout.addWidget(btn_toggle)
        
        layout.addLayout(buttons_layout)
        
        # Info
        info_label = QLabel("💡 Tip: Hodiny 0 = okamžitě, Dny 0 = dnes do půlnoci")
        info_label.setStyleSheet("color: gray; font-style: italic;")
        layout.addWidget(info_label)
        
        self.setLayout(layout)
        self.load_splatnosti_table()
    
    def load_splatnosti_table(self):
        self.splatnosti_table.setRowCount(len(self.splatnosti))
        
        for i, splatnost in enumerate(self.splatnosti):
            self.splatnosti_table.setItem(i, 0, QTableWidgetItem(splatnost["nazev"]))
            self.splatnosti_table.setItem(i, 1, QTableWidgetItem(str(splatnost["hodiny"])))
            self.splatnosti_table.setItem(i, 2, QTableWidgetItem(splatnost.get("jednotka", "hodiny")))
            
            # Aktivní checkbox
            active_checkbox = QCheckBox()
            active_checkbox.setChecked(splatnost.get("aktivni", True))
            active_checkbox.stateChanged.connect(lambda state, idx=i: self.toggle_splatnost_active(idx, state))
            self.splatnosti_table.setCellWidget(i, 3, active_checkbox)
            
            # Tlačítka akcí
            button_layout = QHBoxLayout()
            edit_button = QPushButton("Upravit")
            delete_button = QPushButton("Smazat")
            
            edit_button.clicked.connect(lambda checked, idx=i: self.edit_splatnost_by_index(idx))
            delete_button.clicked.connect(lambda checked, idx=i: self.delete_splatnost_by_index(idx))
            
            button_widget = QWidget()
            button_layout.addWidget(edit_button)
            button_layout.addWidget(delete_button)
            button_widget.setLayout(button_layout)
            
            self.splatnosti_table.setCellWidget(i, 4, button_widget)
    
    def add_splatnost(self):
        dialog = SplatnostEditDialog(self)
        if dialog.exec():
            new_splatnost = dialog.get_splatnost()
            self.splatnosti.append(new_splatnost)
            save_splatnosti(self.splatnosti)
            self.load_splatnosti_table()
    
    def edit_splatnost(self):
        selected = self.splatnosti_table.currentRow()
        if selected >= 0:
            self.edit_splatnost_by_index(selected)
    
    def edit_splatnost_by_index(self, index):
        if 0 <= index < len(self.splatnosti):
            splatnost = self.splatnosti[index]
            dialog = SplatnostEditDialog(self, splatnost)
            if dialog.exec():
                self.splatnosti[index] = dialog.get_splatnost()
                save_splatnosti(self.splatnosti)
                self.load_splatnosti_table()
    
    def delete_splatnost(self):
        selected = self.splatnosti_table.currentRow()
        if selected >= 0:
            self.delete_splatnost_by_index(selected)
    
    def delete_splatnost_by_index(self, index):
        if 0 <= index < len(self.splatnosti):
            if len(self.splatnosti) <= 1:
                QMessageBox.warning(self, "Chyba", "Musí zůstat alespoň jedna splatnost")
                return
            
            splatnost = self.splatnosti[index]
            reply = QMessageBox.question(
                self, "Smazat splatnost",
                f"Opravdu chcete smazat splatnost '{splatnost['nazev']}'?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                del self.splatnosti[index]
                save_splatnosti(self.splatnosti)
                self.load_splatnosti_table()
    
    def toggle_active(self):
        selected = self.splatnosti_table.currentRow()
        if selected >= 0:
            self.toggle_splatnost_active(selected, None)
    
    def toggle_splatnost_active(self, index, state):
        if 0 <= index < len(self.splatnosti):
            self.splatnosti[index]["aktivni"] = not self.splatnosti[index].get("aktivni", True)
            save_splatnosti(self.splatnosti)
            self.load_splatnosti_table()

class SplatnostEditDialog(QDialog):
    def __init__(self, parent=None, splatnost=None):
        super().__init__(parent)
        self.splatnost = splatnost or {}
        self.is_edit = bool(splatnost)
        
        if self.is_edit:
            self.setWindowTitle("Upravit splatnost")
        else:
            self.setWindowTitle("Nová splatnost")
        
        self.setFixedWidth(300)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        # Název
        self.nazev_edit = QLineEdit(self.splatnost.get("nazev", ""))
        form_layout.addRow("Název splatnosti:", self.nazev_edit)
        
        # Počet
        self.hodiny_spin = QSpinBox()
        self.hodiny_spin.setRange(0, 720)
        self.hodiny_spin.setValue(self.splatnost.get("hodiny", 0))
        # bez suffixu, jednotka se vybírá zvlášť
        form_layout.addRow("Počet:", self.hodiny_spin)
        
        # Jednotka
        self.jednotka_combo = QComboBox()
        self.jednotka_combo.addItems(["hodiny", "dny"])
        self.jednotka_combo.setCurrentText(self.splatnost.get("jednotka", "hodiny"))
        form_layout.addRow("Jednotka:", self.jednotka_combo)
        
        # Aktivní
        self.active_checkbox = QCheckBox("Aktivní")
        self.active_checkbox.setChecked(self.splatnost.get("aktivni", True))
        form_layout.addRow("", self.active_checkbox)
        
        layout.addLayout(form_layout)
        
        # Tlačítka
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_splatnost)
        cancel_button.clicked.connect(self.reject)
        
        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        
        layout.addLayout(buttons_layout)
        self.setLayout(layout)
    
    def save_splatnost(self):
        if not self.nazev_edit.text():
            QMessageBox.warning(self, "Chyba", "Zadejte název splatnosti")
            return
        
        # Kontrola duplicitního názvu
        existing_splatnosti = load_splatnosti()
        for s in existing_splatnosti:
            if s["nazev"] == self.nazev_edit.text() and (not self.is_edit or self.splatnost.get("nazev") != self.nazev_edit.text()):
                QMessageBox.warning(self, "Chyba", "Splatnost s tímto názvem již existuje")
                return
        
        self.splatnost = {
            "nazev": self.nazev_edit.text(),
            "hodiny": self.hodiny_spin.value(),
            "jednotka": self.jednotka_combo.currentText(),
            "aktivni": self.active_checkbox.isChecked()
        }
        
        self.accept()
    
    def get_splatnost(self):
        return self.splatnost

# ================= 8. OFFLINE REŽIM A RECOVERY =================
class RecoveryDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Obnova dat")
        self.setMinimumSize(500, 400)
        
        self.load_recovery_data()
        self.init_ui()
    
    def load_recovery_data(self):
        if not os.path.exists(RECOVERY_FILE):
            self.recovery_data = {
                "auto_save": True,
                "save_interval": 5,
                "versions_to_keep": 10,
                "last_backup": None,
                "recovery_points": []
            }
            self.save_recovery_data()
        else:
            with open(RECOVERY_FILE, 'r', encoding='utf-8') as f:
                self.recovery_data = json.load(f)
    
    def save_recovery_data(self):
        with open(RECOVERY_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.recovery_data, f, indent=2, ensure_ascii=False)
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Auto-save settings
        auto_save_group = QGroupBox("Automatické ukládání")
        auto_save_layout = QFormLayout()
        
        self.auto_save_checkbox = QCheckBox("Povolit automatické ukládání")
        self.auto_save_checkbox.setChecked(self.recovery_data["auto_save"])
        auto_save_layout.addRow("", self.auto_save_checkbox)
        
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 60)
        self.interval_spin.setValue(self.recovery_data["save_interval"])
        self.interval_spin.setSuffix(" minut")
        auto_save_layout.addRow("Interval ukládání:", self.interval_spin)
        
        self.versions_spin = QSpinBox()
        self.versions_spin.setRange(1, 50)
        self.versions_spin.setValue(self.recovery_data["versions_to_keep"])
        self.versions_spin.setSuffix(" verzí")
        auto_save_layout.addRow("Uchovat verzí:", self.versions_spin)
        
        auto_save_group.setLayout(auto_save_layout)
        layout.addWidget(auto_save_group)
        
        # Recovery points
        recovery_group = QGroupBox("Body obnovy")
        recovery_layout = QVBoxLayout()
        
        self.recovery_list = QListWidget()
        for point in self.recovery_data["recovery_points"]:
            self.recovery_list.addItem(f"{point['time']} - {point['description']}")
        recovery_layout.addWidget(self.recovery_list)
        
        recovery_buttons = QHBoxLayout()
        create_button = QPushButton("Vytvořit bod obnovy")
        restore_button = QPushButton("Obnovit vybraný")
        delete_button = QPushButton("Smazat")
        
        create_button.clicked.connect(self.create_recovery_point)
        restore_button.clicked.connect(self.restore_recovery_point)
        delete_button.clicked.connect(self.delete_recovery_point)
        
        recovery_buttons.addWidget(create_button)
        recovery_buttons.addWidget(restore_button)
        recovery_buttons.addWidget(delete_button)
        
        recovery_layout.addLayout(recovery_buttons)
        recovery_group.setLayout(recovery_layout)
        layout.addWidget(recovery_group)
        
        # Last backup info
        if self.recovery_data["last_backup"]:
            last_backup_label = QLabel(f"Poslední záloha: {self.recovery_data['last_backup']}")
            layout.addWidget(last_backup_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Uložit nastavení")
        close_button = QPushButton("Zavřít")
        
        save_button.clicked.connect(self.save_settings)
        close_button.clicked.connect(self.accept)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(close_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_settings(self):
        self.recovery_data.update({
            "auto_save": self.auto_save_checkbox.isChecked(),
            "save_interval": self.interval_spin.value(),
            "versions_to_keep": self.versions_spin.value()
        })
        self.save_recovery_data()
        QMessageBox.information(self, "Nastavení", "Nastavení bylo uloženo")
    
    def create_recovery_point(self):
        description, ok = QInputDialog.getText(self, "Bod obnovy", "Popis bodu obnovy:")
        if ok and description:
            # Create recovery point
            recovery_point = {
                "time": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "description": description,
                "data": {
                    "archive": load_archive_data(),
                    "settings": settings,
                    "services": services,
                    "splatnosti": load_splatnosti(),
                    "counter": open(COUNTER_FILE, 'r').read() if os.path.exists(COUNTER_FILE) else ""
                }
            }
            
            self.recovery_data["recovery_points"].append(recovery_point)
            
            # Keep only limited number of versions
            if len(self.recovery_data["recovery_points"]) > self.recovery_data["versions_to_keep"]:
                self.recovery_data["recovery_points"] = self.recovery_data["recovery_points"][-self.recovery_data["versions_to_keep"]:]
            
            self.save_recovery_data()
            self.recovery_list.addItem(f"{recovery_point['time']} - {recovery_point['description']}")
    
    def restore_recovery_point(self):
        selected = self.recovery_list.currentRow()
        if selected >= 0:
            reply = QMessageBox.question(
                self, "Obnovit bod",
                "Opravdu chcete obnovit tento bod obnovy?\nAktuální data budou přepsána.",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                point = self.recovery_data["recovery_points"][selected]
                
                # Restore data
                save_archive_data(point["data"]["archive"])
                save_settings(point["data"]["settings"])
                save_services(point["data"]["services"])
                save_splatnosti(point["data"]["splatnosti"])
                
                if point["data"]["counter"]:
                    with open(COUNTER_FILE, 'w') as f:
                        f.write(point["data"]["counter"])
                
                QMessageBox.information(self, "Obnoveno", "Data byla obnovena")
                self.accept()
    
    def delete_recovery_point(self):
        selected = self.recovery_list.currentRow()
        if selected >= 0:
            reply = QMessageBox.question(
                self, "Smazat bod obnovy",
                "Opravdu chcete smazat tento bod obnovy?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                del self.recovery_data["recovery_points"][selected]
                self.save_recovery_data()
                self.recovery_list.takeItem(selected)

# ================= PŮVODNÍ ZÁSOBNÍK DOKLADŮ =================
class ArchiveDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Zásobník dokladů")
        self.setMinimumSize(1000, 600)
        
        self.init_ui()
        self.load_data()
        
        # Timer pro pravidelnou aktualizaci stavů
        self.timer = QTimer()
        self.timer.timeout.connect(self.refresh_table)
        self.timer.start(60000)

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Tabulka
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "Číslo faktury", 
            "Jméno zákazníka", 
            "Datum konání", 
            "Datum vytvoření", 
            "Stav platby", 
            "Cena (CZK)",
            "Vydal",
            "Služba"
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.open_pdf)
        
        # Tlačítka
        btn_layout = QHBoxLayout()
        
        self.btn_paid = QPushButton("✓ Uhrazeno")
        self.btn_paid.clicked.connect(self.mark_as_paid)
        
        self.btn_unpaid = QPushButton("✗ Neuhrazeno")
        self.btn_unpaid.clicked.connect(self.mark_as_unpaid)
        
        self.btn_delete = QPushButton("🗑 Smazat")
        self.btn_delete.clicked.connect(self.delete_selected)
        
        self.btn_refresh = QPushButton("🔄 Obnovit")
        self.btn_refresh.clicked.connect(self.refresh_table)
        
        self.btn_export = QPushButton("📊 Export CSV")
        self.btn_export.clicked.connect(self.export_csv)
        
        btn_layout.addWidget(self.btn_paid)
        btn_layout.addWidget(self.btn_unpaid)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_refresh)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addStretch()
        
        layout.addWidget(self.table)
        layout.addLayout(btn_layout)
        
        # Kontextové menu
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, position):
        if self.table.selectedItems():
            menu = QMenu()
            open_action = QAction("Otevřít PDF", self)
            open_action.triggered.connect(self.open_pdf)
            menu.addAction(open_action)
            menu.exec(self.table.viewport().mapToGlobal(position))

    def load_data(self):
        # Aktualizovat stavy podle času
        archive_data = update_archive_statuses()
        
        self.table.setRowCount(len(archive_data))
        
        for row, item in enumerate(archive_data):
            # Číslo faktury
            self.table.setItem(row, 0, QTableWidgetItem(item["cislo_faktury"]))
            
            # Jméno
            self.table.setItem(row, 1, QTableWidgetItem(item["jmeno"]))
            
            # Datum konání
            self.table.setItem(row, 2, QTableWidgetItem(item["termin"]))
            
            # Datum vytvoření
            self.table.setItem(row, 3, QTableWidgetItem(item["datum_vytvoreni"]))
            
            # Stav platby
            status_item = QTableWidgetItem()
            if item["stav"] == 0:
                status_item.setText("NEUHRAZENO")
                status_item.setBackground(QColor(255, 200, 200))
            elif item["stav"] == 1:
                status_item.setText("UHRAZENO")
                status_item.setBackground(QColor(200, 255, 200))
            else:
                status_item.setText("K UHRAZENÍ")
                status_item.setBackground(QColor(200, 220, 255))
            self.table.setItem(row, 4, status_item)
            
            # Cena
            self.table.setItem(row, 5, QTableWidgetItem(f"{item['cena']} CZK"))
            
            # Vydal
            self.table.setItem(row, 6, QTableWidgetItem(item.get("vydal", "Neznámý")))
            
            # Služba
            self.table.setItem(row, 7, QTableWidgetItem(item["sluzba"]))
        
        self.table.resizeColumnsToContents()

    def refresh_table(self):
        self.load_data()

    def get_selected_invoice_number(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return None
        row = selected_rows[0].row()
        return self.table.item(row, 0).text()

    def mark_as_paid(self):
        invoice_no = self.get_selected_invoice_number()
        if invoice_no:
            if update_payment_status(invoice_no, 1):
                self.refresh_table()

    def mark_as_unpaid(self):
        invoice_no = self.get_selected_invoice_number()
        if invoice_no:
            if update_payment_status(invoice_no, 0):
                self.refresh_table()

    def delete_selected(self):
        invoice_no = self.get_selected_invoice_number()
        if invoice_no:
            reply = QMessageBox.question(self, "Smazat fakturu", 
                                        f"Opravdu chcete smazat fakturu {invoice_no}?",
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                if remove_from_archive(invoice_no):
                    self.refresh_table()

    def open_pdf(self):
        invoice_no = self.get_selected_invoice_number()
        if invoice_no:
            archive_data = load_archive_data()
            for item in archive_data:
                if item["cislo_faktury"] == invoice_no:
                    if os.path.exists(item["archiv_path"]):
                        os.startfile(item["archiv_path"])
                    else:
                        QMessageBox.warning(self, "Chyba", "Soubor PDF nebyl nalezen.")
    
    def export_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exportovat do CSV", "", "CSV Files (*.csv)"
        )
        if file_path:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(["Číslo faktury", "Jméno zákazníka", "Datum konání", "Datum vytvoření", "Stav platby", "Cena (CZK)", "Vydal", "Služba"])
                
                for row in range(self.table.rowCount()):
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        if item:
                            row_data.append(item.text())
                        else:
                            row_data.append("")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Export", "Data byla exportována do CSV.")

# ================= SPRÁVA SLUŽEB =================
class ServicesDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Správa služeb a splatností")
        self.setMinimumSize(900, 600)
        
        self.services = load_services()
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        tabs = QTabWidget()
        
        # Služby tab
        services_tab = QWidget()
        services_layout = QVBoxLayout()
        
        # TreeWidget pro zobrazení služeb
        self.services_tree = QTreeWidget()
        self.services_tree.setHeaderLabels(["Název", "Cena (CZK)", "Typ", "Status"])
        self.services_tree.setColumnCount(4)
        
        # Naplnění stromu
        self.load_services_tree()
        
        services_layout.addWidget(self.services_tree)
        
        # Tlačítka pro správu služeb
        services_buttons = QHBoxLayout()
        
        btn_add_category = QPushButton("➕ Přidat kategorii")
        btn_add_service = QPushButton("➕ Přidat službu")
        btn_edit = QPushButton("✏️ Upravit")
        btn_delete = QPushButton("🗑 Smazat")
        btn_toggle = QPushButton("🔄 Aktivovat/Deaktivovat")
        
        btn_add_category.clicked.connect(self.add_category)
        btn_add_service.clicked.connect(self.add_service)
        btn_edit.clicked.connect(self.edit_item)
        btn_delete.clicked.connect(self.delete_item)
        btn_toggle.clicked.connect(self.toggle_active)
        
        services_buttons.addWidget(btn_add_category)
        services_buttons.addWidget(btn_add_service)
        services_buttons.addWidget(btn_edit)
        services_buttons.addWidget(btn_delete)
        services_buttons.addWidget(btn_toggle)
        
        services_layout.addLayout(services_buttons)
        
        # Info
        services_info = QLabel("💡 Tipy: Wellness (vířivka, sauna), Ubytování (pokoje), Ostatní (balíčky)")
        services_info.setStyleSheet("color: gray; font-style: italic;")
        services_layout.addWidget(services_info)
        
        services_tab.setLayout(services_layout)
        
        # Splatnosti tab
        splatnosti_tab = QWidget()
        splatnosti_layout = QVBoxLayout()
        
        splatnosti_button = QPushButton("Spravovat splatnosti")
        splatnosti_button.clicked.connect(self.manage_splatnosti)
        splatnosti_layout.addWidget(splatnosti_button)
        
        splatnosti_info = QLabel("💡 Zde můžete přidávat, upravovat a mazat splatnosti faktur")
        splatnosti_info.setStyleSheet("color: gray; font-style: italic;")
        splatnosti_layout.addWidget(splatnosti_info)
        
        splatnosti_tab.setLayout(splatnosti_layout)
        
        tabs.addTab(services_tab, "Služby")
        tabs.addTab(splatnosti_tab, "Splatnosti")
        
        layout.addWidget(tabs)
        
        self.setLayout(layout)
    
    def load_services_tree(self):
        self.services_tree.clear()
        
        for category, services_list in self.services.items():
            category_item = QTreeWidgetItem(self.services_tree, [category, "", "", ""])
            category_item.setData(0, Qt.UserRole, ("category", category))
            
            for service in services_list:
                status = "Aktivní" if service.get("aktivni", True) else "Neaktivní"
                service_item = QTreeWidgetItem(category_item, [
                    service["nazev"], 
                    str(service["cena"]), 
                    service.get("typ", "ostatni"),
                    status
                ])
                service_item.setData(0, Qt.UserRole, ("service", category, service["nazev"]))
                
                if not service.get("aktivni", True):
                    service_item.setForeground(0, QColor("gray"))
            
            category_item.setExpanded(True)
    
    def manage_splatnosti(self):
        dialog = SplatnostiDialog(self)
        dialog.exec()
    
    def add_category(self):
        name, ok = QInputDialog.getText(self, "Nová kategorie", "Název kategorie:")
        if ok and name:
            if name not in self.services:
                self.services[name] = []
                save_services(self.services)
                self.load_services_tree()
    
    def add_service(self):
        selected = self.services_tree.currentItem()
        if not selected:
            QMessageBox.warning(self, "Chyba", "Vyberte kategorii, do které chcete přidat službu")
            return
        
        data = selected.data(0, Qt.UserRole)
        if not data or data[0] != "category":
            QMessageBox.warning(self, "Chyba", "Vyberte kategorii, do které chcete přidat službu")
            return
        
        category = data[1]
        
        dialog = ServiceEditDialog(self, None, category)
        if dialog.exec():
            new_service = dialog.get_service()
            self.services[category].append(new_service)
            save_services(self.services)
            self.load_services_tree()
    
    def edit_item(self):
        selected = self.services_tree.currentItem()
        if not selected:
            return
        
        data = selected.data(0, Qt.UserRole)
        if not data:
            return
        
        if data[0] == "category":
            # Editace kategorie
            old_name = data[1]
            new_name, ok = QInputDialog.getText(self, "Upravit kategorii", "Nový název kategorie:", text=old_name)
            if ok and new_name and new_name != old_name:
                self.services[new_name] = self.services.pop(old_name)
                save_services(self.services)
                self.load_services_tree()
        
        elif data[0] == "service":
            # Editace služby
            category = data[1]
            service_name = data[2]
            
            # Najít službu
            for service in self.services[category]:
                if service["nazev"] == service_name:
                    dialog = ServiceEditDialog(self, service, category)
                    if dialog.exec():
                        updated_service = dialog.get_service()
                        # Nahradit starou službu novou
                        index = self.services[category].index(service)
                        self.services[category][index] = updated_service
                        save_services(self.services)
                        self.load_services_tree()
                    break
    
    def delete_item(self):
        selected = self.services_tree.currentItem()
        if not selected:
            return
        
        data = selected.data(0, Qt.UserRole)
        if not data:
            return
        
        if data[0] == "category":
            category = data[1]
            if len(self.services[category]) > 0:
                reply = QMessageBox.question(self, "Smazat kategorii", 
                                           f"Kategorie '{category}' obsahuje služby. Opravdu chcete smazat celou kategorii včetně všech služeb?",
                                           QMessageBox.Yes | QMessageBox.No)
                if reply != QMessageBox.Yes:
                    return
            
            del self.services[category]
        
        elif data[0] == "service":
            category = data[1]
            service_name = data[2]
            
            self.services[category] = [s for s in self.services[category] if s["nazev"] != service_name]
        
        save_services(self.services)
        self.load_services_tree()
    
    def toggle_active(self):
        selected = self.services_tree.currentItem()
        if not selected:
            return
        
        data = selected.data(0, Qt.UserRole)
        if not data or data[0] != "service":
            return
        
        category = data[1]
        service_name = data[2]
        
        for service in self.services[category]:
            if service["nazev"] == service_name:
                service["aktivni"] = not service.get("aktivni", True)
                break
        
        save_services(self.services)
        self.load_services_tree()

class ServiceEditDialog(QDialog):
    def __init__(self, parent=None, service=None, category=None):
        super().__init__(parent)
        self.service = service or {}
        self.category = category
        self.is_edit = bool(service)
        
        if self.is_edit:
            self.setWindowTitle("Upravit službu")
        else:
            self.setWindowTitle("Nová služba")
        
        self.setFixedWidth(400)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        # Název služby
        self.name_edit = QLineEdit(self.service.get("nazev", ""))
        form_layout.addRow("Název služby:", self.name_edit)
        
        # Cena
        self.price_spin = QDoubleSpinBox()
        self.price_spin.setRange(0, 100000)
        self.price_spin.setValue(self.service.get("cena", 0))
        self.price_spin.setSuffix(" Kč")
        form_layout.addRow("Cena:", self.price_spin)
        
        # Typ služby
        self.type_combo = QComboBox()
        self.type_combo.addItems(["wellness", "ubytovani", "ostatni"])
        if "typ" in self.service:
            self.type_combo.setCurrentText(self.service["typ"])
        form_layout.addRow("Typ služby:", self.type_combo)
        
        # Aktivní
        self.active_checkbox = QCheckBox("Aktivní služba")
        self.active_checkbox.setChecked(self.service.get("aktivni", True))
        form_layout.addRow("", self.active_checkbox)
        
        layout.addLayout(form_layout)
        
        # Popis (volitelný)
        layout.addWidget(QLabel("Popis (volitelné):"))
        self.desc_edit = QTextEdit()
        self.desc_edit.setPlainText(self.service.get("popis", ""))
        self.desc_edit.setMaximumHeight(80)
        layout.addWidget(self.desc_edit)
        
        # Tlačítka
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("Uložit")
        cancel_button = QPushButton("Zrušit")
        
        save_button.clicked.connect(self.save_service)
        cancel_button.clicked.connect(self.reject)
        
        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        
        layout.addLayout(buttons_layout)
        self.setLayout(layout)
    
    def save_service(self):
        if not self.name_edit.text():
            QMessageBox.warning(self, "Chyba", "Zadejte název služby")
            return
        
        self.service = {
            "nazev": self.name_edit.text(),
            "cena": self.price_spin.value(),
            "typ": self.type_combo.currentText(),
            "aktivni": self.active_checkbox.isChecked(),
            "popis": self.desc_edit.toPlainText()
        }
        
        self.accept()
    
    def get_service(self):
        return self.service

# ================= ROZŠÍŘENÉ NASTAVENÍ =================
class SettingsDialog(QDialog):
    def __init__(self, user_manager):
        super().__init__()
        self.user_manager = user_manager
        self.setWindowTitle("Nastavení")
        self.setMinimumWidth(900)
        self.setMinimumHeight(700)

        tabs = QTabWidget()

        # Obecné
        general = QWidget()
        g = QFormLayout(general)
        self.auto = QCheckBox("Automaticky otevřít PDF")
        self.auto.setChecked(settings["open_pdf"])
        g.addRow(self.auto)
        
        # E-mailové notifikace
        self.email_checkbox = QCheckBox("Povolit e-mailové notifikace")
        self.email_checkbox.setChecked(settings.get("email_notifications", False))
        g.addRow(self.email_checkbox)
        
        # Automatické zálohování
        self.backup_checkbox = QCheckBox("Automatické zálohování")
        self.backup_checkbox.setChecked(settings.get("auto_backup", True))
        g.addRow(self.backup_checkbox)

        # Tlačítko pro odhlášení
        logout_btn = QPushButton("Odhlásit")
        logout_btn.clicked.connect(self.logout)
        g.addRow("", logout_btn)

        # Vzhled
        theme = QWidget()
        t = QFormLayout(theme)
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Systémový (default)", "Světlý", "Tmavý"])
        
        current_theme = settings.get("theme", "system")
        if current_theme == "system":
            self.theme_combo.setCurrentIndex(0)
        elif current_theme == "light":
            self.theme_combo.setCurrentIndex(1)
        elif current_theme == "dark":
            self.theme_combo.setCurrentIndex(2)
            
        t.addRow("Režim vzhledu:", self.theme_combo)
        
        note = QLabel("Změna se projeví po restartování aplikace.")
        note.setStyleSheet("color: gray; font-style: italic;")
        t.addRow("", note)

        # Counter
        counter = QWidget()
        cl = QVBoxLayout(counter)

        btn_backup = QPushButton("Záloha counteru na plochu")
        btn_restore = QPushButton("Obnovit counter z plochy")
        btn_reset = QPushButton("Reset counteru")

        btn_backup.clicked.connect(self.backup_counter)
        btn_restore.clicked.connect(self.restore_counter)
        btn_reset.clicked.connect(self.reset_counter)

        cl.addWidget(btn_backup)
        cl.addWidget(btn_restore)
        cl.addWidget(btn_reset)
        cl.addStretch()

        # Archiv
        archive = QWidget()
        al = QVBoxLayout(archive)
        
        btn_open_archive = QPushButton("Otevřít složku s archivem")
        btn_open_archive.clicked.connect(self.open_archive_folder)
        
        btn_clear_archive = QPushButton("Vyčistit archiv")
        btn_clear_archive.clicked.connect(self.clear_archive)
        
        archive_info = QLabel(f"Archiv faktur: {ARCHIVE_DIR}")
        archive_info.setWordWrap(True)
        archive_info.setStyleSheet("color: gray; font-style: italic;")
        
        al.addWidget(archive_info)
        al.addWidget(btn_open_archive)
        al.addWidget(btn_clear_archive)
        al.addStretch()

        # Společnost
        company = QWidget()
        company_layout = QFormLayout()
        
        self.company_name = QLineEdit(settings.get("company_name", ""))
        self.company_address = QLineEdit(settings.get("company_address", ""))
        self.company_id = QLineEdit(settings.get("company_id", ""))
        self.company_vat = QLineEdit(settings.get("company_vat", ""))
        self.tax_rate = QDoubleSpinBox()
        self.tax_rate.setRange(0, 100)
        self.tax_rate.setValue(settings.get("tax_rate", 21))
        self.tax_rate.setSuffix(" %")
        
        company_layout.addRow("Název společnosti:", self.company_name)
        company_layout.addRow("Adresa:", self.company_address)
        company_layout.addRow("IČO:", self.company_id)
        company_layout.addRow("DIČ:", self.company_vat)
        company_layout.addRow("Sazba DPH:", self.tax_rate)
        
        # Provozovna
        company_layout.addWidget(QLabel("<b>Provozovna:</b>"))
        
        self.branch_same_checkbox = QCheckBox("Stejné jako údaje fakturační")
        self.branch_same_checkbox.setChecked(settings.get("branch_same", True))
        company_layout.addRow("", self.branch_same_checkbox)
        
        self.branch_name = QLineEdit(settings.get("branch_name", ""))
        self.branch_address = QLineEdit(settings.get("branch_address", ""))
        
        company_layout.addRow("Název provozovny:", self.branch_name)
        company_layout.addRow("Adresa provozovny:", self.branch_address)
        
        # Povolit/znepřístupnit pole podle checkboxu
        self.branch_same_checkbox.stateChanged.connect(self.toggle_branch_fields)
        self.toggle_branch_fields()
        
        company.setLayout(company_layout)

        # Služby a splatnosti
        services_tab = QWidget()
        services_layout = QVBoxLayout()
        
        services_info = QLabel("Správa služeb a splatností:")
        services_info.setStyleSheet("font-weight: bold;")
        services_layout.addWidget(services_info)
        
        btn_manage_services = QPushButton("Spravovat služby a splatnosti")
        btn_manage_services.clicked.connect(self.manage_services)
        services_layout.addWidget(btn_manage_services)
        
        services_info2 = QLabel("💡 Zde můžete přidávat, upravovat a mazat služby a splatnosti faktur")
        services_info2.setStyleSheet("color: gray; font-style: italic;")
        services_layout.addWidget(services_info2)
        
        services_layout.addStretch()
        services_tab.setLayout(services_layout)

        # O programu
        about = QWidget()
        a = QVBoxLayout(about)
        lbl = QLabel(
            "HEM - Zálohové faktury\n\n"
            "Verze: v2.0.1\n"
            "Vývojář: JAMAsoft\n"
            "Web: www.jamasoft.cz\n\n"
            "Systém pro správu zálohových faktur\n"
            "Wellness, Ubytování a ostatní služby\n\n"
            "© 2025 HEM - Hotel Easy Manager\n\n"
            "Používáno ve Wellness Hotelu Beethoven"
        )
        lbl.setWordWrap(True)
        lbl.setAlignment(Qt.AlignCenter)
        a.addWidget(lbl)

        tabs.addTab(general, "Obecné")
        tabs.addTab(theme, "Vzhled")
        tabs.addTab(counter, "Faktury / Counter")
        tabs.addTab(archive, "Archiv")
        tabs.addTab(company, "Společnost")
        tabs.addTab(services_tab, "Služby a splatnosti")
        tabs.addTab(about, "O programu")

        save = QPushButton("Uložit")
        save.clicked.connect(self.save)

        layout = QVBoxLayout(self)
        layout.addWidget(tabs)
        layout.addWidget(save, alignment=Qt.AlignRight)
    
    def toggle_branch_fields(self):
        enabled = not self.branch_same_checkbox.isChecked()
        self.branch_name.setEnabled(enabled)
        self.branch_address.setEnabled(enabled)
    
    def manage_services(self):
        dialog = ServicesDialog()
        dialog.exec()
    
    def logout(self):
        reply = QMessageBox.question(self, "Odhlásit", 
                                   "Opravdu chcete odhlásit?\n"
                                   "Budete přesměrováni na přihlašovací obrazovku.",
                                   QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            # Vymazat uložené přihlášení
            clear_last_login()
            self.reject()
            # Zavřít aplikaci - v hlavním okně se provede restart
    
    def require_password(self):
        dlg = PasswordDialog()
        if dlg.exec() and dlg.password() == COUNTER_PASSWORD:
            return True
        QMessageBox.critical(self, "Chyba", "Nesprávné heslo.")
        return False

    def backup_counter(self):
        if os.path.exists(COUNTER_FILE):
            shutil.copy(COUNTER_FILE, os.path.join(desktop_path(), "faktury_counter_backup.txt"))
            QMessageBox.information(self, "Hotovo", "Counter byl zazálohován na plochu.")
        else:
            QMessageBox.warning(self, "Chyba", "Counter neexistuje.")

    def restore_counter(self):
        if not self.require_password():
            return
        src = os.path.join(desktop_path(), "faktury_counter_backup.txt")
        if os.path.exists(src):
            shutil.copy(src, COUNTER_FILE)
            QMessageBox.information(self, "Hotovo", "Counter byl obnoven.")
        else:
            QMessageBox.warning(self, "Chyba", "Záloha na ploše nebyla nalezena.")

    def reset_counter(self):
        if not self.require_password():
            return
        with open(COUNTER_FILE, "w", encoding="utf-8") as f:
            f.write(f"{datetime.now().year};0")
        QMessageBox.information(self, "Hotovo", "Counter byl resetován.")

    def open_archive_folder(self):
        ensure_archive_dir()
        os.startfile(ARCHIVE_DIR)

    def clear_archive(self):
        if not self.require_password():
            return
            
        reply = QMessageBox.question(self, "Vyčistit archiv", 
                                    "Opravdu chcete smazat všechny faktury v archivu?\n"
                                    "Tato akce je nevratná!",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                # Smazat soubory v archivu
                for filename in os.listdir(ARCHIVE_DIR):
                    file_path = os.path.join(ARCHIVE_DIR, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                
                # Smazat data archivu
                if os.path.exists(ARCHIVE_DATA_FILE):
                    os.remove(ARCHIVE_DATA_FILE)
                
                QMessageBox.information(self, "Hotovo", "Archiv byl vyčištěn.")
            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se vyčistit archiv: {str(e)}")

    def save(self):
        try:
            settings["open_pdf"] = self.auto.isChecked()
            settings["email_notifications"] = self.email_checkbox.isChecked()
            settings["auto_backup"] = self.backup_checkbox.isChecked()
            
            theme_index = self.theme_combo.currentIndex()
            if theme_index == 0:
                settings["theme"] = "system"
            elif theme_index == 1:
                settings["theme"] = "light"
            elif theme_index == 2:
                settings["theme"] = "dark"
            
            # Company settings
            settings["company_name"] = self.company_name.text()
            settings["company_address"] = self.company_address.text()
            settings["company_id"] = self.company_id.text()
            settings["company_vat"] = self.company_vat.text()
            settings["tax_rate"] = self.tax_rate.value()
            
            # Branch settings
            settings["branch_same"] = self.branch_same_checkbox.isChecked()
            settings["branch_name"] = self.branch_name.text()
            settings["branch_address"] = self.branch_address.text()
            
            save_settings(settings)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při ukládání: {str(e)}")

# ================= HLAVNÍ OKNO S UPRAVENÝMI FUNKCEMI =================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HEM - Zálohové faktury")
        self.resize(900, 700)
        
        # Inicializace managerů
        self.user_manager = UserManager()
        self.backup_manager = BackupManager()
        
        # Nejdříve přihlášení
        if not self.login():
            sys.exit()
        
        self.init_ui()
    
    def login(self):
        dlg = LoginDialog(self.user_manager)
        return dlg.exec() == QDialog.Accepted
    
    def init_ui(self):
        # Hlavní widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        
        # Formulář pro faktury (upravené)
        form_group = QGroupBox("Nová zálohová faktura")
        form_layout = QFormLayout()
        
        self.jmeno = QLineEdit()
        self.email = QLineEdit()
        self.tel = QLineEdit()
        
        # Combo box pro služby - editable
        self.sluzba = QComboBox()
        self.sluzba.setEditable(True)
        self.sluzba.setInsertPolicy(QComboBox.NoInsert)
        self.load_services_to_combo()
        
        # Termín s kalendářem - podporuje i čas
        termin_layout = QHBoxLayout()
        self.termin = QLineEdit()
        self.termin.setPlaceholderText("DD.MM.RRRR nebo DD.MM.RRRR HH:MM")
        termin_layout.addWidget(self.termin)
        
        # Tlačítko kalendáře
        calendar_btn = QPushButton("📅")
        calendar_btn.setFixedWidth(30)
        calendar_btn.clicked.connect(self.show_calendar)
        termin_layout.addWidget(calendar_btn)
        
        # Poznámka
        self.poznamka = QLineEdit()
        self.poznamka.setPlaceholderText("Nepovinná poznámka")
        
        # Ruční zadání ceny
        self.rucni_cena_check = QCheckBox("Ruční zadání ceny")
        self.rucni_cena_spin = QDoubleSpinBox()
        self.rucni_cena_spin.setRange(0, 1000000)
        self.rucni_cena_spin.setSuffix(" CZK")
        self.rucni_cena_spin.setEnabled(False)
        self.rucni_cena_check.stateChanged.connect(lambda: self.rucni_cena_spin.setEnabled(self.rucni_cena_check.isChecked()))
        
        rucni_layout = QHBoxLayout()
        rucni_layout.addWidget(self.rucni_cena_check)
        rucni_layout.addWidget(self.rucni_cena_spin)
        
        form_layout.addRow("Jméno *", self.jmeno)
        form_layout.addRow("E-mail", self.email)
        form_layout.addRow("Telefon", self.tel)
        form_layout.addRow("Služba *", self.sluzba)
        form_layout.addRow("Termín (včetně času) *", termin_layout)
        form_layout.addRow("Poznámka", self.poznamka)
        form_layout.addRow("Cena", rucni_layout)
        
        self.navyseni = QLineEdit("0")
        form_layout.addRow("Navýšení % (jen při výběru služby)", self.navyseni)
        
        # Combo box pro splatnosti
        self.splatnost = QComboBox()
        self.load_splatnosti_to_combo()
        form_layout.addRow("Splatnost *", self.splatnost)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Hlavní tlačítka
        main_buttons = QHBoxLayout()
        gen = QPushButton("Vytvořit zálohovou fakturu")
        gen.clicked.connect(self.generuj)
        main_buttons.addWidget(gen)
        
        # Tlačítko pro obnovení služeb
        refresh_btn = QPushButton("Obnovit seznam služeb")
        refresh_btn.clicked.connect(self.load_services_to_combo)
        main_buttons.addWidget(refresh_btn)
        
        main_buttons.addStretch()
        layout.addLayout(main_buttons)
        
        # Toolbar s upravenými funkcemi podle rolí
        toolbar = QToolBar()
        self.addToolBar(toolbar)
        
        # Základní tlačítka pro všechny role
        archive_action = QAction("📋 Zásobník", self)
        archive_action.triggered.connect(self.show_archive)
        toolbar.addAction(archive_action)
        
        settings_action = QAction("⚙ Nastavení", self)
        settings_action.triggered.connect(self.nastaveni)
        toolbar.addAction(settings_action)
        
        toolbar.addSeparator()
        
        # Podle role uživatele přidáme další tlačítka
        user_role = self.user_manager.users.get(self.user_manager.current_user, {}).get("role", "")
        
        # Admin má všechno
        if user_role == "admin":
            stats_action = QAction("📊 Statistiky", self)
            stats_action.triggered.connect(self.show_statistics)
            toolbar.addAction(stats_action)
            
            email_action = QAction("📧 E-mail", self)
            email_action.triggered.connect(self.show_email_settings)
            toolbar.addAction(email_action)
            
            tax_action = QAction("💰 Daňové reporty", self)
            tax_action.triggered.connect(self.show_tax_reports)
            toolbar.addAction(tax_action)
            
            users_action = QAction("👥 Uživatelé", self)
            users_action.triggered.connect(self.show_user_management)
            toolbar.addAction(users_action)
            
            backup_action = QAction("💾 Zálohy", self)
            backup_action.triggered.connect(self.show_backup)
            toolbar.addAction(backup_action)
            
            services_action = QAction("🛎️ Služby", self)
            services_action.triggered.connect(self.show_services_management)
            toolbar.addAction(services_action)
            
            recurring_action = QAction("🔄 Pravidelné", self)
            recurring_action.triggered.connect(self.show_recurring)
            toolbar.addAction(recurring_action)
            
            recovery_action = QAction("🔄 Obnova", self)
            recovery_action.triggered.connect(self.show_recovery)
            toolbar.addAction(recovery_action)
        
        # Recepční má omezené funkce
        elif user_role == "recepční":
            stats_action = QAction("📊 Statistiky", self)
            stats_action.triggered.connect(self.show_statistics)
            toolbar.addAction(stats_action)
            
            tax_action = QAction("💰 Daňové reporty", self)
            tax_action.triggered.connect(self.show_tax_reports)
            toolbar.addAction(tax_action)
            
            backup_action = QAction("💾 Zálohy", self)
            backup_action.triggered.connect(self.show_backup)
            toolbar.addAction(backup_action)
            
            services_action = QAction("🛎️ Služby", self)
            services_action.triggered.connect(self.show_services_management)
            toolbar.addAction(services_action)
            
            recurring_action = QAction("🔄 Pravidelné", self)
            recurring_action.triggered.connect(self.show_recurring)
            toolbar.addAction(recurring_action)
        
        # Účetní má jiné funkce
        elif user_role == "účetní":
            stats_action = QAction("📊 Statistiky", self)
            stats_action.triggered.connect(self.show_statistics)
            toolbar.addAction(stats_action)
            
            tax_action = QAction("💰 Daňové reporty", self)
            tax_action.triggered.connect(self.show_tax_reports)
            toolbar.addAction(tax_action)
        
        central_widget.setLayout(layout)
        
        # Status bar
        status_bar = QStatusBar()
        self.setStatusBar(status_bar)
        
        user_info = f"Přihlášen jako: {self.user_manager.current_user} ({user_role})"
        status_bar.showMessage(user_info)
    
    def show_calendar(self):
        """Zobrazí kalendář pro výběr data"""
        calendar_dialog = QDialog(self)
        calendar_dialog.setWindowTitle("Vyberte datum")
        calendar_dialog.setFixedSize(300, 300)
        
        layout = QVBoxLayout()
        
        calendar = QCalendarWidget()
        calendar.setGridVisible(True)
        
        select_button = QPushButton("Vybrat")
        select_button.clicked.connect(lambda: self.select_date(calendar.selectedDate(), calendar_dialog))
        
        layout.addWidget(calendar)
        layout.addWidget(select_button)
        
        calendar_dialog.setLayout(layout)
        calendar_dialog.exec()
    
    def select_date(self, date, dialog):
        """Zpracuje vybrané datum z kalendáře (bez času)"""
        self.termin.setText(date.toString("dd.MM.yyyy"))
        dialog.accept()
    
    def load_services_to_combo(self):
        self.sluzba.clear()
        all_services = get_all_active_services()
        
        if not all_services:
            QMessageBox.warning(self, "Upozornění", "Žádné služby nejsou dostupné. Přidejte služby v nastavení.")
            return
        
        # Seřadit služby podle kategorie
        services_by_category = {}
        for service in all_services:
            category = service["kategorie"]
            if category not in services_by_category:
                services_by_category[category] = []
            services_by_category[category].append(service)
        
        # Přidat do combo boxu
        for category, services in services_by_category.items():
            self.sluzba.addItem(f"--- {category} ---")
            self.sluzba.model().item(self.sluzba.count()-1).setEnabled(False)
            
            for service in services:
                self.sluzba.addItem(f"{service['nazev']} - {service['cena']} Kč")
        
        if self.sluzba.count() > 0:
            # Vybrat první skutečnou službu (po oddělovači kategorie)
            if self.sluzba.count() > 1:
                self.sluzba.setCurrentIndex(1)
    
    def load_splatnosti_to_combo(self):
        """Načte splatnosti do combo boxu"""
        self.splatnost.clear()
        active_splatnosti = get_all_active_splatnosti()
        
        if not active_splatnosti:
            # Pokud nejsou žádné aktivní splatnosti, načteme defaultní
            for splatnost in DEFAULT_SPLATNOSTI:
                self.splatnost.addItem(splatnost["nazev"])
        else:
            for splatnost in active_splatnosti:
                self.splatnost.addItem(splatnost["nazev"])
        
        if self.splatnost.count() > 0:
            self.splatnost.setCurrentIndex(0)
    
    def generuj(self):
        try:
            # Získat text služby
            service_text = self.sluzba.currentText().strip()
            if not service_text or "---" in service_text:
                QMessageBox.critical(self, "Chyba", "Vyberte platnou službu nebo zadejte vlastní název.")
                return
            
            # Zjistit, zda je vybraná služba ze seznamu (obsahuje " - " a končí "Kč")
            is_preset = " - " in service_text and service_text.endswith("Kč")
            cena = None
            
            # Ruční zadání ceny
            if self.rucni_cena_check.isChecked():
                cena = int(self.rucni_cena_spin.value())
                # Pokud je ruční cena, ignorujeme navýšení a cenu ze služby
                nav = 0
            else:
                # Cena ze služby
                if is_preset:
                    # Extrahovat cenu z textu
                    cena_text = service_text.split(" - ")[-1].replace(" Kč", "")
                    cena = int(cena_text)
                    # Aplikovat navýšení
                    nav = float(self.navyseni.text() or 0)
                    cena = int(cena * (1 + nav / 100))
                else:
                    # Vlastní služba bez ruční ceny = chyba
                    QMessageBox.critical(self, "Chyba", "Pro vlastní název služby musíte zaškrtnout 'Ruční zadání ceny' a vyplnit částku.")
                    return
            
            # Zpracování termínu (může obsahovat čas)
            termin_text = self.termin.text().strip()
            if not termin_text:
                QMessageBox.critical(self, "Chyba", "Zadejte termín konání.")
                return
            
            # Validace termínu (pokus o parsování)
            try:
                if ' ' in termin_text:
                    datetime.strptime(termin_text, "%d.%m.%Y %H:%M")
                else:
                    datetime.strptime(termin_text, "%d.%m.%Y")
            except ValueError:
                QMessageBox.critical(self, "Chyba", "Neplatný formát termínu. Použijte DD.MM.RRRR nebo DD.MM.RRRR HH:MM.")
                return

            if not self.jmeno.text():
                QMessageBox.critical(self, "Chyba", "Vyplň jméno.")
                return

            # Název služby pro fakturu (bez ceny)
            if is_preset:
                sluzba_nazev = service_text.split(" - ")[0]
            else:
                sluzba_nazev = service_text  # vlastní text

            data = {
                "jmeno": self.jmeno.text(),
                "email": self.email.text(),
                "telefon": self.tel.text(),
                "sluzba": sluzba_nazev,
                "termin": termin_text,
                "cena": cena,
                "splatnost": self.splatnost.currentText(),
                "poznamka": self.poznamka.text().strip()
            }
            
            # Výpočet konkrétního data splatnosti podle jednotky
            splatnosti = load_splatnosti()
            splatnost_obj = next((s for s in splatnosti if s["nazev"] == data["splatnost"]), None)
            if not splatnost_obj:
                QMessageBox.critical(self, "Chyba", "Zvolená splatnost není v seznamu.")
                return
            
            vytvoreni = datetime.now()
            hodiny = splatnost_obj.get("hodiny", 0)
            jednotka = splatnost_obj.get("jednotka", "hodiny")
            
            if jednotka == "hodiny":
                if hodiny == 0:
                    due_date = vytvoreni.date()
                else:
                    due_date = (vytvoreni + timedelta(hours=hodiny)).date()
            else:  # dny
                if hodiny == 0:
                    due_date = vytvoreni.date()
                else:
                    due_date = vytvoreni.date() + timedelta(days=hodiny)
            
            due_date_str = due_date.strftime("%d.%m.%Y")
            # Celý text do PDF
            splatnost_text = f"{due_date_str} - {data['splatnost']}"
            
            filepath, invoice_no = vytvor_pdf(data, splatnost_text, self.user_manager.current_user)
            
            # Odeslání e-mailu pokud je povoleno a je zadán e-mail
            if self.email.text() and settings.get("email_notifications", False):
                # Připravit data pro e-mail (včetně due_date)
                invoice_email_data = {
                    "cislo_faktury": invoice_no,
                    "jmeno": data["jmeno"],
                    "cena": data["cena"],
                    "due_date": due_date_str
                }
                success, message = send_invoice_email(self.email.text(), filepath, invoice_email_data)
                
                if success:
                    QMessageBox.information(self, "Hotovo", 
                                          f"Faktura {invoice_no} vytvořena a odeslána e-mailem (kopie odesílateli).")
                else:
                    QMessageBox.information(self, "Hotovo s upozorněním", 
                                          f"Faktura {invoice_no} vytvořena, ale e-mail se nepodařilo odeslat:\n{message}")
            else:
                QMessageBox.information(self, "Hotovo", 
                                      f"Faktura {invoice_no} vytvořena a přidána do archivu.")
            
            # Vyčistit formulář (jen některá pole)
            self.jmeno.clear()
            self.email.clear()
            self.tel.clear()
            self.poznamka.clear()
            self.navyseni.setText("0")
            self.rucni_cena_check.setChecked(False)
            self.rucni_cena_spin.setValue(0)
            
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se vytvořit fakturu: {str(e)}")
    
    def nastaveni(self):
        if not self.user_manager.has_permission("settings"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění upravovat nastavení.")
            return
        
        dlg = SettingsDialog(self.user_manager)
        if dlg.exec():
            self.load_services_to_combo()
            self.load_splatnosti_to_combo()
    
    def show_archive(self):
        if not self.user_manager.has_permission("view"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění prohlížet faktury.")
            return
        
        dlg = ArchiveDialog(self)
        dlg.exec()
    
    def show_services_management(self):
        if not self.user_manager.has_permission("services"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění spravovat služby.")
            return
        
        dlg = ServicesDialog()
        dlg.exec()
        # Po zavření dialogu obnovit seznam služeb
        self.load_services_to_combo()
        self.load_splatnosti_to_combo()
    
    # Nové metody pro nové funkce
    def show_statistics(self):
        if not self.user_manager.has_permission("reports"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění zobrazovat statistiky.")
            return
        
        dlg = StatisticsDialog(self)
        dlg.exec()
    
    def show_email_settings(self):
        if not self.user_manager.has_permission("settings"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění upravovat nastavení e-mailu.")
            return
        
        dlg = EmailSettingsDialog()
        dlg.exec()
    
    def show_tax_reports(self):
        if not self.user_manager.has_permission("reports"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění generovat daňové reporty.")
            return
        
        dlg = TaxReportDialog(self)
        dlg.exec()
    
    def show_user_management(self):
        if self.user_manager.current_user != "admin":
            QMessageBox.warning(self, "Oprávnění", "Pouze administrátor může spravovat uživatele.")
            return
        
        dlg = UserManagementDialog(self.user_manager)
        dlg.exec()
    
    def show_backup(self):
        if not self.user_manager.has_permission("backup"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění spravovat zálohy.")
            return
        
        dlg = BackupDialog(self.backup_manager)
        dlg.exec()
    
    def show_recurring(self):
        if not self.user_manager.has_permission("recurring"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění spravovat pravidelné faktury.")
            return
        
        dlg = RecurringInvoiceDialog(self)
        dlg.exec()
    
    def show_recovery(self):
        if not self.user_manager.has_permission("settings"):
            QMessageBox.warning(self, "Oprávnění", "Nemáte oprávnění spravovat obnovu.")
            return
        
        dlg = RecoveryDialog()
        dlg.exec()

# ================= PŮVODNÍ DIALOG NA HESLO =================
class PasswordDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Ověření")
        self.setFixedWidth(300)

        self.input = QLineEdit()
        self.input.setEchoMode(QLineEdit.Password)

        btn = QPushButton("Potvrdit")
        btn.clicked.connect(self.accept)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Zadej heslo:"))
        layout.addWidget(self.input)
        layout.addWidget(btn)

    def password(self):
        return self.input.text()

# ================= START =================
if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Nastavení tématu podle uložených nastavení
    theme = settings.get("theme", "system")
    setup_theme(app, theme)

    # Vytvořit archivní složku při startu
    ensure_archive_dir()

    # Načíst splatnosti při startu
    load_splatnosti()

    win = MainWindow()
    win.show()

    sys.exit(app.exec())